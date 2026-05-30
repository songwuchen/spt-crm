"""Excel export utilities using openpyxl."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.responses import StreamingResponse


def build_excel(title: str, headers: list[str], rows: list[list]) -> io.BytesIO:
    """Build an Excel workbook from headers + rows, return BytesIO."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-width (approximate)
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                cell_len = len(str(row_data[col_idx - 1] or ""))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _style_sheet(ws, headers: list[str], rows: list[list]) -> None:
    """Apply the standard header style + borders + auto-width to a worksheet."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                cell_len = len(str(row_data[col_idx - 1] or ""))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)


def build_excel_multi(sheets: list[tuple[str, list[str], list[list]]]) -> io.BytesIO:
    """Build a multi-sheet Excel workbook.

    `sheets` is a list of (title, headers, rows). The first sheet reuses the
    default worksheet; subsequent ones are created. Returns BytesIO.
    """
    wb = Workbook()
    if not sheets:
        # keep a single valid empty sheet
        wb.active.title = "Sheet1"
    for idx, (title, headers, rows) in enumerate(sheets):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = (title or f"Sheet{idx + 1}")[:31]
        _style_sheet(ws, headers, rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_template(title: str, headers: list[str], sample_rows: list[list] | None = None) -> io.BytesIO:
    """Build a template Excel with headers and optional sample data rows."""
    rows = sample_rows or []
    return build_excel(title, headers, rows)


def excel_response(buf: io.BytesIO, filename: str) -> StreamingResponse:
    """Wrap BytesIO in a StreamingResponse for download."""
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
