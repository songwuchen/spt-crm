from fastapi import APIRouter, Depends, Query, UploadFile, File
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
import io
import csv

from app.dependencies import get_db, get_tenant_id, require_permissions
from app.common.schemas import ok
from app.common.exceptions import BusinessException
from app.common.export import build_excel, excel_response
from app.domains.product import service
from app.dependencies import get_current_user
from app.domains.product.models import Product
from app.domains.product.schemas import (
    ProductCreate, ProductUpdate,
    ProductCategoryCreate, ProductCategoryUpdate,
)
from app.domains.quote.models import QuoteLine

router = APIRouter(prefix="/api/v1/products", tags=["产品目录"])


def _product_dict(p) -> dict:
    return {
        "id": p.id, "product_code": p.product_code, "name": p.name,
        "category_id": p.category_id, "item_type": p.item_type,
        "spec": p.spec, "unit": p.unit,
        "unit_price": float(p.unit_price) if p.unit_price is not None else None,
        "cost_price": float(p.cost_price) if p.cost_price is not None else None,
        "leadtime_days": p.leadtime_days,
        "is_active": p.is_active, "remark": p.remark,
        "extra_json": p.extra_json,
        "created_at": p.created_at.isoformat() if p.created_at else "",
        "updated_at": p.updated_at.isoformat() if p.updated_at else "",
    }


def _category_dict(c) -> dict:
    return {
        "id": c.id, "name": c.name, "parent_id": c.parent_id,
        "sort_order": c.sort_order, "description": c.description,
        "created_at": c.created_at.isoformat() if c.created_at else "",
    }


# ---- Category ----

@router.get("/categories")
async def list_categories(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("product:view")),
):
    items = await service.list_categories(db, tenant_id)
    return ok([_category_dict(c) for c in items])


@router.post("/categories")
async def create_category(
    body: ProductCategoryCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("product:edit")),
):
    c = await service.create_category(db, tenant_id, body)
    return ok(_category_dict(c))


@router.put("/categories/{category_id}")
async def update_category(
    category_id: str,
    body: ProductCategoryUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("product:edit")),
):
    c = await service.update_category(db, tenant_id, category_id, body)
    return ok(_category_dict(c))


@router.delete("/categories/{category_id}")
async def delete_category(
    category_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("product:edit")),
):
    await service.delete_category(db, tenant_id, category_id)
    return ok()


_CATEGORY_IMPORT_COLS = ["分类名称", "上级分类", "排序", "描述"]


@router.get("/categories/import/template")
async def category_import_template(_user=Depends(require_permissions("product:view"))):
    """下载产品分类导入模板。上级分类填已存在/同表中靠前的分类名称，留空为顶级分类。"""
    example = ["振动筛", "", 1, "筛分设备大类"]
    buf = build_excel("产品分类导入模板", _CATEGORY_IMPORT_COLS, [example])
    return excel_response(buf, "product_categories_template.xlsx")


@router.post("/categories/import")
async def import_categories(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("product:edit")),
):
    """批量导入产品分类。列顺序：分类名称, 上级分类, 排序, 描述。
    上级分类按名称匹配（已存在或本表中靠前的行）；重名分类跳过。"""
    content = await file.read()
    fname = (file.filename or "").lower()
    try:
        if fname.endswith(".csv"):
            text = content.decode("utf-8-sig", errors="replace")
            all_rows = [tuple(r) for r in csv.reader(io.StringIO(text))]
        else:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True)) if ws is not None else []
            wb.close()
    except Exception:
        raise BusinessException(message="无法解析文件，请使用导入模板（.xlsx 或 .csv）")

    if len(all_rows) < 2:
        return ok({"created": 0, "skipped": 0, "errors": []})
    data_rows = all_rows[1:]

    name_to_id = {c.name: c.id for c in await service.list_categories(db, tenant_id)}
    created, skipped, errors = 0, 0, []
    for idx, row in enumerate(data_rows, 2):
        if not row or not any(row):
            continue

        def _cell(i: int):
            return str(row[i]).strip() if len(row) > i and row[i] is not None and str(row[i]).strip() != "" else None

        name = _cell(0)
        if not name:
            skipped += 1
            continue
        if name in name_to_id:
            skipped += 1
            continue
        parent_name = _cell(1)
        parent_id = name_to_id.get(parent_name) if parent_name else None
        sort_order = 0
        try:
            sort_order = int(float(row[2])) if len(row) > 2 and row[2] is not None and str(row[2]).strip() != "" else 0
        except (ValueError, TypeError):
            pass
        try:
            c = await service.create_category(db, tenant_id, ProductCategoryCreate(
                name=name, parent_id=parent_id, sort_order=sort_order, description=_cell(3)))
            name_to_id[name] = c.id
            created += 1
        except Exception as e:
            await db.rollback()
            errors.append(f"第{idx}行: {str(e)[:80]}")
    return ok({"created": created, "skipped": skipped, "errors": errors})


# ---- Product ----

@router.get("")
async def list_products(
    pageNo: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    keyword: str = Query(None),
    category_id: str = Query(None),
    item_type: str = Query(None),
    is_active: bool = Query(None),
    filter: str = Query(None, description="高级筛选 FilterDsl(JSON)"),
    sort_by: str = Query(None),
    sort_order: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("product:view")),
):
    items, total = await service.list_products(
        db, tenant_id, pageNo, pageSize, keyword, category_id, item_type, is_active,
        current_user=_user, adv_filter=filter, sort_by=sort_by, sort_order=sort_order)
    # Batch usage count by matching item_code to product_code
    codes = [p.product_code for p in items if p.product_code]
    usage_map: dict[str, int] = {}
    if codes:
        rows = (await db.execute(
            select(QuoteLine.item_code, func.count(QuoteLine.id).label("cnt"))
            .where(QuoteLine.tenant_id == tenant_id, QuoteLine.item_code.in_(codes))
            .group_by(QuoteLine.item_code)
        )).all()
        usage_map = {r.item_code: r.cnt for r in rows}
    result_items = []
    for p in items:
        d = _product_dict(p)
        d["usage_count"] = usage_map.get(p.product_code, 0)
        result_items.append(d)
    return ok({"items": result_items, "total": total, "pageNo": pageNo, "pageSize": pageSize})


# 注意：静态路径必须注册在 /{product_id} 之前，否则会被动态路由吞掉（FastAPI 按注册顺序匹配）
@router.get("/check-unique")
async def check_unique_product(
    product_code: str = Query(..., min_length=1),
    exclude_id: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Check if a product code is unique within the tenant."""
    q = select(Product.id).where(
        Product.tenant_id == tenant_id,
        Product.product_code == product_code,
    )
    if exclude_id:
        q = q.where(Product.id != exclude_id)
    exists = (await db.execute(q)).scalar() is not None
    return ok({"unique": not exists})


@router.get("/{product_id}")
async def get_product(
    product_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("product:view")),
):
    p = await service.get_product(db, tenant_id, product_id)
    return ok(_product_dict(p))


@router.post("")
async def create_product(
    body: ProductCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("product:create")),
):
    p = await service.create_product(db, tenant_id, body, current_user)
    return ok(_product_dict(p))


@router.put("/{product_id}")
async def update_product(
    product_id: str,
    body: ProductUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("product:edit")),
):
    p = await service.update_product(db, tenant_id, product_id, body, current_user)
    return ok(_product_dict(p))


@router.delete("/{product_id}")
async def delete_product(
    product_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("product:delete")),
):
    await service.delete_product(db, tenant_id, product_id, current_user)
    return ok()


_PRODUCT_IMPORT_COLS = ["产品编码", "名称", "类型(标准品/非标品/服务/备件)", "规格", "单位", "单价", "成本价", "交期(天)"]

# 类型：兼容中文标签与英文代码，未知值按空处理（不阻断导入）
_ITEM_TYPE_MAP = {
    "standard": "standard", "标准品": "standard", "标准件": "standard", "标准": "standard",
    "nonstandard": "nonstandard", "非标品": "nonstandard", "非标件": "nonstandard", "非标": "nonstandard",
    "service": "service", "服务": "service",
    "spare": "spare", "备件": "spare", "配件": "spare",
}


def _norm_item_type(v):
    if not v:
        return None
    return _ITEM_TYPE_MAP.get(str(v).strip())


@router.get("/import/template")
async def product_import_template(_user=Depends(require_permissions("product:view"))):
    """下载产品导入 Excel 模板（表头 + 示例行，列顺序与导入一致）。"""
    example = ["P-001", "示例产品", "标准品", "Φ100×200", "台", 1200, 800, 15]
    buf = build_excel("产品导入模板", _PRODUCT_IMPORT_COLS, [example])
    return excel_response(buf, "products_template.xlsx")


@router.post("/import/excel")
async def import_products_excel(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("product:create")),
):
    """从 Excel/CSV 导入产品。列顺序：产品编码, 名称, 类型, 规格, 单位, 单价, 成本价, 交期(天)。"""
    content = await file.read()
    fname = (file.filename or "").lower()
    try:
        if fname.endswith(".csv"):
            text = content.decode("utf-8-sig", errors="replace")
            all_rows = [tuple(r) for r in csv.reader(io.StringIO(text))]
        else:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True)) if ws is not None else []
            wb.close()
    except Exception:
        raise BusinessException(message="无法解析文件，请使用导入模板（.xlsx 或 .csv；旧版 .xls 请先另存为 .xlsx）")

    if len(all_rows) < 2:
        return ok({"created": 0, "skipped": 0, "errors": []})

    data_rows = all_rows[1:]

    # Detect existing by product_code
    codes = []
    for row in data_rows:
        if row and len(row) > 0 and row[0]:
            codes.append(str(row[0]).strip())
    existing_codes: set[str] = set()
    if codes:
        result = await db.execute(
            select(Product.product_code).where(
                Product.tenant_id == tenant_id,
                Product.product_code.in_(codes),
            )
        )
        existing_codes = {r[0] for r in result.all()}

    from app.domains.product.schemas import ProductCreate
    created = 0
    skipped = 0
    errors = []
    for idx, row in enumerate(data_rows, 2):
        if not row or not row[0]:
            continue
        code = str(row[0]).strip()
        if code in existing_codes:
            skipped += 1
            continue
        try:
            def _cell(i: int) -> str | None:
                return str(row[i]).strip() if len(row) > i and row[i] is not None else None

            price = None
            cost = None
            lt = None
            try:
                price = float(row[5]) if len(row) > 5 and row[5] is not None else None
            except (ValueError, TypeError):
                pass
            try:
                cost = float(row[6]) if len(row) > 6 and row[6] is not None else None
            except (ValueError, TypeError):
                pass
            try:
                lt = int(float(row[7])) if len(row) > 7 and row[7] is not None else None
            except (ValueError, TypeError):
                pass

            data = ProductCreate(
                product_code=code,
                name=_cell(1) or code,
                item_type=_norm_item_type(_cell(2)),
                spec=_cell(3),
                unit=_cell(4),
                unit_price=price,
                cost_price=cost,
                leadtime_days=lt,
            )
            await service.create_product(db, tenant_id, data, current_user)
            created += 1
            existing_codes.add(code)
        except Exception as e:
            await db.rollback()  # 单行失败不污染会话，后续行可继续
            errors.append(f"第{idx}行: {str(e)[:80]}")
    return ok({"created": created, "skipped": skipped, "errors": errors})
