from datetime import date

from fastapi import APIRouter, Depends, Query, Header as FastAPIHeader, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel
from typing import Optional
from openpyxl import load_workbook
import io

from app.dependencies import get_db, get_tenant_id, require_permissions
from app.common.schemas import ok
from app.common.export import build_excel, build_template, excel_response
from app.domains.lead import service
from app.domains.lowcode.field_permission import strip_entity_dicts

router = APIRouter(prefix="/api/v1/leads", tags=["线索管理"])

# 导入列（顺序即模板列顺序）；与线索列表字段保持一致 (issue #95)：
# 增加「客户类型」「类别」两列，覆盖列表中可录入的业务字段。
LEAD_IMPORT_HEADERS = ["标题", "公司名称", "联系人", "联系电话", "邮箱", "来源",
                       "客户类型", "行业", "类别", "地区", "业务日期", "负责人"]

# 类别列既接受编码(self_reported/distributed)也接受中文标签
_CATEGORY_BY_LABEL = {
    "自报": "self_reported", "自拓": "self_reported", "self_reported": "self_reported",
    "分发": "distributed", "分配": "distributed", "distributed": "distributed",
}


def _norm_category(v):
    """把类别单元格归一化为编码；无法识别则返回 None（避免整行导入失败）。"""
    if v is None or str(v).strip() == "":
        return None
    return _CATEGORY_BY_LABEL.get(str(v).strip())


def _product_dict(p) -> dict:
    return {
        "id": p.id, "product_name": p.product_name, "product_spec": p.product_spec,
        "quantity": float(p.quantity) if p.quantity is not None else None,
        "remark": p.remark,
    }


def _lead_dict(l, products=None, dept_names=None) -> dict:
    """线索出参的唯一序列化入口（没有 Out schema，改字段请只改这里）。

    dept_names: department_id -> name 的批量映射，由调用方预取以避免逐条查询；
    未传时 department_name 为 None，前端退化为不显示部门名。
    """
    return {
        "id": l.id, "lead_code": l.lead_code, "title": l.title, "company_name": l.company_name,
        "contact_name": l.contact_name, "contact_phone": l.contact_phone,
        "contact_email": l.contact_email, "contact_raw_json": l.contact_raw_json,
        "source": l.source, "source_detail_json": l.source_detail_json,
        "demand_summary": l.demand_summary,
        "industry": l.industry,
        "customer_type": l.customer_type,
        "category": l.category,
        "country_type": l.country_type,
        "country_name": l.country_name,
        "region": l.region,
        "province": l.province,
        "city": l.city,
        "district": l.district,
        "region_code": l.region_code,
        "department_id": l.department_id,
        "department_name": (dept_names or {}).get(l.department_id),
        "budget_range": l.budget_range,
        "owner_id": l.owner_id, "owner_name": l.owner_name,
        "created_by_id": l.created_by_id, "created_by_name": l.created_by_name,
        "biz_date": str(l.biz_date) if l.biz_date else None,
        "status": l.status, "score": l.score,
        "review_status": getattr(l, "review_status", "approved"),
        "review_flow_id": getattr(l, "review_flow_id", None),
        "reject_reason": getattr(l, "reject_reason", None),
        "converted_customer_id": l.converted_customer_id,
        "remark": l.remark,
        # 扩展字段值必须回传：strip_entity_dicts 依赖它做字段级权限裁剪，前端编辑表单也据此
        # 回填，缺失会导致保存时以空对象覆盖掉已存值。
        "custom_fields_json": l.custom_fields_json or {},
        "products": [_product_dict(p) for p in products] if products is not None else [],
        "created_at": l.created_at.isoformat() if l.created_at else "",
        "updated_at": l.updated_at.isoformat() if l.updated_at else "",
    }


@router.get("")
async def list_leads(
    pageNo: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    keyword: str = Query(None),
    status: str = Query(None),
    owner_id: str = Query(None),
    customer_type: str = Query(None),
    category: str = Query(None),
    country_type: str = Query(None),
    province: str = Query(None),
    department_id: str = Query(None),
    industry: str = Query(None),
    company_name: str = Query(None),
    start_date: date = Query(None),
    end_date: date = Query(None),
    date_field: str = Query(None, description="日期区间筛选字段：created_at(默认) / biz_date"),
    filter: str = Query(None, description="高级筛选 FilterDsl(JSON)"),
    sort_by: str = Query(None),
    sort_order: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("lead:view")),
):
    # 数据范围「本人」= 负责人/创建人/共享给本人；「部门」= 部门子树成员；「全部」= 不限。
    items, total = await service.list_leads(
        db, tenant_id, pageNo, pageSize, keyword, status, owner_id,
        customer_type=customer_type, category=category, country_type=country_type,
        province=province, department_id=department_id, industry=industry,
        company_name=company_name, start_date=start_date, end_date=end_date,
        date_field=date_field, current_user=_user,
        adv_filter=filter, sort_by=sort_by, sort_order=sort_order,
    )
    dept_names = await _lead_department_names(db, tenant_id, items)  # 一次查询，避免逐条取部门名
    dicts = [_lead_dict(l, dept_names=dept_names) for l in items]
    await strip_entity_dicts(db, tenant_id, "lead", dicts, _user.get("roles"))  # 字段级权限：读取剔除隐藏扩展字段
    return ok({"items": dicts, "total": total, "pageNo": pageNo, "pageSize": pageSize})


async def _lead_department_names(db: AsyncSession, tenant_id: str, items) -> dict:
    """批量取 department_id -> name，供导出回填部门名。"""
    from sqlalchemy import select
    from app.domains.organization.models import Department
    ids = {l.department_id for l in items if l.department_id}
    if not ids:
        return {}
    rows = (await db.execute(select(Department.id, Department.name).where(
        Department.tenant_id == tenant_id, Department.id.in_(ids)))).all()
    return {did: name for did, name in rows}


async def _lead_products_text(db: AsyncSession, tenant_id: str, items) -> dict:
    """批量取各线索的产品明细，拼成一段可读文本供导出（一条线索可有多个产品）。"""
    from sqlalchemy import select
    from app.domains.lead.models import LeadProduct
    ids = {l.id for l in items}
    if not ids:
        return {}
    rows = (await db.execute(select(LeadProduct).where(
        LeadProduct.tenant_id == tenant_id, LeadProduct.lead_id.in_(ids))
        .order_by(LeadProduct.lead_id, LeadProduct.sort_order))).scalars().all()
    grouped: dict = {}
    for p in rows:
        parts = [p.product_name or ""]
        if p.product_spec:
            parts.append(f"({p.product_spec})")
        if p.quantity is not None:
            parts.append(f"x{float(p.quantity):g}")
        if p.remark:
            parts.append(f"[{p.remark}]")
        grouped.setdefault(p.lead_id, []).append("".join(parts))
    return {lid: "; ".join(v) for lid, v in grouped.items()}


@router.get("/export/excel")
async def export_leads_excel(
    keyword: str = Query(None),
    status: str = Query(None),
    owner_id: str = Query(None),
    company_name: str = Query(None),
    start_date: date = Query(None),
    end_date: date = Query(None),
    date_field: str = Query(None, description="日期区间筛选字段：created_at(默认) / biz_date"),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("lead:view")),
):
    from app.config import settings
    items, _ = await service.list_leads(
        db, tenant_id, 1, settings.MAX_EXPORT_ROWS, keyword, status, owner_id,
        company_name=company_name, start_date=start_date, end_date=end_date,
        date_field=date_field, current_user=_user,
    )
    # 导出除列表字段外，补齐详情中「部门/联系人/产品/补充信息」各模块字段 (issue #95)
    dept_names = await _lead_department_names(db, tenant_id, items)
    product_texts = await _lead_products_text(db, tenant_id, items)
    headers = [
        # 列表 & 基本信息
        "线索编码", "标题", "公司名称", "部门", "来源", "类别", "客户类型", "行业",
        "业务日期", "负责人", "状态", "评分",
        # 联系人信息
        "联系人", "联系电话", "邮箱",
        # 地区
        "国别", "国家", "省", "市", "区县", "地区",
        # 产品信息 & 补充信息
        "产品信息", "预算范围", "需求摘要", "备注",
        "创建时间",
    ]
    category_label = {"self_reported": "自报", "distributed": "分发"}
    country_label = {"domestic": "国内", "overseas": "国外"}
    # 导出与列表/详情同口径：隐藏字段导空、脱敏字段导 "***"。
    # 否则「页面看不到但能导出来」就是一条绕过字段权限的后门。
    from app.domains.lowcode.field_permission import entity_field_restrictions, export_cell
    rst = await entity_field_restrictions(db, tenant_id, "lead", _user.get("roles"))
    c = lambda fid, v: export_cell(rst, fid, v)  # noqa: E731
    rows = []
    for l in items:
        rows.append([
            l.lead_code or "", c("title", l.title or ""), c("company_name", l.company_name or ""),
            c("department_id", dept_names.get(l.department_id, "") if l.department_id else ""),
            c("source", l.source or ""),
            c("category", category_label.get(l.category or "", l.category or "")),
            c("customer_type", l.customer_type or ""),
            c("industry", l.industry or ""),
            c("biz_date", str(l.biz_date) if l.biz_date else ""),
            c("owner_id", l.owner_name or ""), l.status or "", l.score or "",
            c("contact_name", l.contact_name or ""), c("contact_phone", l.contact_phone or ""),
            c("contact_email", l.contact_email or ""),
            c("country_type", country_label.get(l.country_type or "", l.country_type or "")),
            c("country_name", l.country_name or ""),
            l.province or "", l.city or "", l.district or "", c("region", l.region or ""),
            product_texts.get(l.id, ""),
            c("budget_range", l.budget_range or ""),
            c("demand_summary", l.demand_summary or ""),
            c("remark", l.remark or ""),
            l.created_at.strftime("%Y-%m-%d %H:%M") if l.created_at else "",
        ])
    buf = build_excel("线索列表", headers, rows)
    return excel_response(buf, "leads.xlsx")


@router.get("/import/template")
async def download_lead_import_template(
    _user=Depends(require_permissions("lead:create")),
):
    """下载线索导入模板（列与线索列表字段保持一致，issue #95）。"""
    sample = [["某某设备采购线索", "示例科技有限公司", "张三", "13800000000",
               "zhangsan@example.com", "展会", "企业客户", "机械制造", "自报",
               "上海市浦东新区", "2026-07-01", "李四"]]
    buf = build_template("线索导入模板", LEAD_IMPORT_HEADERS, sample)
    return excel_response(buf, "lead_import_template.xlsx")


def _parse_date_cell(v):
    """把 Excel 日期单元格(datetime/date/字符串)归一化为 date，无法解析则返回 None。"""
    if v is None or v == "":
        return None
    if hasattr(v, "date") and not isinstance(v, str):  # datetime
        return v.date()
    if isinstance(v, date):
        return v
    return str(v).strip()  # 交给 Pydantic 解析 "YYYY-MM-DD"


async def _resolve_owner_id(db: AsyncSession, tenant_id: str, name):
    """按姓名(real_name 或 username)在租户内匹配用户，返回 user_id；匹配不到返回 None。"""
    if not name or not str(name).strip():
        return None
    from sqlalchemy import select
    from app.domains.auth.models import User as AuthUser
    nm = str(name).strip()
    u = (await db.execute(select(AuthUser).where(
        AuthUser.tenant_id == tenant_id,
        (AuthUser.real_name == nm) | (AuthUser.username == nm)))).scalars().first()
    return u.id if u else None


@router.post("/import/excel")
async def import_leads_excel(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("lead:create")),
):
    """Import leads from Excel. Columns: 标题, 公司名称, 联系人, 联系电话, 邮箱, 来源, 客户类型, 行业, 类别, 地区, 业务日期, 负责人"""
    from app.domains.lead.schemas import LeadCreate
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    errors = []

    def cell(row, i):
        return row[i] if len(row) > i and row[i] not in (None, "") else None

    for idx, row in enumerate(rows, 2):
        if not row or not row[0]:
            continue
        try:
            owner_id = await _resolve_owner_id(db, tenant_id, cell(row, 11))
            data = LeadCreate(
                title=str(row[0]).strip(),
                # company_name is required; fall back to the title when the column is blank
                company_name=str(row[1]).strip() if cell(row, 1) else str(row[0]).strip(),
                contact_name=str(row[2]).strip() if cell(row, 2) else None,
                contact_phone=str(row[3]).strip() if cell(row, 3) else None,
                contact_email=str(row[4]).strip() if cell(row, 4) else None,
                source=str(row[5]).strip() if cell(row, 5) else "import",
                customer_type=str(row[6]).strip() if cell(row, 6) else None,
                industry=str(row[7]).strip() if cell(row, 7) else None,
                category=_norm_category(cell(row, 8)),
                region=str(row[9]).strip() if cell(row, 9) else None,
                biz_date=_parse_date_cell(cell(row, 10)),
                owner_id=owner_id,
            )
            await service.create_lead(db, tenant_id, data, current_user)
            created += 1
        except Exception as e:
            errors.append(f"第{idx}行: {str(e)[:80]}")
    wb.close()
    return ok({"created": created, "errors": errors})


@router.post("")
async def create_lead(
    body: service.LeadCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("lead:create")),
):
    l = await service.create_lead(db, tenant_id, body, current_user)
    products = await service.list_lead_products(db, tenant_id, l.id)
    return ok(_lead_dict(l, products, await _lead_department_names(db, tenant_id, [l])))


@router.get("/{lead_id}")
async def get_lead(
    lead_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("lead:view")),
):
    l = await service.get_lead(db, tenant_id, lead_id)
    products = await service.list_lead_products(db, tenant_id, l.id)
    d = _lead_dict(l, products, await _lead_department_names(db, tenant_id, [l]))
    await strip_entity_dicts(db, tenant_id, "lead", [d], _user.get("roles"))  # 字段级权限：读取剔除隐藏扩展字段
    return ok(d)


@router.put("/{lead_id}")
async def update_lead(
    lead_id: str,
    body: service.LeadUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("lead:edit")),
):
    l = await service.update_lead(db, tenant_id, lead_id, body, current_user)
    products = await service.list_lead_products(db, tenant_id, l.id)
    return ok(_lead_dict(l, products, await _lead_department_names(db, tenant_id, [l])))


class QualifyBody(BaseModel):
    create_opportunity: bool = False


@router.post("/{lead_id}/qualify")
async def qualify_lead(
    lead_id: str,
    body: Optional[QualifyBody] = None,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("lead:qualify")),
):
    create_opp = body.create_opportunity if body else False
    result = await service.qualify_lead(db, tenant_id, lead_id, current_user, create_opportunity=create_opp)
    return ok(result)


@router.post("/{lead_id}/submit_review")
async def submit_lead_review(
    lead_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("lead:edit")),
):
    """被驳回的线索修改后重新提交内勤审核。"""
    l = await service.resubmit_lead_review(db, tenant_id, lead_id, current_user)
    products = await service.list_lead_products(db, tenant_id, l.id)
    return ok(_lead_dict(l, products))


@router.post("/{lead_id}/discard")
async def discard_lead(
    lead_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("lead:discard")),
):
    l = await service.discard_lead(db, tenant_id, lead_id, current_user)
    return ok(_lead_dict(l))


@router.delete("/{lead_id}")
async def delete_lead(
    lead_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("lead:delete")),
):
    await service.delete_lead(db, tenant_id, lead_id, current_user)
    return ok()


# --- Public Lead Capture Webhook ---

class PublicLeadBody(BaseModel):
    company_name: Optional[str] = None
    contact_name: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[str] = None
    demand_summary: Optional[str] = None
    industry: Optional[str] = None
    region: Optional[str] = None
    source: Optional[str] = "inbound"


public_router = APIRouter(prefix="/api/public/v1", tags=["公开接口"])


@public_router.post("/leads")
async def public_lead_capture(
    body: PublicLeadBody,
    x_tenant_id: str = FastAPIHeader(alias="X-Tenant-Id", default="00000000-0000-0000-0000-000000000001"),
    db: AsyncSession = Depends(get_db),
):
    """Public lead capture webhook. No auth required. Requires X-Tenant-Id header."""
    from app.domains.lead.schemas import LeadCreate
    title = body.company_name or body.contact_name or "网页表单线索"
    data = LeadCreate(
        title=title,
        # company_name is required; fall back to the derived title for anonymous submissions
        company_name=body.company_name or title,
        contact_name=body.contact_name,
        contact_phone=body.contact_phone,
        contact_email=body.contact_email,
        demand_summary=body.demand_summary,
        industry=body.industry,
        region=body.region,
        source=body.source or "inbound",
    )
    system_user = {"sub": "system", "real_name": "系统"}
    # 公开表单线索无归属提交人，直接免审进入线索池（由内勤后续分配/跟进）
    lead = await service.create_lead(db, x_tenant_id, data, system_user, auto_review=False)
    return ok({"id": lead.id, "title": lead.title})


# ---- Batch Operations ----

class BatchAssignBody(BaseModel):
    ids: list[str]
    owner_id: str
    owner_name: Optional[str] = None


class BatchStatusBody(BaseModel):
    ids: list[str]
    status: str  # new / following / qualified / discarded


@router.post("/batch_assign")
async def batch_assign(
    body: BatchAssignBody,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("lead:edit")),
):
    """Batch assign leads to a new owner."""
    from sqlalchemy import select, update
    from app.domains.lead.models import Lead
    sample = (await db.execute(
        select(Lead).where(Lead.tenant_id == tenant_id, Lead.id.in_(body.ids), Lead.is_deleted == False).limit(1)
    )).scalar()
    result = await db.execute(
        update(Lead).where(
            Lead.tenant_id == tenant_id,
            Lead.id.in_(body.ids),
            Lead.is_deleted == False,
        ).values(owner_id=body.owner_id, owner_name=body.owner_name)
    )
    await db.commit()
    # 批量改派给他人 → 给新负责人发一条汇总通知
    if body.owner_id and body.owner_id != current_user["sub"] and result.rowcount and sample:
        try:
            from app.common.auto_notify import notify_lead_assigned
            await notify_lead_assigned(
                db, tenant_id, sample.title or sample.company_name or "线索",
                body.owner_id, current_user.get("real_name") or current_user.get("username"),
                sample.id, count=result.rowcount,
            )
        except Exception:
            pass
    return ok({"updated": result.rowcount})


@router.post("/batch_status")
async def batch_status(
    body: BatchStatusBody,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("lead:edit")),
):
    """Batch update lead status."""
    from sqlalchemy import update
    from app.domains.lead.models import Lead
    if body.status not in ("new", "following", "qualified", "discarded"):
        from app.common.exceptions import BusinessException
        raise BusinessException(message="无效状态")
    stmt = update(Lead).where(
        Lead.tenant_id == tenant_id,
        Lead.id.in_(body.ids),
        Lead.is_deleted == False,
    )
    # 批量转化时跳过尚未通过审核的线索，避免绕过审核门禁
    if body.status == "qualified":
        stmt = stmt.where(Lead.review_status == "approved")
    result = await db.execute(stmt.values(status=body.status))
    await db.commit()
    return ok({"updated": result.rowcount})
