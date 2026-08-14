"""线索 Excel 导入：模板列定义、表头映射、单元格归一化。

对齐申报表单必填/常用字段；支持旧版 12 列模板（标题/类别/来源…）按表头兼容。
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

# 模板列：对齐新建页「申报信息（创建时填写）」+「其他（可选）」
# 不含：中标/原因（新建隐藏）、业务反馈、评估区（新/老客户、备注2 等审批填写）
LEAD_IMPORT_HEADERS: list[str] = [
    "项目名称",
    "来源",
    "公司名称",
    "客户类型",
    "国别",
    "国家",
    "省",
    "市",
    "区县",
    "详细地址",
    "是否内部冲突",
    "备注：请示部门经理的结果",
    "行业",
    "委托状态",
    "委托开具日期",
    "委托期限",
    "部门",
    "申报人",
    "申报时间",
    "负责人",
    "项目动态",
    "备注1（线索内容）",
    "联系人",
    "联系电话",
    "联系邮箱",
    "线索来源",
    "业务日期",
    "备注",
]

# 表头别名 → 规范字段名
_HEADER_TO_FIELD: dict[str, str] = {
    "项目名称": "title",
    "标题": "title",
    # 「来源」单独处理：新模板=报备来源(category)；旧模板与「类别」并存时=线索来源
    "类别": "category",
    "报备来源": "category",
    "公司名称": "company_name",
    "客户类型": "customer_type",
    "国别": "country_type",
    "国家": "country_name",
    "省": "province",
    "市": "city",
    "区县": "district",
    "详细地址": "region",
    "地区": "region",
    "是否内部冲突": "has_internal_conflict",
    "备注：请示部门经理的结果": "conflict_note",
    "冲突备注": "conflict_note",
    "行业": "industry",
    "中标情况": "bid_result",
    "原因": "bid_fail_reason",
    "委托状态": "entrust_status",
    "委托开具日期": "entrust_issued_at",
    "委托期限": "entrust_term",
    "部门": "department_name",
    "申报人": "reporter_name",
    "报备人": "reporter_name",
    "申报时间": "reported_at",
    "报备时间": "reported_at",
    "负责人": "owner_name",
    "项目动态": "project_activity",
    "备注1（线索内容）": "demand_summary",
    "备注1": "demand_summary",
    "需求摘要": "demand_summary",
    "线索内容": "demand_summary",
    "联系人": "contact_name",
    "联系电话": "contact_phone",
    "联系邮箱": "contact_email",
    "邮箱": "contact_email",
    "线索来源": "source",
    "业务日期": "biz_date",
    "备注": "remark",
}

_CATEGORY_BY_LABEL = {
    "自报": "self_reported",
    "自拓": "self_reported",
    "self_reported": "self_reported",
    "分发": "distributed",
    "分配": "distributed",
    "distributed": "distributed",
}

_COUNTRY_BY_LABEL = {
    "国内": "domestic",
    "国外": "overseas",
    "海外": "overseas",
    "domestic": "domestic",
    "overseas": "overseas",
}

_YES_NO = {"是": "是", "否": "否", "yes": "是", "no": "否", "y": "是", "n": "否", "true": "是", "false": "否"}

# 客户类型 / 行业：中文标签 → 编码（与 seed_lead_dicts / leadForm 兜底一致）
_CUSTOMER_TYPE_BY_LABEL = {
    "终端客户-央企/国企": "terminal_soe",
    "终端客户-大型民企（注册资本10亿以上）": "terminal_large_private",
    "终端客户-一般民企": "terminal_private",
    "设计院": "design_institute",
    "总包商": "general_contractor",
    "配套商、贸易商": "supporting_trader",
    "配套商/贸易商": "supporting_trader",
    "其他": "other",
    "企业客户": "other",  # 旧样例兼容
}
_CUSTOMER_TYPE_BY_CODE = {v: v for v in _CUSTOMER_TYPE_BY_LABEL.values()}

_INDUSTRY_BY_LABEL = {
    "筛分分选-冶金": "screening_metallurgy",
    "筛分分选-矿山": "screening_mining",
    "筛分分选-砂石": "screening_aggregate",
    "筛分分选-焦化": "screening_coking",
    "筛分分选-煤炭": "screening_coal",
    "筛分分选-电力": "screening_power",
    "筛分分选-化工": "screening_chemical",
    "筛分分选-医药": "screening_pharma",
    "筛分分选-食品": "screening_food",
    "筛分分选-备件": "screening_spare_parts",
    "循环经济": "circular_economy",
    "废钢利用": "scrap_steel",
    "智能化大宗物料管理": "bulk_material_intelligent",
    "机械制造": "screening_metallurgy",  # 旧样例兼容
}
_INDUSTRY_BY_CODE = {v: v for v in _INDUSTRY_BY_LABEL.values()}

_SOURCE_BY_LABEL = {
    "展会": "expo",
    "转介绍": "referral",
    "广告": "ad",
    "官网/入站": "inbound",
    "官网": "inbound",
    "入站": "inbound",
    "合作伙伴": "partner",
    "电话": "call",
    "expo": "expo",
    "referral": "referral",
    "ad": "ad",
    "inbound": "inbound",
    "partner": "partner",
    "call": "call",
    "import": "import",
}

_PROJECT_ACTIVITY = {"技术交流", "出方案", "报价", "投标", "拟建"}
_ENTRUST = {"已开", "未开"}
_BID_RESULT = {"中标", "结果未出", "项目取消", "项目延期", "落标", "流标", "未参与"}
_REPORT_STATUS = {"进行中", "暂缓", "暂停", "取消", "落标", "中标", "已签合同"}


def lead_import_sample_row() -> list[Any]:
    """与 LEAD_IMPORT_HEADERS 顺序对应的样例行。"""
    return [
        "某某设备采购线索",  # 项目名称
        "自报",  # 来源
        "示例科技有限公司",  # 公司名称
        "设计院",  # 客户类型
        "国内",  # 国别
        "",  # 国家
        "浙江省",  # 省
        "嘉兴市",  # 市
        "海盐县",  # 区县
        "某某路 1 号",  # 详细地址
        "否",  # 是否内部冲突
        "",  # 冲突备注
        "筛分分选-冶金",  # 行业
        "未开",  # 委托状态
        "",  # 委托开具日期
        "",  # 委托期限
        "信息情报部",  # 部门（按名称匹配，请改成实际部门名）
        "",  # 申报人（空=导入人）
        "",  # 申报时间（空=当前）
        "",  # 负责人（空=导入人）
        "拟建",  # 项目动态
        "需湿法钢渣处理方案",  # 备注1
        "张三",  # 联系人
        "13800000000",  # 联系电话
        "zhangsan@example.com",  # 联系邮箱
        "展会",  # 线索来源
        "2026-07-01",  # 业务日期
        "",  # 备注
    ]


def lead_import_guide_rows() -> list[list[str]]:
    return [
        ["范围", "仅新建可填字段", "不含中标/原因、业务反馈、评估区（新/老客户、备注2 等由情报审批填写）"],
        ["项目名称", "必填", "与表单「项目名称」一致"],
        ["来源", "必填", "自报 / 分发"],
        ["公司名称", "必填", ""],
        ["客户类型", "必填", "设计院 / 总包商 / 终端客户-一般民企 等（可填中文或编码）；不是审批里的「新/老客户」"],
        ["国别", "必填", "国内 / 国外"],
        ["国家", "国别=国外时填", "如 越南"],
        ["省/市/区县", "国内必填", "如 浙江省 / 嘉兴市 / 海盐县"],
        ["详细地址", "选填", "街道门牌等"],
        ["是否内部冲突", "必填", "是 / 否；为「是」时须填冲突备注列"],
        ["行业", "必填", "筛分分选-冶金 等"],
        ["部门", "必填", "填系统中的部门全名"],
        ["申报人/负责人", "选填", "填姓名；空则默认为导入操作人"],
        ["项目动态", "必填", "技术交流 / 出方案 / 报价 / 投标 / 拟建"],
        ["联系人/线索来源等", "选填", "对应表单「其他（可选）」"],
        ["业务日期", "选填", "YYYY-MM-DD"],
        ["兼容说明", "", "旧模板列「标题」「类别」「来源」「地区」「邮箱」「中标情况」仍可识别，但不进新模板"],
    ]


def map_header_row(header_cells: tuple | list) -> dict[str, int]:
    """表头单元格 → {field: col_index}。"""
    out: dict[str, int] = {}
    ambiguous_laiyuan: list[int] = []
    for i, raw in enumerate(header_cells):
        if raw is None:
            continue
        name = str(raw).strip().replace("\n", "")
        if not name:
            continue
        if name == "来源":
            ambiguous_laiyuan.append(i)
            continue
        field = _HEADER_TO_FIELD.get(name)
        if field and field not in out:
            out[field] = i
    # 「来源」歧义：已有「类别」→ 当线索来源；否则当报备来源(自报/分发)
    for i in ambiguous_laiyuan:
        if "category" in out and "source" not in out:
            out["source"] = i
        elif "category" not in out:
            out["category"] = i
        elif "source" not in out:
            out["source"] = i
    return out


def parse_upload_rows(content: bytes, filename: str | None = None) -> list[tuple]:
    """解析上传的 xlsx/csv 为原始行列表（含表头）。"""
    fname = (filename or "").lower()
    if fname.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return [tuple(r) for r in csv.reader(io.StringIO(text))]
    wb = load_workbook(io.BytesIO(content), read_only=True)
    try:
        ws = wb.active
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def rows_for_preview(all_rows: list[tuple]) -> tuple[list[str], list[list[str]]]:
    """预览用：表头展示名 + 数据行（字符串）。"""
    if not all_rows:
        return [], []
    headers = [str(c).strip() if c is not None and str(c).strip() else f"列{i + 1}"
               for i, c in enumerate(all_rows[0])]
    data_rows: list[list[str]] = []
    for row in all_rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        while len(cells) < len(headers):
            cells.append("")
        data_rows.append(cells[: len(headers)])
    return headers, data_rows


def _cell(row: tuple | list, idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    v = row[idx]
    if v is None or v == "":
        return None
    return v


def _str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_date_cell(v: Any) -> date | str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return str(v).strip()


def parse_datetime_cell(v: Any) -> datetime | str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    return str(v).strip()


def _norm_category(v: Any) -> str | None:
    if v is None or str(v).strip() == "":
        return None
    return _CATEGORY_BY_LABEL.get(str(v).strip())


def _norm_country(v: Any) -> str | None:
    if v is None or str(v).strip() == "":
        return None
    return _COUNTRY_BY_LABEL.get(str(v).strip())


def _norm_yes_no(v: Any) -> str | None:
    if v is None or str(v).strip() == "":
        return None
    return _YES_NO.get(str(v).strip().lower()) or _YES_NO.get(str(v).strip())


def _norm_customer_type(v: Any) -> str | None:
    s = _str(v)
    if not s:
        return None
    if s in _CUSTOMER_TYPE_BY_CODE:
        return s
    return _CUSTOMER_TYPE_BY_LABEL.get(s, s)


def _norm_industry(v: Any) -> str | None:
    s = _str(v)
    if not s:
        return None
    if s in _INDUSTRY_BY_CODE:
        return s
    return _INDUSTRY_BY_LABEL.get(s, s)


def _norm_source(v: Any) -> str | None:
    s = _str(v)
    if not s:
        return None
    return _SOURCE_BY_LABEL.get(s, s)


def row_to_payload(row: tuple | list, colmap: dict[str, int]) -> dict[str, Any]:
    """把一行 Excel 转为 create_lead 可用的字段 dict（含 *_name 待解析）。"""
    def g(field: str) -> Any:
        return _cell(row, colmap.get(field))

    # 旧模板「来源」列可能是线索来源：若不是自报/分发，则归入 source
    category_raw = g("category")
    category = _norm_category(category_raw)
    source = _norm_source(g("source"))
    if category is None and category_raw is not None and colmap.get("source") is None:
        # 仅有「来源」列且不是类别值 → 当线索来源
        maybe_src = _norm_source(category_raw)
        if maybe_src and _norm_category(category_raw) is None:
            source = source or maybe_src

    payload: dict[str, Any] = {
        "title": _str(g("title")),
        "company_name": _str(g("company_name")),
        "category": category,
        "customer_type": _norm_customer_type(g("customer_type")),
        "country_type": _norm_country(g("country_type")),
        "country_name": _str(g("country_name")),
        "province": _str(g("province")),
        "city": _str(g("city")),
        "district": _str(g("district")),
        "region": _str(g("region")),
        "has_internal_conflict": _norm_yes_no(g("has_internal_conflict")),
        "conflict_note": _str(g("conflict_note")),
        "industry": _norm_industry(g("industry")),
        "bid_result": _str(g("bid_result")),
        "bid_fail_reason": _str(g("bid_fail_reason")),
        "entrust_status": _str(g("entrust_status")),
        "entrust_issued_at": parse_datetime_cell(g("entrust_issued_at")),
        "entrust_term": _str(g("entrust_term")),
        "department_name": _str(g("department_name")),
        "reporter_name": _str(g("reporter_name")),
        "owner_name": _str(g("owner_name")),
        "reported_at": parse_datetime_cell(g("reported_at")),
        "project_activity": _str(g("project_activity")),
        "demand_summary": _str(g("demand_summary")),
        "contact_name": _str(g("contact_name")),
        "contact_phone": _str(g("contact_phone")),
        "contact_email": _str(g("contact_email")),
        "source": source,
        "biz_date": parse_date_cell(g("biz_date")),
        "remark": _str(g("remark")),
    }
    # 去掉全 None，避免覆盖默认
    return {k: v for k, v in payload.items() if v is not None}
