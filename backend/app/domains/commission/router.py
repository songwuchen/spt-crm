import io
import csv

from fastapi import APIRouter, Depends, Query, UploadFile, File
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_db, get_tenant_id, require_permissions
from app.common.schemas import ok
from app.common.exceptions import BusinessException
from app.common.export import build_excel, excel_response
from app.domains.commission import service
from app.domains.commission.schemas import (
    CommissionRecordCreate, CommissionRecordUpdate, CommissionPayoutCreate,
    CommissionRuleCreate, CommissionRuleUpdate,
)

router = APIRouter(prefix="/api/v1/commissions", tags=["提成管理"])

STATUS_LABEL = {"draft": "草稿", "submitted": "待审", "approved": "已核准", "paid": "已结清"}


def _rec_dict(r) -> dict:
    f = lambda v: float(v) if v is not None else 0  # noqa: E731
    return {
        "id": r.id, "record_no": r.record_no,
        "project_id": r.project_id, "contract_id": r.contract_id,
        "customer_id": r.customer_id, "customer_name": r.customer_name,
        "owner_id": r.owner_id, "owner_name": r.owner_name,
        "department_id": r.department_id, "department_name": r.department_name,
        "signed_date": str(r.signed_date) if r.signed_date else None,
        "contract_amount": f(r.contract_amount), "received_amount": f(r.received_amount),
        "deduction_freight": f(r.deduction_freight), "deduction_service": f(r.deduction_service),
        "deduction_entertain": f(r.deduction_entertain), "deduction_rebate": f(r.deduction_rebate),
        "commission_mode": getattr(r, "commission_mode", "rate") or "rate",
        "commission_rate": f(r.commission_rate), "commission_amount": f(getattr(r, "commission_amount", 0)),
        "settle_rate": f(r.settle_rate),
        "accrued_amount": f(r.accrued_amount), "paid_amount": f(r.paid_amount),
        "current_amount": f(r.current_amount), "status": r.status, "remark": r.remark,
        "created_by_name": r.created_by_name,
        "created_at": r.created_at.isoformat() if r.created_at else "",
    }


def _payout_dict(p) -> dict:
    return {
        "id": p.id, "commission_id": p.commission_id,
        "paid_at": str(p.paid_at) if p.paid_at else None,
        "amount": float(p.amount) if p.amount is not None else 0,
        "method": p.method, "remark": p.remark, "created_by_name": p.created_by_name,
        "created_at": p.created_at.isoformat() if p.created_at else "",
    }


def _rule_dict(r) -> dict:
    return {
        "id": r.id, "name": r.name, "scope_type": r.scope_type,
        "department_id": r.department_id, "department_name": r.department_name,
        "rate": float(r.rate) if r.rate is not None else 0,
        "min_amount": float(r.min_amount) if r.min_amount is not None else None,
        "enabled": r.enabled, "sort_order": r.sort_order, "remark": r.remark,
    }


# ---------- Rules (declared before /{id} to avoid path capture) ----------
@router.get("/rules")
async def list_rules(tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
                     _u=Depends(require_permissions("commission:view"))):
    return ok([_rule_dict(r) for r in await service.list_rules(db, tenant_id)])


@router.post("/rules")
async def create_rule(body: CommissionRuleCreate, tenant_id: str = Depends(get_tenant_id),
                      db: AsyncSession = Depends(get_db), u=Depends(require_permissions("commission:manage"))):
    return ok(_rule_dict(await service.create_rule(db, tenant_id, body, u)))


@router.put("/rules/{rule_id}")
async def update_rule(rule_id: str, body: CommissionRuleUpdate, tenant_id: str = Depends(get_tenant_id),
                      db: AsyncSession = Depends(get_db), u=Depends(require_permissions("commission:manage"))):
    return ok(_rule_dict(await service.update_rule(db, tenant_id, rule_id, body, u)))


@router.delete("/rules/{rule_id}")
async def delete_rule(rule_id: str, tenant_id: str = Depends(get_tenant_id),
                      db: AsyncSession = Depends(get_db), u=Depends(require_permissions("commission:manage"))):
    await service.delete_rule(db, tenant_id, rule_id, u)
    return ok()


# ---------- Summary ----------
@router.get("/summary")
async def summary(tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
                  _u=Depends(require_permissions("commission:view"))):
    return ok(await service.summary_by_owner(db, tenant_id))


# ---------- Export ----------
@router.get("/export/excel")
async def export_excel(owner_id: str = Query(None), status: str = Query(None), keyword: str = Query(None),
                       tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
                       _u=Depends(require_permissions("commission:view"))):
    from app.config import settings
    items, _ = await service.list_records(db, tenant_id, 1, settings.MAX_EXPORT_ROWS, owner_id, status, keyword)
    headers = ["提成单号", "客户", "业务员", "部门", "合同额", "累计回款", "结算比例", "提成比例",
               "应计奖金", "已提奖金", "本次可提", "状态", "签订日期"]
    rows = []
    for r in items:
        rows.append([
            r.record_no, r.customer_name or "", r.owner_name or "", r.department_name or "",
            float(r.contract_amount or 0), float(r.received_amount or 0),
            f"{float(r.settle_rate or 0) * 100:.1f}%", f"{float(r.commission_rate or 0) * 100:.2f}%",
            float(r.accrued_amount or 0), float(r.paid_amount or 0), float(r.current_amount or 0),
            STATUS_LABEL.get(r.status or "", r.status or ""),
            str(r.signed_date) if r.signed_date else "",
        ])
    buf = build_excel("提成台账", headers, rows)
    return excel_response(buf, "commissions.xlsx")


# ---------- 批量导入（声明在 /{record_id} 之前，避免路径捕获） ----------
_COMM_IMPORT_COLS = ["客户名称", "负责人", "部门", "签约日期", "合同金额", "回款金额",
                     "运费扣减", "服务扣减", "招待扣减", "返利扣减", "提成比例(0-1)", "备注"]


@router.get("/import/template")
async def commission_import_template(_u=Depends(require_permissions("commission:view"))):
    """下载提成单导入模板。提成比例填小数（如 0.05 表示 5%）；留空则按提成政策自动套用。"""
    example = ["示例客户有限公司", "张三", "销售一部", "2026-06-01", 1000000, 600000,
               0, 0, 0, 0, 0.05, "首批结算"]
    buf = build_excel("提成单导入模板", _COMM_IMPORT_COLS, [example])
    return excel_response(buf, "commissions_template.xlsx")


@router.post("/import")
async def import_commissions(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    u=Depends(require_permissions("commission:edit")),
):
    """批量导入提成单。列顺序：客户名称, 负责人, 部门, 签约日期, 合同金额, 回款金额,
    运费扣减, 服务扣减, 招待扣减, 返利扣减, 提成比例(0-1), 备注。单行失败不中断其余行。"""
    content = await file.read()
    fname = (file.filename or "").lower()
    try:
        if fname.endswith(".csv"):
            text = content.decode("utf-8-sig", errors="replace")
            all_rows = [tuple(r) for r in csv.reader(io.StringIO(text))]
        else:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True)) if ws is not None else []
            wb.close()
    except Exception:
        raise BusinessException(message="无法解析文件，请使用导入模板（.xlsx 或 .csv）")

    if len(all_rows) < 2:
        return ok({"created": 0, "skipped": 0, "errors": []})
    data_rows = all_rows[1:]

    def _num(v):
        try:
            return float(v) if v is not None and str(v).strip() != "" else 0
        except (ValueError, TypeError):
            return None

    created, skipped, errors = 0, 0, []
    for idx, row in enumerate(data_rows, 2):
        if not row or not any(row):
            continue

        def _cell(i: int):
            return str(row[i]).strip() if len(row) > i and row[i] is not None and str(row[i]).strip() != "" else None

        if not _cell(0):  # 客户名称为空视为空行
            skipped += 1
            continue
        amounts = [_num(row[i]) if len(row) > i else 0 for i in (4, 5, 6, 7, 8, 9, 10)]
        if any(a is None for a in amounts):
            errors.append(f"第{idx}行: 金额/比例格式错误")
            continue
        contract_amt, received_amt, dfreight, dservice, dentertain, drebate, rate = amounts
        try:
            body = CommissionRecordCreate(
                customer_name=_cell(0), owner_name=_cell(1), department_name=_cell(2),
                signed_date=_cell(3),
                contract_amount=contract_amt or 0, received_amount=received_amt or 0,
                deduction_freight=dfreight or 0, deduction_service=dservice or 0,
                deduction_entertain=dentertain or 0, deduction_rebate=drebate or 0,
                commission_rate=rate or 0, remark=_cell(11),
            )
            await service.create_record(db, tenant_id, body, u)
            created += 1
        except Exception as e:
            await db.rollback()
            errors.append(f"第{idx}行: {str(e)[:80]}")
    return ok({"created": created, "skipped": skipped, "errors": errors})


# ---------- Records ----------
@router.get("")
async def list_records(pageNo: int = Query(1, ge=1), pageSize: int = Query(20, ge=1, le=100),
                       owner_id: str = Query(None), status: str = Query(None), keyword: str = Query(None),
                       filter: str = Query(None, description="高级筛选 FilterDsl(JSON)"),
                       sort_by: str = Query(None), sort_order: str = Query(None),
                       tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
                       _u=Depends(require_permissions("commission:view"))):
    items, total = await service.list_records(db, tenant_id, pageNo, pageSize, owner_id, status, keyword,
                                              adv_filter=filter, sort_by=sort_by, sort_order=sort_order, current_user=_u)
    return ok({"items": [_rec_dict(r) for r in items], "total": total, "pageNo": pageNo, "pageSize": pageSize})


@router.post("")
async def create_record(body: CommissionRecordCreate, tenant_id: str = Depends(get_tenant_id),
                        db: AsyncSession = Depends(get_db), u=Depends(require_permissions("commission:edit"))):
    return ok(_rec_dict(await service.create_record(db, tenant_id, body, u)))


@router.post("/generate/from-contract/{contract_id}")
async def generate_from_contract(contract_id: str, tenant_id: str = Depends(get_tenant_id),
                                 db: AsyncSession = Depends(get_db), u=Depends(require_permissions("commission:edit"))):
    return ok(_rec_dict(await service.generate_from_contract(db, tenant_id, contract_id, u)))


@router.get("/{record_id}")
async def get_record(record_id: str, tenant_id: str = Depends(get_tenant_id),
                     db: AsyncSession = Depends(get_db), _u=Depends(require_permissions("commission:view"))):
    return ok(_rec_dict(await service.get_record(db, tenant_id, record_id)))


@router.put("/{record_id}")
async def update_record(record_id: str, body: CommissionRecordUpdate, tenant_id: str = Depends(get_tenant_id),
                        db: AsyncSession = Depends(get_db), u=Depends(require_permissions("commission:edit"))):
    return ok(_rec_dict(await service.update_record(db, tenant_id, record_id, body, u)))


@router.delete("/{record_id}")
async def delete_record(record_id: str, tenant_id: str = Depends(get_tenant_id),
                        db: AsyncSession = Depends(get_db), u=Depends(require_permissions("commission:edit"))):
    await service.delete_record(db, tenant_id, record_id, u)
    return ok()


@router.post("/{record_id}/recalc")
async def recalc(record_id: str, tenant_id: str = Depends(get_tenant_id),
                 db: AsyncSession = Depends(get_db), u=Depends(require_permissions("commission:edit"))):
    return ok(_rec_dict(await service.recalc_received(db, tenant_id, record_id, u)))


@router.get("/{record_id}/payouts")
async def list_payouts(record_id: str, tenant_id: str = Depends(get_tenant_id),
                       db: AsyncSession = Depends(get_db), _u=Depends(require_permissions("commission:view"))):
    return ok([_payout_dict(p) for p in await service.list_payouts(db, tenant_id, record_id)])


@router.post("/{record_id}/payouts")
async def add_payout(record_id: str, body: CommissionPayoutCreate, tenant_id: str = Depends(get_tenant_id),
                     db: AsyncSession = Depends(get_db), u=Depends(require_permissions("commission:edit"))):
    return ok(_payout_dict(await service.add_payout(db, tenant_id, record_id, body, u)))
