import io
import csv

from fastapi import APIRouter, Depends, Query, UploadFile, File
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_db, get_tenant_id, require_permissions
from app.common.schemas import ok
from app.common.export import build_excel, excel_response
from app.common.code_generator import generate_code
from app.database import generate_uuid
from app.domains.measurement import service
from app.domains.measurement.models import ServiceMeasurement
from app.domains.measurement.schemas import MeasurementCreate, MeasurementUpdate

router = APIRouter(prefix="/api/v1/measurements", tags=["售后实测数据"])

# 导入表头 -> 模型字段（兼容中文表头与英文字段名；与导出表头一致，可用导出文件回填）
_MEAS_HEADER_MAP = {
    "记录号": "record_no", "record_no": "record_no",
    "客户": "customer_name", "客户名称": "customer_name", "customer_name": "customer_name",
    "服务日期": "service_date", "service_date": "service_date",
    "行业": "industry", "industry": "industry",
    "设备名称": "equipment_name", "equipment_name": "equipment_name",
    "设备型号": "equipment_model", "equipment_model": "equipment_model",
    "产品编号": "product_no", "product_no": "product_no",
    "物料": "material_name", "物料名称": "material_name", "material_name": "material_name",
    "电机功率": "motor_power_kw", "motor_power_kw": "motor_power_kw",
    "振幅": "amplitude_mm", "amplitude_mm": "amplitude_mm",
    "料层厚度": "layer_thickness_mm", "layer_thickness_mm": "layer_thickness_mm",
    "给料粒度": "feed_size_mm", "feed_size_mm": "feed_size_mm",
    "筛分效率%": "screen_efficiency", "筛分效率": "screen_efficiency", "screen_efficiency": "screen_efficiency",
    "处理量t/h": "throughput_tph", "处理量": "throughput_tph", "throughput_tph": "throughput_tph",
    "振源温度": "source_temp_c", "source_temp_c": "source_temp_c",
    "环境温度": "ambient_temp_c", "ambient_temp_c": "ambient_temp_c",
    "运行电流A": "running_current_a", "运行电流": "running_current_a", "running_current_a": "running_current_a",
    "日运行h": "daily_run_hours", "日运行": "daily_run_hours", "daily_run_hours": "daily_run_hours",
    "服务人员": "engineer_name", "engineer_name": "engineer_name",
    "结果描述": "result_desc", "问题": "issues", "备注": "remark",
}
_MEAS_TEMPLATE_HEADERS = ["记录号", "客户", "服务日期", "行业", "设备名称", "设备型号", "物料",
                          "筛分效率%", "处理量t/h", "运行电流A", "振源温度", "日运行h", "服务人员"]


def _m_dict(m) -> dict:
    f = lambda v: float(v) if v is not None else None  # noqa: E731
    return {
        "id": m.id, "record_no": m.record_no, "ticket_id": m.ticket_id,
        "customer_id": m.customer_id, "customer_name": m.customer_name,
        "service_date": str(m.service_date) if m.service_date else None,
        "engineer_id": m.engineer_id, "engineer_name": m.engineer_name,
        "industry": m.industry,
        "equipment_name": m.equipment_name, "equipment_model": m.equipment_model, "product_no": m.product_no,
        "motor_power_kw": f(m.motor_power_kw), "amplitude_mm": f(m.amplitude_mm),
        "material_name": m.material_name, "layer_thickness_mm": f(m.layer_thickness_mm),
        "feed_size_mm": f(m.feed_size_mm), "screen_efficiency": f(m.screen_efficiency),
        "throughput_tph": f(m.throughput_tph), "source_temp_c": f(m.source_temp_c),
        "ambient_temp_c": f(m.ambient_temp_c), "running_current_a": f(m.running_current_a),
        "daily_run_hours": f(m.daily_run_hours),
        "service_rating": m.service_rating, "product_rating": m.product_rating,
        "result_desc": m.result_desc, "issues": m.issues, "remark": m.remark,
        "created_by_name": m.created_by_name,
        "created_at": m.created_at.isoformat() if m.created_at else "",
    }


@router.get("/stats")
async def stats(tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
                _u=Depends(require_permissions("service:view"))):
    return ok(await service.stats_by_model(db, tenant_id))


@router.get("/export/excel")
async def export_excel(customer_id: str = Query(None), equipment_model: str = Query(None),
                       industry: str = Query(None), keyword: str = Query(None),
                       tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
                       _u=Depends(require_permissions("service:view"))):
    from app.config import settings
    items, _ = await service.list_measurements(db, tenant_id, 1, settings.MAX_EXPORT_ROWS,
                                               customer_id, None, equipment_model, industry, keyword)
    headers = ["记录号", "客户", "服务日期", "行业", "设备名称", "设备型号", "物料",
               "筛分效率%", "处理量t/h", "运行电流A", "振源温度", "日运行h", "服务人员"]
    rows = []
    for m in items:
        rows.append([
            m.record_no, m.customer_name or "", str(m.service_date) if m.service_date else "",
            m.industry or "", m.equipment_name or "", m.equipment_model or "", m.material_name or "",
            float(m.screen_efficiency) if m.screen_efficiency is not None else "",
            float(m.throughput_tph) if m.throughput_tph is not None else "",
            float(m.running_current_a) if m.running_current_a is not None else "",
            float(m.source_temp_c) if m.source_temp_c is not None else "",
            float(m.daily_run_hours) if m.daily_run_hours is not None else "",
            m.engineer_name or "",
        ])
    buf = build_excel("售后实测数据", headers, rows)
    return excel_response(buf, "measurements.xlsx")


@router.get("/template")
async def measurement_template(_u=Depends(require_permissions("service:view"))):
    """下载实测数据导入 Excel 模板（表头 + 示例行）。"""
    example = ["", "示例客户有限公司", "2026-06-01", "矿山", "圆振动筛", "YK1854", "铁矿石",
               92.5, 320, 18.6, 65, 16, "李工"]
    buf = build_excel("实测数据导入模板", _MEAS_TEMPLATE_HEADERS, [example])
    return excel_response(buf, "measurements_template.xlsx")


@router.post("/import")
async def import_measurements(file: UploadFile = File(...), tenant_id: str = Depends(get_tenant_id),
                              db: AsyncSession = Depends(get_db), u=Depends(require_permissions("service:edit"))):
    """批量导入实测数据（Excel/CSV）。表头兼容中文与英文字段名，可用导出/模板文件回填。"""
    content = await file.read()
    fname = (file.filename or "").lower()
    if fname.endswith((".xlsx", ".xls")):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header_cells = [str(c).strip() if c is not None else "" for c in (all_rows[0] if all_rows else [])]
        body_rows = [list(r) for r in all_rows[1:]] if len(all_rows) > 1 else []
    else:
        reader = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
        header_cells = [h.strip() for h in (reader[0] if reader else [])]
        body_rows = reader[1:] if len(reader) > 1 else []

    col_field = {}
    for i, h in enumerate(header_cells):
        field = _MEAS_HEADER_MAP.get(h) or _MEAS_HEADER_MAP.get(h.replace(" ", ""))
        if field:
            col_field[i] = field
    if not col_field:
        return ok({"success": 0, "failed": 0, "total": 0,
                   "errors": [{"row": 1, "reason": "未识别到有效表头，请使用导入模板"}]})

    success, errors = 0, []
    for idx, row in enumerate(body_rows, start=2):
        rec = {}
        for i, field in col_field.items():
            if i < len(row) and row[i] is not None and str(row[i]).strip() != "":
                rec[field] = str(row[i]).strip()
        if not rec:
            continue  # 跳过空行
        try:
            mc = MeasurementCreate(**rec)
        except ValidationError as ve:
            first = ve.errors()[0]
            loc = ".".join(str(x) for x in first.get("loc", []))
            errors.append({"row": idx, "reason": f"{loc}: {first.get('msg', '')}"})
            continue
        dump = mc.model_dump(exclude_unset=True)
        if not dump.get("record_no"):
            dump["record_no"] = await generate_code(db, tenant_id, "measurement")
        db.add(ServiceMeasurement(
            id=generate_uuid(), tenant_id=tenant_id,
            created_by_id=u["sub"], created_by_name=u.get("real_name") or u.get("username"),
            **dump,
        ))
        success += 1

    if success > 0:
        await db.commit()
    return ok({"success": success, "failed": len(errors), "total": success + len(errors), "errors": errors})


@router.get("")
async def list_measurements(pageNo: int = Query(1, ge=1), pageSize: int = Query(20, ge=1, le=100),
                            customer_id: str = Query(None), ticket_id: str = Query(None),
                            equipment_model: str = Query(None), industry: str = Query(None), keyword: str = Query(None),
                            tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
                            _u=Depends(require_permissions("service:view"))):
    items, total = await service.list_measurements(db, tenant_id, pageNo, pageSize, customer_id,
                                                   ticket_id, equipment_model, industry, keyword)
    return ok({"items": [_m_dict(m) for m in items], "total": total, "pageNo": pageNo, "pageSize": pageSize})


@router.post("")
async def create_measurement(body: MeasurementCreate, tenant_id: str = Depends(get_tenant_id),
                             db: AsyncSession = Depends(get_db), u=Depends(require_permissions("service:edit"))):
    return ok(_m_dict(await service.create_measurement(db, tenant_id, body, u)))


@router.get("/{mid}")
async def get_measurement(mid: str, tenant_id: str = Depends(get_tenant_id),
                          db: AsyncSession = Depends(get_db), _u=Depends(require_permissions("service:view"))):
    return ok(_m_dict(await service.get_measurement(db, tenant_id, mid)))


@router.put("/{mid}")
async def update_measurement(mid: str, body: MeasurementUpdate, tenant_id: str = Depends(get_tenant_id),
                             db: AsyncSession = Depends(get_db), u=Depends(require_permissions("service:edit"))):
    return ok(_m_dict(await service.update_measurement(db, tenant_id, mid, body, u)))


@router.delete("/{mid}")
async def delete_measurement(mid: str, tenant_id: str = Depends(get_tenant_id),
                             db: AsyncSession = Depends(get_db), u=Depends(require_permissions("service:edit"))):
    await service.delete_measurement(db, tenant_id, mid, u)
    return ok()
