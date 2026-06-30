"""Standalone contact list — cross-customer contact search and management."""
import csv
import io

from fastapi import APIRouter, Depends, Query, UploadFile, File
from openpyxl import load_workbook
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_db, get_tenant_id, get_current_user, require_permissions
from app.common.schemas import ok
from app.common.export import build_excel, excel_response
from app.database import generate_uuid
from app.domains.customer.models import Contact, Customer

router = APIRouter(prefix="/api/v1/contacts", tags=["联系人"])


@router.get("")
async def list_contacts(
    pageNo: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    keyword: str = Query(None),
    role_type: str = Query(None),
    customer_id: str = Query(None),
    filter: str = Query(None, description="高级筛选 FilterDsl(JSON)"),
    sort_by: str = Query(None),
    sort_order: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("contact:view")),
):
    q = select(Contact).where(Contact.tenant_id == tenant_id)
    count_q = select(func.count(Contact.id)).where(Contact.tenant_id == tenant_id)

    if keyword:
        like = f"%{keyword}%"
        q = q.where(
            (Contact.name.ilike(like)) | (Contact.phone.ilike(like)) |
            (Contact.mobile.ilike(like)) | (Contact.email.ilike(like))
        )
        count_q = count_q.where(
            (Contact.name.ilike(like)) | (Contact.phone.ilike(like)) |
            (Contact.mobile.ilike(like)) | (Contact.email.ilike(like))
        )
    if role_type:
        q = q.where(Contact.role_type == role_type)
        count_q = count_q.where(Contact.role_type == role_type)
    if customer_id:
        q = q.where(Contact.customer_id == customer_id)
        count_q = count_q.where(Contact.customer_id == customer_id)

    # 高级筛选（多字段/多条件）
    from app.common.search import filter_clause_or_400, resolve_sort
    clause = filter_clause_or_400("contact", filter, {"user_id": _user.get("sub")})
    if clause is not None:
        q = q.where(clause)
        count_q = count_q.where(clause)

    total = (await db.execute(count_q)).scalar() or 0
    order = resolve_sort("contact", sort_by, sort_order) or Contact.created_at.desc()
    items = (await db.execute(
        q.order_by(order)
        .offset((pageNo - 1) * pageSize).limit(pageSize)
    )).scalars().all()

    # Batch-fetch customer names
    cust_ids = list({c.customer_id for c in items if c.customer_id})
    cust_map: dict[str, str] = {}
    if cust_ids:
        rows = (await db.execute(
            select(Customer.id, Customer.name).where(Customer.id.in_(cust_ids), Customer.tenant_id == tenant_id)
        )).all()
        cust_map = {r[0]: r[1] for r in rows}

    return ok({
        "items": [{
            "id": c.id, "customer_id": c.customer_id,
            "customer_name": cust_map.get(c.customer_id, ""),
            "name": c.name, "title": c.title, "role_type": c.role_type,
            "phone": c.phone, "mobile": c.mobile, "email": c.email,
            "is_primary": c.is_primary, "remark": c.remark,
            "created_at": c.created_at.isoformat() if c.created_at else "",
        } for c in items],
        "total": total, "pageNo": pageNo, "pageSize": pageSize,
    })


@router.get("/export")
async def export_contacts(
    keyword: str = Query(None),
    role_type: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("contact:view")),
):
    """Export contacts as Excel file."""
    q = select(Contact).where(Contact.tenant_id == tenant_id)
    if keyword:
        like = f"%{keyword}%"
        q = q.where(
            (Contact.name.ilike(like)) | (Contact.phone.ilike(like)) |
            (Contact.mobile.ilike(like)) | (Contact.email.ilike(like))
        )
    if role_type:
        q = q.where(Contact.role_type == role_type)

    items = (await db.execute(q.order_by(Contact.created_at.desc()))).scalars().all()

    # Batch-fetch customer names
    cust_ids = list({c.customer_id for c in items if c.customer_id})
    cust_map: dict[str, str] = {}
    if cust_ids:
        rows = (await db.execute(
            select(Customer.id, Customer.name).where(Customer.id.in_(cust_ids), Customer.tenant_id == tenant_id)
        )).all()
        cust_map = {r[0]: r[1] for r in rows}

    headers = ["customer_name", "name", "title", "role_type", "phone", "mobile", "email", "is_primary", "remark"]
    data_rows = []
    for c in items:
        data_rows.append([
            cust_map.get(c.customer_id, ""),
            c.name or "",
            c.title or "",
            c.role_type or "",
            c.phone or "",
            c.mobile or "",
            c.email or "",
            "true" if c.is_primary else "false",
            c.remark or "",
        ])

    buf = build_excel("联系人列表", headers, data_rows)
    return excel_response(buf, "contacts.xlsx")


@router.get("/template")
async def contact_import_template(
    _user=Depends(require_permissions("contact:view")),
):
    """下载联系人导入 Excel 模板（表头 + 示例行），列与导入接口一致。"""
    headers = ["customer_name", "name", "title", "role_type", "phone", "mobile", "email", "is_primary", "remark"]
    example = [
        "示例客户有限公司", "张三", "采购经理", "procurement",
        "010-12345678", "13800000000", "zhangsan@example.com", "true", "主要采购对接人",
    ]
    buf = build_excel("联系人导入模板", headers, [example])
    return excel_response(buf, "contacts_template.xlsx")


@router.post("/import")
async def import_contacts(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("contact:create")),
):
    """Import contacts from CSV or Excel file. Matches customers by name."""
    content = await file.read()
    fname = (file.filename or "").lower()

    if fname.endswith((".xlsx", ".xls")):
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return ok({"success": 0, "failed": 0, "total": 0, "errors": []})
        headers_row = [str(c).strip() if c else "" for c in all_rows[0]]
        rows_iter = []
        for row in all_rows[1:]:
            if not row or not any(row):
                continue
            cells = [str(c).strip() if c is not None else "" for c in row]
            rows_iter.append(dict(zip(headers_row, cells)))
    else:
        text = content.decode("utf-8-sig")
        rows_iter = list(csv.DictReader(io.StringIO(text)))

    # Pre-fetch all customer name -> id mapping for this tenant
    cust_rows = (await db.execute(
        select(Customer.id, Customer.name).where(
            Customer.tenant_id == tenant_id,
            Customer.is_deleted == False,
        )
    )).all()
    cust_map: dict[str, str] = {r[1].strip(): r[0] for r in cust_rows}

    success = 0
    errors: list[dict] = []
    valid_role_types = {"decision_maker", "influencer", "user", "finance", "procurement"}

    for idx, row in enumerate(rows_iter, start=2):  # row 1 is header
        customer_name = (row.get("customer_name") or "").strip()
        name = (row.get("name") or "").strip()

        if not customer_name:
            errors.append({"row": idx, "reason": "缺少 customer_name"})
            continue
        if not name:
            errors.append({"row": idx, "reason": "缺少 name"})
            continue

        customer_id = cust_map.get(customer_name)
        if not customer_id:
            errors.append({"row": idx, "reason": f"未找到客户: {customer_name}"})
            continue

        role_type = (row.get("role_type") or "").strip() or None
        if role_type and role_type not in valid_role_types:
            errors.append({"row": idx, "reason": f"无效的角色类型: {role_type}"})
            continue

        is_primary_raw = (row.get("is_primary") or "").strip().lower()
        is_primary = is_primary_raw in ("true", "1", "是", "yes")

        contact = Contact(
            id=generate_uuid(),
            tenant_id=tenant_id,
            customer_id=customer_id,
            name=name,
            title=(row.get("title") or "").strip() or None,
            role_type=role_type,
            phone=(row.get("phone") or "").strip() or None,
            mobile=(row.get("mobile") or "").strip() or None,
            email=(row.get("email") or "").strip() or None,
            is_primary=is_primary,
            remark=(row.get("remark") or "").strip() or None,
        )
        db.add(contact)
        success += 1

    if success > 0:
        await db.commit()

    return ok({
        "success": success,
        "failed": len(errors),
        "total": success + len(errors),
        "errors": errors,
    })
