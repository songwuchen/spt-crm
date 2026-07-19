from fastapi import APIRouter, Depends, Query, UploadFile, File
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from datetime import datetime, timezone, timedelta
import io

from app.dependencies import get_db, get_tenant_id, get_current_user, require_permissions
from app.common.schemas import ok
from app.common.export import build_excel, build_excel_multi, build_template, excel_response
from app.domains.customer.models import Customer
from app.domains.lowcode.field_permission import strip_entity_dicts
from app.domains.customer.schemas import (
    CustomerCreate, CustomerUpdate, CustomerOut,
    ContactCreate, ContactUpdate, ContactOut,
    RelationCreate, ShareCreate, BatchReleaseRequest,
    CustomerPoolCreate, CustomerPoolUpdate,
)
from app.domains.customer import service

router = APIRouter(prefix="/api/v1/customers", tags=["客户管理"])


def _format_region(c) -> str:
    """展示用地区串：结构化省·市·区优先，回退 legacy 自由文本 region。"""
    parts = [p for p in (c.province, c.city, c.district) if p]
    return " · ".join(parts) if parts else (c.region or "")


async def _tenant_pool_rules(db: AsyncSession, tenant_id: str) -> dict | None:
    """租户级公海回收规则(security_policy_json.pool_rules)，用于计算「预计回收时间」。"""
    from app.domains.admin.models import TenantProfile
    p = (await db.execute(
        select(TenantProfile).where(TenantProfile.tenant_id == tenant_id)
    )).scalar_one_or_none()
    return (p.security_policy_json or {}).get("pool_rules") if p else None


def _expected_recycle_iso(c, rules: dict | None) -> str | None:
    """在职有主客户「预计回收时间」= 最后活动(或创建)时间 + 该级别闲置天数。规则未启用/公海客户返回 None。"""
    if not rules or not rules.get("enabled"):
        return None
    if c.status != "active" or not c.owner_id:
        return None
    base = c.last_activity_at or c.created_at
    if not base:
        return None
    idle = (rules.get("idle_days") or {}).get(c.level or "D", rules.get("default_idle_days", 30))
    try:
        idle = int(idle)
    except (TypeError, ValueError):
        idle = 30
    return (base + timedelta(days=idle)).isoformat()


def _effective_recycle_rules(c, tenant_rules: dict | None, pools) -> dict | None:
    """客户实际生效的回收规则：其(将)归属区域公海的规则优先，回退租户级——与 reminder_worker 归池逻辑保持一致，
    避免列表展示的「预计回收时间」与实际自动回收行为不一致。"""
    routed = service.match_pool(c, pools) if pools else None
    if routed and routed.rules_json and routed.rules_json.get("enabled"):
        return routed.rules_json
    if tenant_rules and tenant_rules.get("enabled"):
        return tenant_rules
    return None


def _customer_dict(c, tenant_rules: dict | None = None, pools=None) -> dict:
    base_activity = c.last_activity_at or c.created_at
    idle_days = max(0, (datetime.now(timezone.utc) - base_activity).days) if base_activity else None
    eff_rules = _effective_recycle_rules(c, tenant_rules, pools)
    return {
        "id": c.id, "customer_code": c.customer_code,
        "name": c.name, "short_name": c.short_name,
        "industry": c.industry, "scale_level": c.scale_level,
        "region": c.region,
        "province": c.province, "city": c.city, "district": c.district, "region_code": c.region_code,
        "address": c.address, "website": c.website,
        "owner_id": c.owner_id, "owner_name": c.owner_name,
        "source": c.source, "level": c.level, "status": c.status,
        "tags_json": c.tags_json, "remark": c.remark, "custom_fields_json": c.custom_fields_json,
        # 商机要素 / 采购意向
        "intent_level": c.intent_level, "key_contact_id": c.key_contact_id,
        "demand": c.demand, "need_match_level": c.need_match_level,
        "budget_amount": float(c.budget_amount) if c.budget_amount is not None else None,
        "expected_purchase_date": c.expected_purchase_date.isoformat() if c.expected_purchase_date else None,
        "headcount": c.headcount,
        # 公司档案
        "industry_l1": c.industry_l1, "industry_l2": c.industry_l2, "industry_l3": c.industry_l3,
        "country": c.country, "postal_code": c.postal_code, "currency": c.currency,
        # 归属 / 审计
        "department_id": c.department_id, "department_name": c.department_name,
        "created_by_id": c.created_by_id, "created_by_name": c.created_by_name,
        "updated_by_id": c.updated_by_id, "updated_by_name": c.updated_by_name,
        # 跟进 / 公海生命周期（含派生指标）
        "last_activity_at": c.last_activity_at.isoformat() if c.last_activity_at else None,
        "last_activity_by_id": c.last_activity_by_id,
        "last_activity_by_name": c.last_activity_by_name,
        "idle_days": idle_days,
        "expected_recycle_at": _expected_recycle_iso(c, eff_rules),
        "won_deal_count": c.won_deal_count or 0,
        "pool_id": c.pool_id, "pool_source": c.pool_source,
        "pool_entered_at": c.pool_entered_at.isoformat() if c.pool_entered_at else None,
        "created_at": c.created_at.isoformat() if c.created_at else "",
        "updated_at": c.updated_at.isoformat() if c.updated_at else "",
    }


@router.get("")
async def list_customers(
    pageNo: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    keyword: str = Query(None),
    industry: str = Query(None),
    region: str = Query(None),
    region_code: str = Query(None, description="行政区划编码前缀，支持逗号分隔多个前缀(大区)"),
    owner_id: str = Query(None),
    tag: str = Query(None),
    filter: str = Query(None, description="高级筛选 FilterDsl(JSON)"),
    sort_by: str = Query(None),
    sort_order: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    # 数据范围「本人」= 负责人/创建人/共享给本人；「部门」= 部门子树成员所属；「全部」= 不限。
    # 由 service 内 apply_data_scope 统一处理（owner_id 为前端显式筛选，仍受数据范围约束）。
    items, total = await service.list_customers(
        db, tenant_id, pageNo, pageSize, keyword, industry, region, owner_id, tag=tag, current_user=_user,
        adv_filter=filter, sort_by=sort_by, sort_order=sort_order, region_code=region_code)
    tenant_rules = await _tenant_pool_rules(db, tenant_id)
    pools = await service.list_active_pools(db, tenant_id)
    dicts = [_customer_dict(c, tenant_rules, pools) for c in items]
    await strip_entity_dicts(db, tenant_id, "customer", dicts, _user.get("roles"))  # 字段级权限：读取剔除隐藏扩展字段
    return ok({"items": dicts, "total": total, "pageNo": pageNo, "pageSize": pageSize})


@router.post("")
async def create_customer(
    body: CustomerCreate,
    to_pool: bool = Query(False, description="若为 true，则创建到公海（无负责人，status=pool）"),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:create")),
):
    if to_pool:
        # Strip any owner the form might have included and post-mark as pool.
        body = body.model_copy(update={"owner_id": None})
    c = await service.create_customer(db, tenant_id, body, current_user)
    if to_pool:
        # create_customer always assigns an owner (creator fallback) — flip it to pool state.
        c.owner_id = None
        c.owner_name = None
        c.department_id = None
        c.department_name = None
        c.status = "pool"
        c.pool_source = "self_built"
        c.pool_entered_at = datetime.now(timezone.utc)
        c.pool_id = await service._route_pool_id(db, tenant_id, c)
        await db.commit()
        await db.refresh(c)
    return ok(_customer_dict(c))


@router.get("/export/excel")
async def export_customers_excel(
    keyword: str = Query(None),
    industry: str = Query(None),
    region: str = Query(None),
    owner_id: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    from app.config import settings
    items, _ = await service.list_customers(db, tenant_id, 1, settings.MAX_EXPORT_ROWS, keyword, industry, region, owner_id, current_user=_user)
    headers = ["客户编码", "客户名称", "简称", "行业", "规模", "地区", "地址", "负责人", "来源", "级别", "状态", "创建时间"]
    # 导出与列表/详情同口径脱敏（隐藏→空、脱敏→***）
    from app.domains.lowcode.field_permission import entity_field_restrictions, export_cell
    rst = await entity_field_restrictions(db, tenant_id, "customer", _user.get("roles"))
    rows = []
    for c in items:
        rows.append([
            c.customer_code, export_cell(rst, "name", c.name), c.short_name or "", c.industry or "",
            c.scale_level or "", _format_region(c), export_cell(rst, "address", c.address or ""),
            c.owner_name or "", c.source or "", c.level or "", c.status or "",
            c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
        ])
    buf = build_excel("客户列表", headers, rows)
    return excel_response(buf, "customers.xlsx")


# ---- Region Distribution ----

@router.get("/region-distribution")
async def region_distribution(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    """Get customer count grouped by region.

    优先按结构化省份(province)分组，回退到 legacy 自由文本 region，
    保证前端地图/统计基于规范化的省级数据。"""
    from sqlalchemy import func
    region_expr = func.coalesce(Customer.province, Customer.region)
    rows = (await db.execute(
        select(region_expr.label("region"), func.count(Customer.id).label("count")).where(
            Customer.tenant_id == tenant_id,
            Customer.is_deleted == False,
            region_expr.isnot(None),
            region_expr != "",
        ).group_by(region_expr)
    )).all()
    return ok([{"region": r.region, "count": r.count} for r in rows])


# ---- Customer Pool (公海池) ----

@router.get("/pool")
async def list_pool_customers(
    pageNo: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    keyword: str = Query(None),
    industry: str = Query(None),
    region: str = Query(None),
    pool_id: str = Query(None, description="按区域公海筛选；'__default__' 表示默认公海(未归池)"),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    items, total = await service.list_pool_customers(
        db, tenant_id, pageNo, pageSize, keyword, industry, region, pool_id=pool_id)
    return ok({"items": [_customer_dict(c) for c in items], "total": total, "pageNo": pageNo, "pageSize": pageSize})


@router.get("/pool/export/excel")
async def export_pool_customers_excel(
    keyword: str = Query(None),
    industry: str = Query(None),
    region: str = Query(None),
    pool_id: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    """Export pool customers to Excel."""
    from app.config import settings
    items, _ = await service.list_pool_customers(
        db, tenant_id, 1, settings.MAX_EXPORT_ROWS, keyword, industry, region, pool_id=pool_id)
    headers = ["客户编码", "客户名称", "简称", "行业", "规模", "地区", "地址", "来源", "级别", "释放时间"]
    rows = []
    for c in items:
        rows.append([
            c.customer_code, c.name, c.short_name or "", c.industry or "",
            c.scale_level or "", _format_region(c), c.address or "",
            c.source or "", c.level or "",
            c.updated_at.strftime("%Y-%m-%d %H:%M") if c.updated_at else "",
        ])
    buf = build_excel("客户公海", headers, rows)
    return excel_response(buf, "customer_pool.xlsx")


@router.post("/{customer_id}/release")
async def release_to_pool(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    c = await service.release_to_pool(db, tenant_id, customer_id, current_user)
    return ok(_customer_dict(c))


@router.post("/{customer_id}/claim")
async def claim_from_pool(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    c = await service.claim_from_pool(db, tenant_id, customer_id, current_user)
    return ok(_customer_dict(c))


@router.post("/batch_release")
async def batch_release(
    body: BatchReleaseRequest,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    ids = body.customer_ids
    released = await service.batch_release_to_pool(db, tenant_id, ids, current_user)
    return ok({"released": released})


@router.get("/import/template")
async def download_import_template():
    """Download an Excel template for customer import."""
    headers = ["客户名称", "简称", "行业", "规模等级", "区域", "地址", "来源", "级别"]
    sample = [["示例客户A", "示例A", "电子制造", "大型", "华东", "上海市浦东新区XX路", "官网", "A"]]
    buf = build_template("客户导入模板", headers, sample)
    return excel_response(buf, "customer_import_template.xlsx")


@router.post("/import/preview")
async def import_preview(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:create")),
):
    """Parse Excel/CSV and return headers + rows + duplicate detection."""
    from sqlalchemy import select as sa_select
    content = await file.read()
    fname = (file.filename or "").lower()
    if fname.endswith(".csv"):
        import csv as csv_mod
        text = content.decode("utf-8-sig")
        reader = csv_mod.reader(text.splitlines())
        all_rows = [tuple(r) for r in reader]
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    if not all_rows:
        return ok({"headers": [], "rows": [], "duplicates": []})

    headers = [str(c).strip() if c else f"列{i+1}" for i, c in enumerate(all_rows[0])]
    data_rows = []
    names = []
    for row in all_rows[1:]:
        if not row or not any(row):
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        # pad to header length
        while len(cells) < len(headers):
            cells.append("")
        data_rows.append(cells[:len(headers)])
        if cells:
            names.append(cells[0])

    # Detect duplicates by name
    from app.domains.customer.models import Customer
    existing = set()
    if names:
        result = await db.execute(
            sa_select(Customer.name).where(
                Customer.tenant_id == tenant_id,
                Customer.is_deleted == False,
                Customer.name.in_(names),
            )
        )
        existing = {r[0] for r in result.all()}

    duplicates = [i for i, name in enumerate(names) if name in existing]

    # Row-level validation errors
    errors: dict[int, str] = {}
    for i, row in enumerate(data_rows):
        if not row or not row[0] or not row[0].strip():
            errors[i] = "客户名称不能为空"

    return ok({"headers": headers, "rows": data_rows, "duplicates": duplicates, "errors": errors})


@router.post("/import/excel")
async def import_customers_excel(
    file: UploadFile = File(...),
    to_pool: bool = Query(False, description="若为 true，则导入到公海（无负责人）"),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:create")),
):
    """Import customers from Excel/CSV with field mapping and duplicate handling."""
    from starlette.datastructures import UploadFile as StarletteUpload
    import json as json_mod

    content = await file.read()
    fname = (file.filename or "").lower()
    if fname.endswith(".csv"):
        import csv as csv_mod
        text = content.decode("utf-8-sig")
        reader = csv_mod.reader(text.splitlines())
        all_rows = [tuple(r) for r in reader]
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    if len(all_rows) < 2:
        return ok({"created": 0, "skipped": 0, "errors": []})

    # Parse field_mapping and skip_duplicates from form multipart (sent as query or form fields)
    from starlette.requests import Request
    # Fallback: default column mapping
    default_fields = ["name", "short_name", "industry", "scale_level", "region", "address", "source", "level"]

    headers = all_rows[0]
    data_rows = all_rows[1:]

    # Detect duplicates
    from sqlalchemy import select as sa_select
    from app.domains.customer.models import Customer
    names = []
    for row in data_rows:
        if row and row[0]:
            names.append(str(row[0]).strip())
    existing_names = set()
    if names:
        result = await db.execute(
            sa_select(Customer.name).where(
                Customer.tenant_id == tenant_id,
                Customer.is_deleted == False,
                Customer.name.in_(names),
            )
        )
        existing_names = {r[0] for r in result.all()}

    created = 0
    skipped = 0
    errors = []
    for idx, row in enumerate(data_rows, 2):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if name in existing_names:
            skipped += 1
            continue
        try:
            data = CustomerCreate(
                name=name,
                short_name=str(row[1]).strip() if len(row) > 1 and row[1] else None,
                industry=str(row[2]).strip() if len(row) > 2 and row[2] else None,
                scale_level=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                region=str(row[4]).strip() if len(row) > 4 and row[4] else None,
                address=str(row[5]).strip() if len(row) > 5 and row[5] else None,
                source=str(row[6]).strip() if len(row) > 6 and row[6] else None,
                level=str(row[7]).strip() if len(row) > 7 and row[7] else None,
            )
            c = await service.create_customer(db, tenant_id, data, current_user)
            if to_pool:
                c.owner_id = None
                c.owner_name = None
                c.status = "pool"
            created += 1
            existing_names.add(name)  # prevent duplicates within same batch
        except Exception as e:
            errors.append(f"第{idx}行: {str(e)[:80]}")
    if to_pool:
        await db.commit()
    return ok({"created": created, "skipped": skipped, "errors": errors})


@router.get("/{customer_id}/stats")
async def customer_stats(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    """Customer 360 stats: project/quote/contract/ticket counts + totals."""
    from sqlalchemy import select, func
    from app.domains.project.models import OpportunityProject
    from app.domains.quote.models import Quote
    from app.domains.contract.models import Contract
    from app.domains.service_ticket.models import ServiceTicket

    project_count = (await db.execute(
        select(func.count(OpportunityProject.id)).where(
            OpportunityProject.tenant_id == tenant_id, OpportunityProject.customer_id == customer_id)
    )).scalar() or 0

    project_amount = (await db.execute(
        select(func.sum(OpportunityProject.amount_expect)).where(
            OpportunityProject.tenant_id == tenant_id, OpportunityProject.customer_id == customer_id,
            OpportunityProject.status == "active")
    )).scalar() or 0

    # Get project IDs for this customer
    proj_ids = (await db.execute(
        select(OpportunityProject.id).where(
            OpportunityProject.tenant_id == tenant_id, OpportunityProject.customer_id == customer_id)
    )).scalars().all()

    quote_count = 0
    contract_count = 0
    contract_amount = 0
    if proj_ids:
        quote_count = (await db.execute(
            select(func.count(Quote.id)).where(Quote.tenant_id == tenant_id, Quote.project_id.in_(proj_ids))
        )).scalar() or 0
        contract_count = (await db.execute(
            select(func.count(Contract.id)).where(Contract.tenant_id == tenant_id, Contract.project_id.in_(proj_ids))
        )).scalar() or 0
        contract_amount = (await db.execute(
            select(func.sum(Contract.amount_total)).where(
                Contract.tenant_id == tenant_id, Contract.project_id.in_(proj_ids), Contract.status == "signed")
        )).scalar() or 0

    ticket_count = (await db.execute(
        select(func.count(ServiceTicket.id)).where(
            ServiceTicket.tenant_id == tenant_id, ServiceTicket.customer_id == customer_id)
    )).scalar() or 0

    ticket_open = (await db.execute(
        select(func.count(ServiceTicket.id)).where(
            ServiceTicket.tenant_id == tenant_id, ServiceTicket.customer_id == customer_id,
            ServiceTicket.status.in_(["open", "in_progress"]))
    )).scalar() or 0

    # Won/lost counts
    won_count = (await db.execute(
        select(func.count(OpportunityProject.id)).where(
            OpportunityProject.tenant_id == tenant_id, OpportunityProject.customer_id == customer_id,
            OpportunityProject.status == "won")
    )).scalar() or 0
    lost_count = (await db.execute(
        select(func.count(OpportunityProject.id)).where(
            OpportunityProject.tenant_id == tenant_id, OpportunityProject.customer_id == customer_id,
            OpportunityProject.status == "lost")
    )).scalar() or 0

    # Payment collection rate
    collection_rate = 0.0
    if proj_ids:
        from app.domains.payment.models import PaymentPlan, PaymentRecord
        plan_total = float((await db.execute(
            select(func.coalesce(func.sum(PaymentPlan.amount), 0)).where(
                PaymentPlan.tenant_id == tenant_id, PaymentPlan.project_id.in_(proj_ids))
        )).scalar() or 0)
        received_total = float((await db.execute(
            select(func.coalesce(func.sum(PaymentRecord.amount), 0)).where(
                PaymentRecord.tenant_id == tenant_id, PaymentRecord.project_id.in_(proj_ids))
        )).scalar() or 0)
        collection_rate = (received_total / plan_total) if plan_total > 0 else 0.0

    win_rate = (won_count / (won_count + lost_count)) if (won_count + lost_count) > 0 else 0.0

    return ok({
        "project_count": project_count,
        "project_amount": float(project_amount),
        "quote_count": quote_count,
        "contract_count": contract_count,
        "contract_amount": float(contract_amount or 0),
        "ticket_count": ticket_count,
        "ticket_open": ticket_open,
        "won_count": won_count,
        "lost_count": lost_count,
        "win_rate": round(win_rate, 3),
        "collection_rate": round(collection_rate, 3),
    })


async def _gather_customer_report(db: AsyncSession, tenant_id: str, customer_id: str) -> dict:
    """Collect a customer's related business entities (商机/报价/合同/订单/标书/回款/工单/交付)."""
    from sqlalchemy import func
    from app.domains.project.models import OpportunityProject
    from app.domains.quote.models import Quote, QuoteVersion
    from app.domains.contract.models import Contract
    from app.domains.service_ticket.models import ServiceTicket
    from app.domains.payment.models import PaymentPlan, PaymentRecord
    from app.domains.delivery.models import DeliveryMilestone
    from app.domains.order.models import Order
    from app.domains.tender.models import Tender

    customer = await service.get_customer(db, tenant_id, customer_id)

    projects = (await db.execute(
        select(OpportunityProject).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.customer_id == customer_id,
            OpportunityProject.is_deleted == False,
        ).order_by(OpportunityProject.created_at.desc())
    )).scalars().all()
    proj_ids = [p.id for p in projects]

    quotes, contracts, plans, records, deliveries = [], [], [], [], []
    if proj_ids:
        quotes = (await db.execute(
            select(Quote).where(Quote.tenant_id == tenant_id, Quote.project_id.in_(proj_ids))
            .order_by(Quote.created_at.desc())
        )).scalars().all()
        contracts = (await db.execute(
            select(Contract).where(Contract.tenant_id == tenant_id, Contract.project_id.in_(proj_ids))
            .order_by(Contract.created_at.desc())
        )).scalars().all()
        plans = (await db.execute(
            select(PaymentPlan).where(PaymentPlan.tenant_id == tenant_id, PaymentPlan.project_id.in_(proj_ids))
        )).scalars().all()
        records = (await db.execute(
            select(PaymentRecord).where(PaymentRecord.tenant_id == tenant_id, PaymentRecord.project_id.in_(proj_ids))
        )).scalars().all()
        deliveries = (await db.execute(
            select(DeliveryMilestone).where(DeliveryMilestone.tenant_id == tenant_id, DeliveryMilestone.project_id.in_(proj_ids))
            .order_by(DeliveryMilestone.sort_order)
        )).scalars().all()

    # current-version amount lookup for quotes
    quote_amount: dict[str, float] = {}
    if quotes:
        qv_rows = (await db.execute(
            select(QuoteVersion.quote_id, QuoteVersion.version_no, QuoteVersion.price_total)
            .where(QuoteVersion.tenant_id == tenant_id, QuoteVersion.quote_id.in_([q.id for q in quotes]))
        )).all()
        cur_ver = {q.id: q.current_version_no for q in quotes}
        for qid, vno, price in qv_rows:
            if cur_ver.get(qid) == vno and price is not None:
                quote_amount[qid] = float(price)

    orders = (await db.execute(
        select(Order).where(Order.tenant_id == tenant_id, Order.customer_id == customer_id, Order.is_deleted == False)
        .order_by(Order.created_at.desc())
    )).scalars().all()
    tenders = (await db.execute(
        select(Tender).where(Tender.tenant_id == tenant_id, Tender.customer_id == customer_id, Tender.is_deleted == False)
        .order_by(Tender.created_at.desc())
    )).scalars().all()
    tickets = (await db.execute(
        select(ServiceTicket).where(ServiceTicket.tenant_id == tenant_id, ServiceTicket.customer_id == customer_id)
        .order_by(ServiceTicket.created_at.desc())
    )).scalars().all()

    return {
        "customer": customer,
        "projects": projects,
        "quotes": quotes,
        "quote_amount": quote_amount,
        "contracts": contracts,
        "orders": orders,
        "tenders": tenders,
        "payment_plans": plans,
        "payment_records": records,
        "tickets": tickets,
        "deliveries": deliveries,
    }


@router.get("/{customer_id}/report")
async def customer_report(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    """客户关联报表：商机/报价/合同/订单/标书/回款/工单/交付 的明细 + 汇总。"""
    data = await _gather_customer_report(db, tenant_id, customer_id)

    c = data["customer"]
    projects = data["projects"]
    quotes = data["quotes"]
    qa = data["quote_amount"]
    contracts = data["contracts"]
    orders = data["orders"]
    tenders = data["tenders"]
    plans = data["payment_plans"]
    records = data["payment_records"]
    tickets = data["tickets"]
    deliveries = data["deliveries"]

    def _f(v):
        return float(v) if v is not None else None

    summary = {
        "project_count": len(projects),
        "quote_count": len(quotes),
        "contract_count": len(contracts),
        "contract_amount": sum(_f(x.amount_total) or 0 for x in contracts if x.status == "signed"),
        "order_count": len(orders),
        "order_amount": sum(_f(o.amount) or 0 for o in orders),
        "tender_count": len(tenders),
        "tender_won": sum(1 for t in tenders if t.status == "won"),
        "payment_plan_total": sum(_f(p.amount) or 0 for p in plans),
        "payment_received_total": sum(_f(r.amount) or 0 for r in records),
        "ticket_count": len(tickets),
        "delivery_count": len(deliveries),
    }

    return ok({
        "customer": {"id": c.id, "name": c.name, "customer_code": c.customer_code},
        "summary": summary,
        "projects": [{
            "id": p.id, "project_code": p.project_code, "name": p.name,
            "stage_code": p.stage_code, "amount_expect": _f(p.amount_expect),
            "status": p.status,
            "created_at": p.created_at.isoformat() if p.created_at else "",
        } for p in projects],
        "quotes": [{
            "id": q.id, "quote_no": q.quote_no, "project_id": q.project_id,
            "current_version_no": q.current_version_no, "status": q.status,
            "amount": qa.get(q.id),
            "created_at": q.created_at.isoformat() if q.created_at else "",
        } for q in quotes],
        "contracts": [{
            "id": x.id, "contract_no": x.contract_no, "project_id": x.project_id,
            "status": x.status, "signed_date": str(x.signed_date) if x.signed_date else None,
            "amount_total": _f(x.amount_total),
            "created_at": x.created_at.isoformat() if x.created_at else "",
        } for x in contracts],
        "orders": [{
            "id": o.id, "order_no": o.order_no, "title": o.title,
            "amount": _f(o.amount), "currency": o.currency, "status": o.status,
            "order_date": str(o.order_date) if o.order_date else None,
            "delivery_date": str(o.delivery_date) if o.delivery_date else None,
        } for o in orders],
        "tenders": [{
            "id": t.id, "tender_no": t.tender_no, "title": t.title,
            "bid_amount": _f(t.bid_amount), "budget_amount": _f(t.budget_amount),
            "status": t.status, "result": t.result,
            "submit_date": str(t.submit_date) if t.submit_date else None,
            "open_date": str(t.open_date) if t.open_date else None,
        } for t in tenders],
        "payment_plans": [{
            "id": p.id, "plan_no": p.plan_no, "due_date": str(p.due_date) if p.due_date else None,
            "amount": _f(p.amount), "status": p.status,
        } for p in plans],
        "payment_records": [{
            "id": r.id, "received_date": str(r.received_date) if r.received_date else None,
            "amount": _f(r.amount), "channel": r.channel, "reference_no": r.reference_no,
        } for r in records],
        "tickets": [{
            "id": t.id, "ticket_no": t.ticket_no, "type": t.type, "status": t.status,
            "priority": t.priority,
            "created_at": t.created_at.isoformat() if t.created_at else "",
        } for t in tickets],
        "deliveries": [{
            "id": d.id, "milestone_code": d.milestone_code, "name": d.name,
            "plan_date": str(d.plan_date) if d.plan_date else None,
            "actual_date": str(d.actual_date) if d.actual_date else None,
            "status": d.status,
        } for d in deliveries],
    })


@router.get("/{customer_id}/report/export")
async def export_customer_report(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    """导出客户关联报表为多 Sheet Excel。"""
    data = await _gather_customer_report(db, tenant_id, customer_id)
    qa = data["quote_amount"]

    def _f(v):
        return float(v) if v is not None else ""

    def _dt(v):
        return v.strftime("%Y-%m-%d %H:%M") if v else ""

    sheets = [
        ("商机", ["项目编码", "名称", "阶段", "预期金额", "状态", "创建时间"],
         [[p.project_code, p.name or "", p.stage_code or "", _f(p.amount_expect), p.status or "", _dt(p.created_at)]
          for p in data["projects"]]),
        ("报价", ["报价单号", "版本", "金额", "状态", "创建时间"],
         [[q.quote_no, q.current_version_no, qa.get(q.id, ""), q.status or "", _dt(q.created_at)]
          for q in data["quotes"]]),
        ("合同", ["合同编号", "状态", "签约日期", "金额", "创建时间"],
         [[x.contract_no, x.status or "", str(x.signed_date) if x.signed_date else "", _f(x.amount_total), _dt(x.created_at)]
          for x in data["contracts"]]),
        ("订单", ["订单号", "标题", "金额", "币种", "状态", "下单日期", "交付日期", "负责人"],
         [[o.order_no, o.title or "", _f(o.amount), o.currency or "", o.status or "",
           str(o.order_date) if o.order_date else "", str(o.delivery_date) if o.delivery_date else "", o.owner_name or ""]
          for o in data["orders"]]),
        ("标书", ["标书号", "标题", "投标金额", "预算金额", "状态", "提交日期", "开标日期", "结果", "负责人"],
         [[t.tender_no, t.title or "", _f(t.bid_amount), _f(t.budget_amount), t.status or "",
           str(t.submit_date) if t.submit_date else "", str(t.open_date) if t.open_date else "", t.result or "", t.owner_name or ""]
          for t in data["tenders"]]),
        ("回款计划", ["计划号", "到期日", "金额", "状态"],
         [[p.plan_no, str(p.due_date) if p.due_date else "", _f(p.amount), p.status or ""]
          for p in data["payment_plans"]]),
        ("回款记录", ["收款日期", "金额", "渠道", "参考号"],
         [[str(r.received_date) if r.received_date else "", _f(r.amount), r.channel or "", r.reference_no or ""]
          for r in data["payment_records"]]),
        ("工单", ["工单号", "类型", "状态", "优先级", "创建时间"],
         [[t.ticket_no, t.type or "", t.status or "", t.priority or "", _dt(t.created_at)]
          for t in data["tickets"]]),
        ("交付", ["里程碑", "名称", "计划日期", "实际日期", "状态"],
         [[d.milestone_code, d.name or "", str(d.plan_date) if d.plan_date else "",
           str(d.actual_date) if d.actual_date else "", d.status or ""]
          for d in data["deliveries"]]),
    ]
    buf = build_excel_multi(sheets)
    return excel_response(buf, f"customer_report_{customer_id}.xlsx")


@router.delete("/{customer_id}")
async def delete_customer(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:delete")),
):
    await service.delete_customer(db, tenant_id, customer_id, current_user)
    return ok()


# ---- Contact sub-resource ----
@router.get("/{customer_id}/contacts")
async def list_contacts(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("contact:view")),
):
    contacts = await service.list_contacts(db, tenant_id, customer_id)
    dicts = [ContactOut.model_validate(c).model_dump() for c in contacts]
    await strip_entity_dicts(db, tenant_id, "contact", dicts, _user.get("roles"))  # 字段级权限：读取剔除隐藏扩展字段
    return ok(dicts)


@router.post("/{customer_id}/contacts")
async def create_contact(
    customer_id: str,
    body: ContactCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("contact:create")),
):
    c = await service.create_contact(db, tenant_id, customer_id, body, current_user)
    return ok(ContactOut.model_validate(c).model_dump())


@router.put("/{customer_id}/contacts/{contact_id}")
async def update_contact(
    customer_id: str,
    contact_id: str,
    body: ContactUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("contact:edit")),
):
    c = await service.update_contact(db, tenant_id, contact_id, body, current_user)
    return ok(ContactOut.model_validate(c).model_dump())


@router.delete("/{customer_id}/contacts/{contact_id}")
async def delete_contact(
    customer_id: str,
    contact_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("contact:delete")),
):
    await service.delete_contact(db, tenant_id, contact_id, current_user)
    return ok()


# ---- Customer Relations ----
@router.get("/{customer_id}/relations")
async def list_relations(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    items = await service.list_relations(db, tenant_id, customer_id)
    return ok([{
        "id": r.id, "from_customer_id": r.from_customer_id,
        "to_customer_id": r.to_customer_id, "relation_type": r.relation_type,
        "note": r.note, "created_at": r.created_at.isoformat() if r.created_at else "",
    } for r in items])


@router.post("/{customer_id}/relations")
async def create_relation(
    customer_id: str,
    body: RelationCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    data = {"from_customer_id": customer_id, **body.model_dump()}
    r = await service.create_relation(db, tenant_id, data, current_user)
    return ok({"id": r.id, "relation_type": r.relation_type})


@router.delete("/{customer_id}/relations/{relation_id}")
async def delete_relation(
    customer_id: str,
    relation_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    await service.delete_relation(db, tenant_id, relation_id, current_user)
    return ok()


# ---- ACL Shares (generic) ----
@router.get("/{customer_id}/shares")
async def list_shares(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    items = await service.list_shares(db, tenant_id, "customer", customer_id)
    return ok([{
        "id": s.id, "biz_type": s.biz_type, "biz_id": s.biz_id,
        "shared_to_type": s.shared_to_type, "shared_to_id": s.shared_to_id,
        "shared_to_name": s.shared_to_name, "permission": s.permission,
        "shared_by_name": s.shared_by_name,
        "created_at": s.created_at.isoformat() if s.created_at else "",
    } for s in items])


@router.post("/{customer_id}/shares")
async def create_share(
    customer_id: str,
    body: ShareCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    data = {"biz_type": "customer", "biz_id": customer_id, **body.model_dump()}
    s = await service.create_share(db, tenant_id, data, current_user)
    return ok({"id": s.id, "shared_to_name": s.shared_to_name})


@router.delete("/{customer_id}/shares/{share_id}")
async def delete_share(
    customer_id: str,
    share_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    await service.delete_share(db, tenant_id, share_id, current_user)
    return ok()


@router.get("/check-similar")
async def check_similar(
    name: str = Query(None, min_length=2),
    phone: str = Query(None, min_length=4),
    exclude_id: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Find customers with similar names or matching contact phone for duplicate detection."""
    from sqlalchemy import or_
    from app.domains.customer.models import Contact

    results: list[dict] = []
    seen_ids: set[str] = set()

    # Match by name / short_name
    if name:
        q = select(Customer.id, Customer.name, Customer.short_name, Customer.industry, Customer.owner_name).where(
            Customer.tenant_id == tenant_id,
            Customer.is_deleted == False,
            or_(Customer.name.ilike(f"%{name}%"), Customer.short_name.ilike(f"%{name}%")),
        )
        if exclude_id:
            q = q.where(Customer.id != exclude_id)
        items = (await db.execute(q.limit(5))).all()
        for r in items:
            seen_ids.add(r.id)
            results.append({"id": r.id, "name": r.name, "short_name": r.short_name,
                            "industry": r.industry, "owner_name": r.owner_name, "match_type": "name"})

    # Match by contact phone
    if phone and len(phone) >= 4:
        phone_q = (
            select(Contact.customer_id, Contact.phone, Contact.name.label("contact_name"))
            .where(
                Contact.tenant_id == tenant_id,
                or_(Contact.phone.ilike(f"%{phone}%"), Contact.mobile.ilike(f"%{phone}%")),
            )
        )
        if exclude_id:
            phone_q = phone_q.where(Contact.customer_id != exclude_id)
        phone_matches = (await db.execute(phone_q.limit(5))).all()
        cust_ids = [m.customer_id for m in phone_matches if m.customer_id not in seen_ids]
        if cust_ids:
            custs = (await db.execute(
                select(Customer).where(Customer.id.in_(cust_ids), Customer.tenant_id == tenant_id, Customer.is_deleted == False)
            )).scalars().all()
            cust_map = {c.id: c for c in custs}
            for m in phone_matches:
                if m.customer_id in seen_ids:
                    continue
                c = cust_map.get(m.customer_id)
                if c:
                    seen_ids.add(c.id)
                    results.append({"id": c.id, "name": c.name, "short_name": c.short_name,
                                    "industry": c.industry, "owner_name": c.owner_name,
                                    "match_type": "phone", "match_phone": m.phone,
                                    "match_contact": m.contact_name})

    return ok(results[:10])


@router.get("/check-unique")
async def check_unique(
    field: str = Query(..., pattern=r"^(name|customer_code)$"),
    value: str = Query(..., min_length=1),
    exclude_id: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Check if a customer field value is unique within the tenant."""
    col = Customer.name if field == "name" else Customer.customer_code
    q = select(Customer.id).where(
        Customer.tenant_id == tenant_id,
        Customer.is_deleted == False,
        col == value,
    )
    if exclude_id:
        q = q.where(Customer.id != exclude_id)
    exists = (await db.execute(q)).scalar() is not None
    return ok({"unique": not exists})


# ---- Batch Operations ----

from pydantic import BaseModel as PydanticBaseModel


class CustomerMergeBody(PydanticBaseModel):
    primary_id: str
    secondary_id: str


@router.post("/merge")
async def merge_customers(
    body: CustomerMergeBody,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    """Merge secondary customer into primary, moving all related data."""
    c = await service.merge_customers(db, tenant_id, body.primary_id, body.secondary_id, current_user)
    return ok(_customer_dict(c))


class BatchTransferBody(PydanticBaseModel):
    ids: list[str]
    owner_id: str
    owner_name: str | None = None


@router.post("/batch_transfer")
async def batch_transfer(
    body: BatchTransferBody,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    """Batch transfer customers to a new owner. Also lifts pool customers back to active
    so this endpoint doubles as 'assign from pool to salesperson'."""
    from sqlalchemy import update, case
    result = await db.execute(
        update(Customer).where(
            Customer.tenant_id == tenant_id,
            Customer.id.in_(body.ids),
            Customer.is_deleted == False,
        ).values(
            owner_id=body.owner_id,
            owner_name=body.owner_name,
            status=case((Customer.status == "pool", "active"), else_=Customer.status),
        )
    )
    await db.commit()
    return ok({"updated": result.rowcount})


@router.get("/{customer_id}/health")
async def customer_health(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Compute a health score (0-100, grade A-D) for a customer."""
    from datetime import datetime, timedelta
    from sqlalchemy import func
    from app.domains.project.models import OpportunityProject
    from app.domains.payment.models import PaymentPlan, PaymentRecord
    from app.domains.service_ticket.models import ServiceTicket
    from app.domains.activity.models import Activity

    now = datetime.utcnow()
    d90 = now - timedelta(days=90)

    # 1. Recent activity count (max 30 pts)
    activity_count = (await db.execute(
        select(func.count()).where(
            Activity.tenant_id == tenant_id,
            Activity.biz_type == "customer",
            Activity.biz_id == customer_id,
            Activity.created_at >= d90,
        )
    )).scalar_one() or 0
    activity_score = min(30, activity_count * 5)

    # 2. Active projects (max 25 pts)
    active_projects = (await db.execute(
        select(func.count()).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.customer_id == customer_id,
            OpportunityProject.is_deleted == False,
            OpportunityProject.stage_code.notin_(["S6", "lost"]),
        )
    )).scalar_one() or 0
    project_score = min(25, active_projects * 10)

    # 3. Payment health (max 25 pts) — ratio of on-time payments
    total_plans = (await db.execute(
        select(func.count()).select_from(PaymentPlan).join(
            OpportunityProject, OpportunityProject.id == PaymentPlan.project_id
        ).where(
            PaymentPlan.tenant_id == tenant_id,
            OpportunityProject.customer_id == customer_id,

        )
    )).scalar_one() or 0
    overdue_plans = (await db.execute(
        select(func.count()).select_from(PaymentPlan).join(
            OpportunityProject, OpportunityProject.id == PaymentPlan.project_id
        ).where(
            PaymentPlan.tenant_id == tenant_id,
            OpportunityProject.customer_id == customer_id,

            PaymentPlan.status == "overdue",
        )
    )).scalar_one() or 0
    if total_plans > 0:
        payment_score = int(25 * (1 - overdue_plans / total_plans))
    else:
        payment_score = 15  # neutral if no plans

    # 4. Service ticket health (max 20 pts) — fewer open tickets = better
    open_tickets = (await db.execute(
        select(func.count()).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.customer_id == customer_id,

            ServiceTicket.status.in_(["open", "in_progress"]),
        )
    )).scalar_one() or 0
    ticket_score = max(0, 20 - open_tickets * 5)

    total = activity_score + project_score + payment_score + ticket_score
    grade = "A" if total >= 80 else "B" if total >= 60 else "C" if total >= 40 else "D"

    return ok({
        "score": total,
        "grade": grade,
        "breakdown": {
            "activity": {"score": activity_score, "max": 30, "detail": f"近90天{activity_count}次互动"},
            "project": {"score": project_score, "max": 25, "detail": f"{active_projects}个活跃商机"},
            "payment": {"score": payment_score, "max": 25, "detail": f"{overdue_plans}/{total_plans}逾期"},
            "service": {"score": ticket_score, "max": 20, "detail": f"{open_tickets}个待处理工单"},
        },
    })


# ---- Batch Messaging ----

class BatchMessageBody(PydanticBaseModel):
    customer_ids: list[str]
    channel: str  # email / sms
    subject: str | None = None
    content: str


@router.post("/batch_message")
async def batch_message(
    body: BatchMessageBody,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    """Send batch messages (email/sms) to primary contacts of selected customers."""
    from app.domains.customer.models import Contact

    # Get primary contacts for selected customers
    contacts = (await db.execute(
        select(Contact).where(
            Contact.tenant_id == tenant_id,
            Contact.customer_id.in_(body.customer_ids),
            Contact.is_primary == True,
        )
    )).scalars().all()

    # Fallback: if no primary contact, pick first contact per customer
    covered_ids = {c.customer_id for c in contacts}
    missing_ids = [cid for cid in body.customer_ids if cid not in covered_ids]
    if missing_ids:
        fallback = (await db.execute(
            select(Contact).where(
                Contact.tenant_id == tenant_id,
                Contact.customer_id.in_(missing_ids),
            ).order_by(Contact.created_at)
        )).scalars().all()
        seen = set()
        for c in fallback:
            if c.customer_id not in seen:
                contacts.append(c)
                seen.add(c.customer_id)

    sent = 0
    failed = 0
    results = []
    for contact in contacts:
        target = contact.email if body.channel == "email" else contact.mobile or contact.phone
        if not target:
            results.append({"customer_id": contact.customer_id, "contact": contact.name, "status": "skipped", "reason": "无联系方式"})
            failed += 1
            continue

        if body.channel == "email":
            try:
                from app.common.email_service import send_email
                await send_email(
                    to=target,
                    subject=body.subject or "来自CRM的消息",
                    body=body.content,
                )
                results.append({"customer_id": contact.customer_id, "contact": contact.name, "target": target, "status": "sent"})
                sent += 1
            except Exception as e:
                results.append({"customer_id": contact.customer_id, "contact": contact.name, "target": target, "status": "failed", "reason": str(e)[:100]})
                failed += 1
        else:
            # SMS: log only (no real SMS provider integrated)
            results.append({"customer_id": contact.customer_id, "contact": contact.name, "target": target, "status": "queued"})
            sent += 1

    # Audit log
    from app.domains.audit.service import log_action
    await log_action(db, tenant_id=tenant_id, user_id=current_user["id"],
                     user_name=current_user.get("real_name", ""), action="batch_message",
                     resource_type="customer", detail=f"channel={body.channel}, sent={sent}, failed={failed}")

    return ok({"sent": sent, "failed": failed, "results": results})


# ---- Single customer CRUD (must be AFTER all fixed-path routes to avoid /{customer_id} swallowing /pool etc.) ----

@router.get("/{customer_id}")
async def get_customer(
    customer_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    c = await service.get_customer(db, tenant_id, customer_id)
    d = _customer_dict(c, await _tenant_pool_rules(db, tenant_id), await service.list_active_pools(db, tenant_id))
    await strip_entity_dicts(db, tenant_id, "customer", [d], _user.get("roles"))  # 字段级权限：读取剔除隐藏扩展字段
    return ok(d)


@router.put("/{customer_id}")
async def update_customer(
    customer_id: str,
    body: CustomerUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("customer:edit")),
):
    c = await service.update_customer(db, tenant_id, customer_id, body, current_user)
    return ok(_customer_dict(c, await _tenant_pool_rules(db, tenant_id), await service.list_active_pools(db, tenant_id)))


# ==================== 区域公海配置 (customer_pools) ====================
# 读取用 customer:view（公海页筛选需要）；增删改属租户配置，用 role:manage（管理员）。

pool_router = APIRouter(prefix="/api/v1/customer-pools", tags=["区域公海"])


def _pool_dict(p, counts: dict | None = None) -> dict:
    return {
        "id": p.id, "name": p.name, "description": p.description,
        "region_scope": p.region_scope, "rules_json": p.rules_json,
        "is_default": p.is_default, "is_active": p.is_active, "sort_order": p.sort_order,
        "customer_count": (counts or {}).get(p.id, 0),
        "created_at": p.created_at.isoformat() if p.created_at else "",
    }


@pool_router.get("")
async def list_customer_pools(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("customer:view")),
):
    pools = await service.list_pools(db, tenant_id)
    counts = await service.pool_counts(db, tenant_id)
    return ok({"items": [_pool_dict(p, counts) for p in pools],
               "default_count": counts.get("__default__", 0)})


@pool_router.post("")
async def create_customer_pool(
    body: CustomerPoolCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("role:manage")),
):
    p = await service.create_pool(db, tenant_id, body.model_dump(), current_user)
    return ok(_pool_dict(p))


@pool_router.put("/{pool_id}")
async def update_customer_pool(
    pool_id: str,
    body: CustomerPoolUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("role:manage")),
):
    p = await service.update_pool(db, tenant_id, pool_id, body.model_dump(exclude_unset=True), current_user)
    return ok(_pool_dict(p))


@pool_router.delete("/{pool_id}")
async def delete_customer_pool(
    pool_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("role:manage")),
):
    await service.delete_pool(db, tenant_id, pool_id, current_user)
    return ok({"deleted": True})
