import io
from datetime import datetime, timezone, date
from typing import Optional
from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy import select, func, extract, or_, and_
from sqlalchemy.ext.asyncio import AsyncSession

from pydantic import BaseModel, Field
from typing import Optional
from app.dependencies import get_db, get_tenant_id, get_current_user, require_permissions
from app.common.schemas import ok
from app.common.export import build_template, excel_response
from app.common.data_scope import resolve_owner_scope, visible_customer_ids_select
from app.domains.customer.models import Customer
from app.domains.lead.models import Lead
from app.domains.project.models import OpportunityProject
from app.domains.quote.models import Quote
from app.domains.solution.models import Solution
from app.domains.delivery.models import DeliveryMilestone
from app.domains.payment.models import Invoice, PaymentRecord
from app.domains.change.models import ChangeRequest
from app.domains.service_ticket.models import ServiceTicket
from app.domains.activity.models import Activity
from app.domains.ai_center.models import AiTask
from app.domains.contract.models import Contract
from app.domains.payment.models import PaymentPlan
from app.domains.dashboard.models import SalesTarget, DashboardSnapshot

router = APIRouter(prefix="/api/v1/dashboard", tags=["工作台"])


# ---------- 数据可见范围（工作台 / 报表中心的统一口径） ----------
#
# 本域每个聚合都必须按调用者的数据范围过滤：否则 data_scope=self 的业务员即使在
# 客户/商机列表里查不到任何数据，也能从「总数、业绩排行、到期预警、风险预警」里
# 反推出全租户的客户名、同事姓名与成交额；预警卡片还带 biz_id 可点开，等于给出了
# 一条越权进详情的点击路径。报表中心的 Excel/PDF 导出走的是同一批聚合，同理。
#
# resolve_owner_scope 返回 None = 不限（管理员 / data_scope=all）——这类用户的
# 每个 _*_scope_where 都返回空列表，SQL 与改造前逐字相同，数字不会有任何变化。
# 各 helper 的判定口径分别对齐 apply_data_scope / apply_project_child_scope /
# visible_customer_ids_select，保证「聚合里算得到的，点开也读得到」。


async def _scope_owner_ids(db: AsyncSession, tenant_id: str, user: dict) -> list[str] | None:
    """本次请求可见的 owner_id 集合；None 表示不限。每个接口只解析一次后向下透传。"""
    return await resolve_owner_scope(db, user, tenant_id)


def _project_scope_where(tenant_id: str, user: dict, scope: list[str] | None) -> list:
    """商机可见条件，口径对齐 apply_data_scope(OpportunityProject, "project")：
    归属在范围内 / 本人创建 / 共享给我 / 我是项目成员。"""
    if scope is None:
        return []
    uid = user.get("sub", "")
    conds = [
        OpportunityProject.owner_id.in_(scope),
        OpportunityProject.created_by_id == uid,
    ]
    try:
        from app.domains.customer.models import AclShare
        conds.append(OpportunityProject.id.in_(select(AclShare.biz_id).where(
            AclShare.tenant_id == tenant_id,
            AclShare.biz_type == "project",
            or_(AclShare.shared_to_id == uid, AclShare.shared_to_type == "all"),
        )))
    except Exception:
        pass
    try:
        from app.domains.project.models import ProjectMember
        conds.append(OpportunityProject.id.in_(select(ProjectMember.project_id).where(
            ProjectMember.tenant_id == tenant_id,
            ProjectMember.user_id == uid,
        )))
    except Exception:
        pass
    return [or_(*conds)]


def _visible_project_ids(tenant_id: str, user: dict, scope: list[str] | None):
    """可见商机 id 子查询，供「挂在商机下、自身没有 owner_id」的实体做父级过滤。"""
    return select(OpportunityProject.id).where(
        OpportunityProject.tenant_id == tenant_id,
        *_project_scope_where(tenant_id, user, scope),
    )


def _child_scope_where(model, tenant_id: str, user: dict, scope: list[str] | None) -> list:
    """商机子实体（报价/方案/里程碑/发票/回款计划/回款记录/变更/合同…）的可见条件。

    这些表没有 owner_id，可见性只能由父商机决定，口径对齐 apply_project_child_scope：
    父商机可见 / 本行由本人创建 / 本行指派给本人。合同的 project_id 可为空（外部导入的
    独立合同），这种行只能靠 assignee/created_by 命中，与合同列表的行为一致。
    """
    if scope is None:
        return []
    uid = user.get("sub", "")
    conds = [model.project_id.in_(_visible_project_ids(tenant_id, user, scope))]
    if hasattr(model, "created_by_id"):
        conds.append(model.created_by_id == uid)
    if hasattr(model, "assignee_id"):
        conds.append(model.assignee_id == uid)
    return [or_(*conds)]


def _lead_scope_where(tenant_id: str, user: dict, scope: list[str] | None) -> list:
    """线索可见条件，口径对齐 apply_data_scope(Lead, "lead")。"""
    if scope is None:
        return []
    uid = user.get("sub", "")
    conds = [Lead.owner_id.in_(scope), Lead.created_by_id == uid]
    try:
        from app.domains.customer.models import AclShare
        conds.append(Lead.id.in_(select(AclShare.biz_id).where(
            AclShare.tenant_id == tenant_id,
            AclShare.biz_type == "lead",
            or_(AclShare.shared_to_id == uid, AclShare.shared_to_type == "all"),
        )))
    except Exception:
        pass
    return [or_(*conds)]


async def _customer_scope_where(db: AsyncSession, tenant_id: str, user: dict) -> list:
    """客户可见条件。直接复用共享的 visible_customer_ids_select（含公海/共享口径），
    避免工作台自己再造一套判定后与客户列表对不上。"""
    visible_ids = await visible_customer_ids_select(db, tenant_id, user)
    return [] if visible_ids is None else [Customer.id.in_(visible_ids)]


async def _signed_contract_amount(db: AsyncSession, tenant_id: str, *extra_where) -> float:
    """已签合同总额（赢单成交额统一口径）。

    金额取 contracts.amount_total，仅统计 status='signed' 的合同。
    extra_where 可追加按 signed_date 的年月/区间等过滤，用于趋势、导出等按期统计。
    """
    q = select(func.coalesce(func.sum(Contract.amount_total), 0)).where(
        Contract.tenant_id == tenant_id,
        Contract.status == "signed",
        *extra_where,
    )
    return float((await db.execute(q)).scalar() or 0)


async def _signed_contract_count(db: AsyncSession, tenant_id: str, *extra_where) -> int:
    """已签合同数（赢单数统一口径，与成交额取同一份合同集）。

    extra_where 可追加按 signed_date 的年月/区间过滤，用于趋势等按期统计。
    """
    q = select(func.count(Contract.id)).where(
        Contract.tenant_id == tenant_id,
        Contract.status == "signed",
        *extra_where,
    )
    return int((await db.execute(q)).scalar() or 0)


@router.get("/stats")
async def stats(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    now = datetime.now(timezone.utc)
    scope = await _scope_owner_ids(db, tenant_id, _user)
    cust_where = await _customer_scope_where(db, tenant_id, _user)
    lead_where = _lead_scope_where(tenant_id, _user, scope)
    proj_where = _project_scope_where(tenant_id, _user, scope)

    customer_total = (await db.execute(
        select(func.count(Customer.id)).where(Customer.tenant_id == tenant_id, *cust_where)
    )).scalar() or 0

    lead_total = (await db.execute(
        select(func.count(Lead.id)).where(Lead.tenant_id == tenant_id, *lead_where)
    )).scalar() or 0

    monthly_new_customers = (await db.execute(
        select(func.count(Customer.id)).where(
            Customer.tenant_id == tenant_id,
            extract("year", Customer.created_at) == now.year,
            extract("month", Customer.created_at) == now.month,
            *cust_where,
        )
    )).scalar() or 0

    pending_leads = (await db.execute(
        select(func.count(Lead.id)).where(
            Lead.tenant_id == tenant_id,
            Lead.status.in_(["new", "following"]),
            *lead_where,
        )
    )).scalar() or 0

    project_total = (await db.execute(
        select(func.count(OpportunityProject.id)).where(
            OpportunityProject.tenant_id == tenant_id, *proj_where)
    )).scalar() or 0

    active_projects = (await db.execute(
        select(func.count(OpportunityProject.id)).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.status == "active",
            *proj_where,
        )
    )).scalar() or 0

    quote_total = (await db.execute(
        select(func.count(Quote.id)).where(
            Quote.tenant_id == tenant_id, *_child_scope_where(Quote, tenant_id, _user, scope))
    )).scalar() or 0

    solution_total = (await db.execute(
        select(func.count(Solution.id)).where(
            Solution.tenant_id == tenant_id, *_child_scope_where(Solution, tenant_id, _user, scope))
    )).scalar() or 0

    milestone_where = _child_scope_where(DeliveryMilestone, tenant_id, _user, scope)
    milestone_total = (await db.execute(
        select(func.count(DeliveryMilestone.id)).where(
            DeliveryMilestone.tenant_id == tenant_id, *milestone_where)
    )).scalar() or 0

    milestone_delayed = (await db.execute(
        select(func.count(DeliveryMilestone.id)).where(
            DeliveryMilestone.tenant_id == tenant_id,
            DeliveryMilestone.status == "delayed",
            *milestone_where,
        )
    )).scalar() or 0

    invoice_total = (await db.execute(
        select(func.count(Invoice.id)).where(
            Invoice.tenant_id == tenant_id, *_child_scope_where(Invoice, tenant_id, _user, scope))
    )).scalar() or 0

    payment_received = (await db.execute(
        select(func.coalesce(func.sum(PaymentRecord.amount), 0)).where(
            PaymentRecord.tenant_id == tenant_id,
            *_child_scope_where(PaymentRecord, tenant_id, _user, scope),
        )
    )).scalar() or 0

    change_total = (await db.execute(
        select(func.count(ChangeRequest.id)).where(
            ChangeRequest.tenant_id == tenant_id,
            *_child_scope_where(ChangeRequest, tenant_id, _user, scope))
    )).scalar() or 0

    ticket_total = (await db.execute(
        select(func.count(ServiceTicket.id)).where(ServiceTicket.tenant_id == tenant_id)
    )).scalar() or 0

    ticket_open = (await db.execute(
        select(func.count(ServiceTicket.id)).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.status.in_(["open", "assigned", "in_progress"]),
        )
    )).scalar() or 0

    # 动态/AI 任务只记 created_by_id，按「本人产生的记录」计数即可；
    # 受限用户看自己的活动量，不限范围的用户仍是全租户口径。
    activity_where = [] if scope is None else [Activity.created_by_id == _user.get("sub")]
    activity_total = (await db.execute(
        select(func.count(Activity.id)).where(Activity.tenant_id == tenant_id, *activity_where)
    )).scalar() or 0

    ai_where = [] if scope is None else [AiTask.created_by_id == _user.get("sub")]
    ai_task_total = (await db.execute(
        select(func.count(AiTask.id)).where(AiTask.tenant_id == tenant_id, *ai_where)
    )).scalar() or 0

    # Pipeline value
    pipeline_value = (await db.execute(
        select(func.coalesce(func.sum(OpportunityProject.amount_expect), 0)).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.status == "active",
            *proj_where,
        )
    )).scalar() or 0

    contract_total = (await db.execute(
        select(func.count(Contract.id)).where(
            Contract.tenant_id == tenant_id, *_child_scope_where(Contract, tenant_id, _user, scope))
    )).scalar() or 0

    return ok({
        "customer_total": customer_total,
        "lead_total": lead_total,
        "monthly_new_customers": monthly_new_customers,
        "pending_leads": pending_leads,
        "project_total": project_total,
        "active_projects": active_projects,
        "quote_total": quote_total,
        "solution_total": solution_total,
        "milestone_total": milestone_total,
        "milestone_delayed": milestone_delayed,
        "invoice_total": invoice_total,
        "payment_received": payment_received,
        "change_total": change_total,
        "ticket_total": ticket_total,
        "ticket_open": ticket_open,
        "activity_total": activity_total,
        "ai_task_total": ai_task_total,
        "pipeline_value": float(pipeline_value),
        "contract_total": contract_total,
    })


@router.get("/trends")
async def trends(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Month-over-month trends for key metrics."""
    now = datetime.now(timezone.utc)
    # Current month
    cur_year, cur_month = now.year, now.month
    # Previous month
    prev_month = cur_month - 1
    prev_year = cur_year
    if prev_month <= 0:
        prev_month += 12
        prev_year -= 1

    scope = await _scope_owner_ids(db, tenant_id, _user)
    cust_where = await _customer_scope_where(db, tenant_id, _user)
    lead_where = _lead_scope_where(tenant_id, _user, scope)
    proj_where = _project_scope_where(tenant_id, _user, scope)

    def _month_filter(model_cls, date_col, y, m, scope_where=()):
        col = getattr(model_cls, date_col)
        return select(func.count(model_cls.id)).where(
            model_cls.tenant_id == tenant_id,
            extract("year", col) == y,
            extract("month", col) == m,
            *scope_where,
        )

    cur_customers = (await db.execute(_month_filter(Customer, "created_at", cur_year, cur_month, cust_where))).scalar() or 0
    prev_customers = (await db.execute(_month_filter(Customer, "created_at", prev_year, prev_month, cust_where))).scalar() or 0

    cur_leads = (await db.execute(_month_filter(Lead, "created_at", cur_year, cur_month, lead_where))).scalar() or 0
    prev_leads = (await db.execute(_month_filter(Lead, "created_at", prev_year, prev_month, lead_where))).scalar() or 0

    cur_projects = (await db.execute(_month_filter(OpportunityProject, "created_at", cur_year, cur_month, proj_where))).scalar() or 0
    prev_projects = (await db.execute(_month_filter(OpportunityProject, "created_at", prev_year, prev_month, proj_where))).scalar() or 0

    # 工单没有归属维度（service_tickets 全租户共享，服务台按 service:view 授权），保持原口径
    cur_tickets = (await db.execute(_month_filter(ServiceTicket, "created_at", cur_year, cur_month))).scalar() or 0
    prev_tickets = (await db.execute(_month_filter(ServiceTicket, "created_at", prev_year, prev_month))).scalar() or 0

    def _trend(cur: int, prev: int) -> dict:
        diff = cur - prev
        if prev == 0:
            pct = 100 if cur > 0 else 0
        else:
            pct = round((diff / prev) * 100)
        return {"current": cur, "previous": prev, "diff": diff, "pct": pct}

    return ok({
        "customers": _trend(cur_customers, prev_customers),
        "leads": _trend(cur_leads, prev_leads),
        "projects": _trend(cur_projects, prev_projects),
        "tickets": _trend(cur_tickets, prev_tickets),
    })


@router.get("/alerts")
async def alerts(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Dashboard alerts: stalled projects, overdue milestones, high-risk projects."""
    now = datetime.now(timezone.utc)
    alerts_list = []

    # 预警卡片带 biz_id 且前端可点开，等于一条进详情的入口——必须按数据范围过滤，
    # 否则会把别人的商机名/回款金额直接摆到受限用户面前。
    scope = await _scope_owner_ids(db, tenant_id, _user)
    proj_where = _project_scope_where(tenant_id, _user, scope)

    # Stalled projects (no update in 14+ days)
    from sqlalchemy import text
    stalled = (await db.execute(
        select(OpportunityProject).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.status == "active",
            OpportunityProject.updated_at < func.now() - text("interval '14 days'"),
            *proj_where,
        ).order_by(OpportunityProject.updated_at.asc()).limit(10)
    )).scalars().all()
    for p in stalled:
        days = (now - p.updated_at.replace(tzinfo=timezone.utc)).days if p.updated_at else 0
        alerts_list.append({
            "type": "stalled", "severity": "warning" if days < 30 else "critical",
            "title": f"商机停滞 {days} 天",
            "content": p.name,
            "biz_type": "project", "biz_id": p.id,
        })

    # Delayed milestones
    delayed = (await db.execute(
        select(DeliveryMilestone).where(
            DeliveryMilestone.tenant_id == tenant_id,
            DeliveryMilestone.status == "delayed",
            *_child_scope_where(DeliveryMilestone, tenant_id, _user, scope),
        ).limit(10)
    )).scalars().all()
    for m in delayed:
        alerts_list.append({
            "type": "milestone_delayed", "severity": "warning",
            "title": f"里程碑延期: {m.name or m.milestone_code}",
            "content": f"计划日期: {m.plan_date}",
            "biz_type": "project", "biz_id": m.project_id,
        })

    # High risk projects
    high_risk = (await db.execute(
        select(OpportunityProject).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.status == "active",
            OpportunityProject.risk_level == "H",
            *proj_where,
        ).limit(10)
    )).scalars().all()
    for p in high_risk:
        alerts_list.append({
            "type": "high_risk", "severity": "critical",
            "title": "高风险商机",
            "content": p.name,
            "biz_type": "project", "biz_id": p.id,
        })

    # Open tickets —— 工单本身没有归属维度（服务台按 service:view 授权、列表就是全租户可见），
    # 这里不额外收窄，否则工作台会比工单列表还严，服务台同事看不到自己该处理的告警。
    open_tickets = (await db.execute(
        select(ServiceTicket).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.status.in_(["open"]),
            ServiceTicket.priority.in_(["high", "critical"]),
        ).limit(10)
    )).scalars().all()
    for t in open_tickets:
        alerts_list.append({
            "type": "urgent_ticket", "severity": "warning",
            "title": f"紧急工单: {t.ticket_no}",
            "content": t.description[:50] if t.description else "",
            "biz_type": "service_ticket", "biz_id": t.id,
        })

    # Overdue payments
    from datetime import timedelta
    overdue_plans = (await db.execute(
        select(PaymentPlan).where(
            PaymentPlan.tenant_id == tenant_id,
            PaymentPlan.status == "pending",
            PaymentPlan.due_date != None,
            PaymentPlan.due_date < date.today(),
            *_child_scope_where(PaymentPlan, tenant_id, _user, scope),
        ).limit(10)
    )).scalars().all()
    for p in overdue_plans:
        overdue_days = (date.today() - p.due_date).days
        alerts_list.append({
            "type": "payment_overdue", "severity": "critical" if overdue_days > 30 else "warning",
            "title": f"回款逾期 {overdue_days} 天: {p.plan_no}",
            "content": f"金额: ¥{float(p.amount or 0):,.0f}，到期: {p.due_date}",
            "biz_type": "project", "biz_id": p.project_id,
        })

    # Large amount projects with no quote
    big_no_quote = (await db.execute(
        select(OpportunityProject).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.status == "active",
            OpportunityProject.amount_expect > 100000,
            OpportunityProject.stage_code.in_(["S3", "S4"]),
            *proj_where,
        ).limit(10)
    )).scalars().all()
    for p in big_no_quote:
        has_quote = (await db.execute(
            select(func.count()).where(Quote.project_id == p.id, Quote.tenant_id == tenant_id)
        )).scalar()
        if not has_quote:
            alerts_list.append({
                "type": "no_quote", "severity": "info",
                "title": f"大额商机未报价: {p.name}",
                "content": f"预期金额 ¥{float(p.amount_expect or 0):,.0f}，阶段 {p.stage_code}",
                "biz_type": "project", "biz_id": p.id,
            })

    return ok(alerts_list)


@router.get("/funnel")
async def funnel(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Sales funnel: count + amount by stage for active projects."""
    stages = ["S1", "S2", "S3", "S4", "S5", "S6"]
    stage_labels = {"S1": "线索确认", "S2": "需求分析", "S3": "方案报价", "S4": "商务谈判", "S5": "合同签订", "S6": "交付验收"}
    scope = await _scope_owner_ids(db, tenant_id, _user)
    proj_where = _project_scope_where(tenant_id, _user, scope)
    result = []
    for s in stages:
        row = await db.execute(
            select(
                func.count(OpportunityProject.id),
                func.coalesce(func.sum(OpportunityProject.amount_expect), 0),
            ).where(
                OpportunityProject.tenant_id == tenant_id,
                OpportunityProject.status == "active",
                OpportunityProject.stage_code == s,
                *proj_where,
            )
        )
        count, amount = row.one()
        result.append({
            "stage": s, "label": stage_labels[s],
            "count": count, "amount": float(amount),
        })
    return ok(result)


@router.get("/win_loss")
async def win_loss(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Win/loss stats."""
    scope = await _scope_owner_ids(db, tenant_id, _user)
    proj_where = _project_scope_where(tenant_id, _user, scope)
    contract_where = _child_scope_where(Contract, tenant_id, _user, scope)

    won_count = (await db.execute(
        select(func.count(OpportunityProject.id)).where(
            OpportunityProject.tenant_id == tenant_id, OpportunityProject.status == "won",
            *proj_where,
        )
    )).scalar() or 0
    lost_count = (await db.execute(
        select(func.count(OpportunityProject.id)).where(
            OpportunityProject.tenant_id == tenant_id, OpportunityProject.status == "lost",
            *proj_where,
        )
    )).scalar() or 0
    # 赢单金额 = 已签合同额（签单额口径），不再用商机预计额
    won_amount = await _signed_contract_amount(db, tenant_id, *contract_where)
    lost_amount = (await db.execute(
        select(func.coalesce(func.sum(OpportunityProject.amount_expect), 0)).where(
            OpportunityProject.tenant_id == tenant_id, OpportunityProject.status == "lost",
            *proj_where,
        )
    )).scalar() or 0
    total = won_count + lost_count
    win_rate = round(won_count / total * 100, 1) if total > 0 else 0

    return ok({
        "won_count": won_count, "lost_count": lost_count,
        "won_amount": float(won_amount), "lost_amount": float(lost_amount),
        "win_rate": win_rate,
    })


@router.get("/top_customers")
async def top_customers(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Top 10 customers by project pipeline value."""
    from app.domains.customer.models import Customer
    # 返回客户 id + 名称（前端可点开），两侧都要卡：客户本身可见 且 计入的商机可见，
    # 否则受限用户会从「TOP 客户」里读到别人的客户名和管道金额。
    scope = await _scope_owner_ids(db, tenant_id, _user)
    cust_where = await _customer_scope_where(db, tenant_id, _user)
    proj_where = _project_scope_where(tenant_id, _user, scope)
    result = await db.execute(
        select(
            Customer.id, Customer.name,
            func.count(OpportunityProject.id).label("project_count"),
            func.coalesce(func.sum(OpportunityProject.amount_expect), 0).label("total_amount"),
        ).join(
            OpportunityProject, OpportunityProject.customer_id == Customer.id
        ).where(
            Customer.tenant_id == tenant_id,
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.status == "active",
            *cust_where,
            *proj_where,
        ).group_by(Customer.id, Customer.name)
        .order_by(func.sum(OpportunityProject.amount_expect).desc())
        .limit(10)
    )
    rows = result.all()
    return ok([{
        "id": r.id, "name": r.name,
        "project_count": r.project_count,
        "total_amount": float(r.total_amount),
    } for r in rows])


@router.get("/payment_overview")
async def payment_overview(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Payment overview: total planned, received, overdue."""
    from app.domains.payment.service import mark_overdue_plans
    await mark_overdue_plans(db, tenant_id)

    from sqlalchemy import text, cast, Date as SaDate
    now = datetime.now(timezone.utc).date()

    scope = await _scope_owner_ids(db, tenant_id, _user)
    plan_where = _child_scope_where(PaymentPlan, tenant_id, _user, scope)
    record_where = _child_scope_where(PaymentRecord, tenant_id, _user, scope)
    contract_where = _child_scope_where(Contract, tenant_id, _user, scope)

    # Total planned
    total_planned = (await db.execute(
        select(func.coalesce(func.sum(PaymentPlan.amount), 0)).where(
            PaymentPlan.tenant_id == tenant_id,
            *plan_where,
        )
    )).scalar() or 0

    # Total received
    total_received = (await db.execute(
        select(func.coalesce(func.sum(PaymentRecord.amount), 0)).where(
            PaymentRecord.tenant_id == tenant_id,
            *record_where,
        )
    )).scalar() or 0

    # Total contract receivable (合同应收) — 回款率的标准分母，避免「计划回款」未录全导致回款率虚高
    # （Contract 已在模块顶部导入；这里原有的函数内 import 会让 Contract 变成局部名，
    #  导致上面算 contract_where 时报 UnboundLocalError，故删除）
    total_receivable = (await db.execute(
        select(func.coalesce(func.sum(Contract.amount_total), 0)).where(
            Contract.tenant_id == tenant_id,
            *contract_where,
        )
    )).scalar() or 0

    # Overdue plans
    overdue_count = (await db.execute(
        select(func.count(PaymentPlan.id)).where(
            PaymentPlan.tenant_id == tenant_id,
            PaymentPlan.status.in_(["pending", "overdue"]),
            PaymentPlan.due_date < now,
            *plan_where,
        )
    )).scalar() or 0

    overdue_amount = (await db.execute(
        select(func.coalesce(func.sum(PaymentPlan.amount), 0)).where(
            PaymentPlan.tenant_id == tenant_id,
            PaymentPlan.status.in_(["pending", "overdue"]),
            PaymentPlan.due_date < now,
            *plan_where,
        )
    )).scalar() or 0

    # Upcoming 30 days
    from datetime import timedelta
    upcoming_date = now + timedelta(days=30)
    upcoming_amount = (await db.execute(
        select(func.coalesce(func.sum(PaymentPlan.amount), 0)).where(
            PaymentPlan.tenant_id == tenant_id,
            PaymentPlan.status == "pending",
            PaymentPlan.due_date >= now,
            PaymentPlan.due_date <= upcoming_date,
            *plan_where,
        )
    )).scalar() or 0

    # 回款率 = 已回款 / 合同应收，封顶 100%（合同应收为 0 时回退到计划总额）
    denom = float(total_receivable) or float(total_planned)
    collection_rate = round(min(float(total_received) / denom * 100, 100.0), 1) if denom else 0
    return ok({
        "total_planned": float(total_planned),
        "total_receivable": float(total_receivable),
        "total_received": float(total_received),
        "overdue_count": overdue_count,
        "overdue_amount": float(overdue_amount),
        "upcoming_30d_amount": float(upcoming_amount),
        "collection_rate": collection_rate,
    })


@router.get("/milestone_overview")
async def milestone_overview(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Delivery milestone overview by status."""
    # Model uses: not_start/doing/done/delayed
    # Map to readable keys for frontend
    status_map = {
        "not_start": "pending",
        "doing": "in_progress",
        "done": "completed",
        "delayed": "delayed",
    }
    scope = await _scope_owner_ids(db, tenant_id, _user)
    milestone_where = _child_scope_where(DeliveryMilestone, tenant_id, _user, scope)
    result = {}
    for db_status, display_key in status_map.items():
        count = (await db.execute(
            select(func.count(DeliveryMilestone.id)).where(
                DeliveryMilestone.tenant_id == tenant_id,
                DeliveryMilestone.status == db_status,
                *milestone_where,
            )
        )).scalar() or 0
        result[display_key] = count
    total = sum(result.values())
    result["total"] = total
    result["completion_rate"] = round(result.get("completed", 0) / total * 100, 1) if total else 0
    return ok(result)


@router.get("/monthly_revenue")
async def monthly_revenue(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Monthly revenue (payment received) for the last 6 months."""
    now = datetime.now(timezone.utc)
    scope = await _scope_owner_ids(db, tenant_id, _user)
    record_where = _child_scope_where(PaymentRecord, tenant_id, _user, scope)
    months = []
    for i in range(5, -1, -1):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        amount = (await db.execute(
            select(func.coalesce(func.sum(PaymentRecord.amount), 0)).where(
                PaymentRecord.tenant_id == tenant_id,
                extract("year", PaymentRecord.received_date) == y,
                extract("month", PaymentRecord.received_date) == m,
                *record_where,
            )
        )).scalar() or 0
        months.append({
            "year": y, "month": m,
            "label": f"{y}-{str(m).zfill(2)}",
            "amount": float(amount),
        })
    return ok(months)


@router.get("/leaderboard")
async def leaderboard(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Performance leaderboard: top salespeople by won deals（签单数/成交额=已签合同）。"""
    # 排行榜逐行就是「同事姓名 + 成交额」，是本域泄露面最大的一个接口：
    # 按业绩归属人（credited_owner）落在数据范围内来卡——self 只剩自己一行，
    # dept 剩本部门子树，data_scope=all/管理员仍是全员榜。功能保留，只收窄行集。
    scope = await _scope_owner_ids(db, tenant_id, _user)

    # 签单数 + 成交额：已签合同，归到商机负责人；无关联商机的外部合同回退到合同负责人
    credited_owner_id = func.coalesce(OpportunityProject.owner_id, Contract.assignee_id)
    credited_owner_name = func.coalesce(OpportunityProject.owner_name, Contract.assignee_name)
    won_scope_where = [] if scope is None else [credited_owner_id.in_(scope)]
    won_rows = (await db.execute(
        select(
            credited_owner_id.label("owner_id"),
            credited_owner_name.label("owner_name"),
            func.count(Contract.id).label("won_count"),
            func.coalesce(func.sum(Contract.amount_total), 0).label("won_amount"),
        )
        .select_from(Contract)
        .outerjoin(
            OpportunityProject,
            and_(
                OpportunityProject.id == Contract.project_id,
                OpportunityProject.tenant_id == tenant_id,
            ),
        )
        .where(
            Contract.tenant_id == tenant_id,
            Contract.status == "signed",
            credited_owner_id.isnot(None),
            *won_scope_where,
        )
        .group_by(credited_owner_id, credited_owner_name)
    )).all()
    won_map = {r.owner_id: {"owner_name": r.owner_name, "won_count": r.won_count, "won_amount": float(r.won_amount)} for r in won_rows}

    # Active pipeline by owner
    pipeline_scope_where = [] if scope is None else [OpportunityProject.owner_id.in_(scope)]
    pipeline_result = await db.execute(
        select(
            OpportunityProject.owner_id,
            func.count(OpportunityProject.id).label("active_count"),
            func.coalesce(func.sum(OpportunityProject.amount_expect), 0).label("pipeline_amount"),
        ).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.status == "active",
            OpportunityProject.owner_id.isnot(None),
            *pipeline_scope_where,
        ).group_by(OpportunityProject.owner_id)
    )
    pipeline_map = {r.owner_id: {"active_count": r.active_count, "pipeline_amount": float(r.pipeline_amount)} for r in pipeline_result.all()}

    # 合并：按成交额排序取 Top10
    board = []
    for oid, w in won_map.items():
        p = pipeline_map.get(oid, {"active_count": 0, "pipeline_amount": 0})
        board.append({
            "owner_id": oid,
            "owner_name": w["owner_name"] or "未知",
            "won_count": w["won_count"],
            "won_amount": w["won_amount"],
            "active_count": p["active_count"],
            "pipeline_amount": p["pipeline_amount"],
        })
    board.sort(key=lambda x: x["won_amount"], reverse=True)
    return ok(board[:10])


@router.get("/trend")
async def trend(
    period: str = Query("month"),
    months: int = Query(6, ge=1, le=24),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Monthly trend: new opportunities, won, lost, and pipeline amount."""
    from datetime import timedelta
    now = datetime.now(timezone.utc)
    scope = await _scope_owner_ids(db, tenant_id, _user)
    proj_where = _project_scope_where(tenant_id, _user, scope)
    contract_where = _child_scope_where(Contract, tenant_id, _user, scope)
    result = []
    for i in range(months - 1, -1, -1):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1

        new_count = (await db.execute(
            select(func.count(OpportunityProject.id)).where(
                OpportunityProject.tenant_id == tenant_id,
                extract("year", OpportunityProject.created_at) == y,
                extract("month", OpportunityProject.created_at) == m,
                *proj_where,
            )
        )).scalar() or 0

        # 赢单数 = 当月已签合同数（按签约日 signed_date 归期）
        won_count = await _signed_contract_count(
            db, tenant_id,
            Contract.signed_date.isnot(None),
            extract("year", Contract.signed_date) == y,
            extract("month", Contract.signed_date) == m,
            *contract_where,
        )

        lost_count = (await db.execute(
            select(func.count(OpportunityProject.id)).where(
                OpportunityProject.tenant_id == tenant_id,
                OpportunityProject.status == "lost",
                extract("year", OpportunityProject.updated_at) == y,
                extract("month", OpportunityProject.updated_at) == m,
                *proj_where,
            )
        )).scalar() or 0

        # 赢单金额 = 当月已签合同额（按签约日 signed_date 归期）
        won_amount = await _signed_contract_amount(
            db, tenant_id,
            Contract.signed_date.isnot(None),
            extract("year", Contract.signed_date) == y,
            extract("month", Contract.signed_date) == m,
            *contract_where,
        )

        result.append({
            "label": f"{y}-{str(m).zfill(2)}",
            "year": y, "month": m,
            "new": new_count, "won": won_count, "lost": lost_count,
            "won_amount": float(won_amount),
        })
    return ok(result)


@router.get("/collection")
async def collection(
    months: int = Query(6, ge=1, le=24),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Monthly collection analysis: receivable, received, overdue."""
    now = datetime.now(timezone.utc)
    scope = await _scope_owner_ids(db, tenant_id, _user)
    plan_where = _child_scope_where(PaymentPlan, tenant_id, _user, scope)
    record_where = _child_scope_where(PaymentRecord, tenant_id, _user, scope)
    result = []
    for i in range(months - 1, -1, -1):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1

        # Planned (due in this month)
        receivable = (await db.execute(
            select(func.coalesce(func.sum(PaymentPlan.amount), 0)).where(
                PaymentPlan.tenant_id == tenant_id,
                extract("year", PaymentPlan.due_date) == y,
                extract("month", PaymentPlan.due_date) == m,
                *plan_where,
            )
        )).scalar() or 0

        # Received in this month
        received = (await db.execute(
            select(func.coalesce(func.sum(PaymentRecord.amount), 0)).where(
                PaymentRecord.tenant_id == tenant_id,
                extract("year", PaymentRecord.received_date) == y,
                extract("month", PaymentRecord.received_date) == m,
                *record_where,
            )
        )).scalar() or 0

        # Overdue plans in this month (due in this month, status overdue or pending past due)
        import calendar
        last_day = calendar.monthrange(y, m)[1]
        from datetime import date as date_type
        month_end = date_type(y, m, last_day)
        today = now.date()
        overdue = 0
        if month_end <= today:
            overdue = (await db.execute(
                select(func.coalesce(func.sum(PaymentPlan.amount), 0)).where(
                    PaymentPlan.tenant_id == tenant_id,
                    PaymentPlan.status.in_(["pending", "overdue"]),
                    extract("year", PaymentPlan.due_date) == y,
                    extract("month", PaymentPlan.due_date) == m,
                    *plan_where,
                )
            )).scalar() or 0

        result.append({
            "label": f"{y}-{str(m).zfill(2)}",
            "receivable": float(receivable),
            "received": float(received),
            "overdue": float(overdue),
        })
    return ok(result)


@router.get("/my_overview")
async def my_overview(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """Personal dashboard: my customers, projects, expiring contracts, pending items."""
    from datetime import timedelta
    uid = user["sub"]
    now = datetime.now(timezone.utc)

    # My customers
    my_customer_count = (await db.execute(
        select(func.count(Customer.id)).where(
            Customer.tenant_id == tenant_id, Customer.owner_id == uid,
        )
    )).scalar() or 0

    # My active projects
    my_active_projects = (await db.execute(
        select(func.count(OpportunityProject.id)).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.owner_id == uid,
            OpportunityProject.status == "active",
        )
    )).scalar() or 0

    # My pipeline value
    my_pipeline = (await db.execute(
        select(func.coalesce(func.sum(OpportunityProject.amount_expect), 0)).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.owner_id == uid,
            OpportunityProject.status == "active",
        )
    )).scalar() or 0

    # My won deals this month = 本月归到我的已签合同数（按签约日 signed_date）
    my_won_month = (await db.execute(
        select(func.count(Contract.id))
        .select_from(Contract)
        .outerjoin(OpportunityProject, and_(
            OpportunityProject.id == Contract.project_id,
            OpportunityProject.tenant_id == tenant_id,
        ))
        .where(
            Contract.tenant_id == tenant_id,
            Contract.status == "signed",
            func.coalesce(OpportunityProject.owner_id, Contract.assignee_id) == uid,
            extract("year", Contract.signed_date) == now.year,
            extract("month", Contract.signed_date) == now.month,
        )
    )).scalar() or 0

    # My pending leads
    my_pending_leads = (await db.execute(
        select(func.count(Lead.id)).where(
            Lead.tenant_id == tenant_id,
            Lead.owner_id == uid,
            Lead.status.in_(["new", "following"]),
        )
    )).scalar() or 0

    # My open tickets
    my_open_tickets = (await db.execute(
        select(func.count(ServiceTicket.id)).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.assigned_to_id == uid,
            ServiceTicket.status.in_(["open", "assigned", "in_progress"]),
        )
    )).scalar() or 0

    # Expiring contracts (next 30 days) — contracts linked to my projects
    from sqlalchemy import text as sa_text
    expiring_rows = (await db.execute(
        select(
            Contract.id, Contract.contract_no, Contract.amount_total,
            Contract.signed_date, OpportunityProject.name.label("project_name"),
        ).join(
            OpportunityProject, OpportunityProject.id == Contract.project_id
        ).where(
            Contract.tenant_id == tenant_id,
            OpportunityProject.owner_id == uid,
            Contract.status == "signed",
        ).order_by(Contract.signed_date.desc()).limit(5)
    )).all()
    expiring_contracts = [{
        "id": r.id, "contract_no": r.contract_no,
        "amount_total": float(r.amount_total) if r.amount_total else 0,
        "project_name": r.project_name,
        "signed_date": str(r.signed_date) if r.signed_date else None,
    } for r in expiring_rows]

    # My stalled projects (no update in 7+ days)
    stalled_rows = (await db.execute(
        select(
            OpportunityProject.id, OpportunityProject.name,
            OpportunityProject.stage_code, OpportunityProject.updated_at,
        ).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.owner_id == uid,
            OpportunityProject.status == "active",
            OpportunityProject.updated_at < func.now() - sa_text("interval '7 days'"),
        ).order_by(OpportunityProject.updated_at.asc()).limit(5)
    )).all()
    stalled_projects = [{
        "id": r.id, "name": r.name, "stage_code": r.stage_code,
        "days_stalled": (now - r.updated_at.replace(tzinfo=timezone.utc)).days if r.updated_at else 0,
    } for r in stalled_rows]

    return ok({
        "my_customer_count": my_customer_count,
        "my_active_projects": my_active_projects,
        "my_pipeline": float(my_pipeline),
        "my_won_month": my_won_month,
        "my_pending_leads": my_pending_leads,
        "my_open_tickets": my_open_tickets,
        "expiring_contracts": expiring_contracts,
        "stalled_projects": stalled_projects,
    })


@router.get("/search")
async def global_search(
    q: str = Query(..., min_length=1, max_length=100),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Global search across customers, leads, projects, contacts, tickets."""
    keyword = f"%{q}%"
    results = []
    LIMIT = 5

    # 全局搜索直接吐 id + 名称 + 详情 url，不按数据范围过滤等于给了一个枚举全租户的入口。
    scope = await _scope_owner_ids(db, tenant_id, _user)
    visible_cust_ids = await visible_customer_ids_select(db, tenant_id, _user)  # None=不限
    cust_where = [] if visible_cust_ids is None else [Customer.id.in_(visible_cust_ids)]

    # Customers
    rows = (await db.execute(
        select(Customer.id, Customer.name, Customer.customer_code).where(
            Customer.tenant_id == tenant_id,
            or_(Customer.name.ilike(keyword), Customer.customer_code.ilike(keyword), Customer.short_name.ilike(keyword)),
            *cust_where,
        ).limit(LIMIT)
    )).all()
    for r in rows:
        results.append({"type": "customer", "id": r.id, "title": r.name, "subtitle": r.customer_code or "", "url": f"/customers/{r.id}"})

    # Leads
    rows = (await db.execute(
        select(Lead.id, Lead.lead_code, Lead.title, Lead.company_name, Lead.contact_name).where(
            Lead.tenant_id == tenant_id,
            or_(Lead.title.ilike(keyword), Lead.company_name.ilike(keyword), Lead.contact_name.ilike(keyword), Lead.lead_code.ilike(keyword)),
            *_lead_scope_where(tenant_id, _user, scope),
        ).limit(LIMIT)
    )).all()
    for r in rows:
        results.append({"type": "lead", "id": r.id, "title": r.title or r.company_name or r.contact_name, "subtitle": r.lead_code or "", "url": f"/leads/{r.id}"})

    # Projects
    rows = (await db.execute(
        select(OpportunityProject.id, OpportunityProject.name, OpportunityProject.project_code).where(
            OpportunityProject.tenant_id == tenant_id,
            or_(OpportunityProject.name.ilike(keyword), OpportunityProject.project_code.ilike(keyword)),
            *_project_scope_where(tenant_id, _user, scope),
        ).limit(LIMIT)
    )).all()
    for r in rows:
        results.append({"type": "project", "id": r.id, "title": r.name, "subtitle": r.project_code or "", "url": f"/opportunities/{r.id}"})

    # Contacts —— 联系人没有 owner_id，可见性跟着父客户走（与联系人列表同口径）
    from app.domains.customer.models import Contact
    contact_where = [] if visible_cust_ids is None else [Contact.customer_id.in_(visible_cust_ids)]
    rows = (await db.execute(
        select(Contact.id, Contact.name, Contact.phone, Contact.customer_id).where(
            Contact.tenant_id == tenant_id,
            or_(Contact.name.ilike(keyword), Contact.phone.ilike(keyword), Contact.email.ilike(keyword)),
            *contact_where,
        ).limit(LIMIT)
    )).all()
    for r in rows:
        results.append({"type": "contact", "id": r.id, "title": r.name, "subtitle": r.phone or "", "url": f"/customers/{r.customer_id}"})

    # Service Tickets —— 工单无归属维度（服务台全租户可见），保持原口径
    rows = (await db.execute(
        select(ServiceTicket.id, ServiceTicket.ticket_no, ServiceTicket.description).where(
            ServiceTicket.tenant_id == tenant_id,
            or_(ServiceTicket.ticket_no.ilike(keyword), ServiceTicket.description.ilike(keyword)),
        ).limit(LIMIT)
    )).all()
    for r in rows:
        results.append({"type": "ticket", "id": r.id, "title": r.ticket_no, "subtitle": (r.description or "")[:50], "url": f"/service-tickets/{r.id}"})

    return ok(results)


# ---- Sales Targets ----

class SalesTargetBody(BaseModel):
    user_id: Optional[str] = None
    user_name: Optional[str] = None
    department_id: Optional[str] = None
    department_name: Optional[str] = None
    year: int = Field(..., ge=2020, le=2100)
    month: int = Field(..., ge=1, le=12)
    target_amount: float = Field(..., ge=0)
    target_count: Optional[int] = Field(None, ge=0)


@router.get("/targets")
async def list_targets(
    year: int = Query(...),
    month: int = Query(None),
    user_id: str = Query(None),
    department_id: str = Query(None),
    target_type: str = Query(None),  # "user" or "department"
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    q = select(SalesTarget).where(SalesTarget.tenant_id == tenant_id, SalesTarget.year == year)
    if month:
        q = q.where(SalesTarget.month == month)
    if user_id:
        q = q.where(SalesTarget.user_id == user_id)
    if department_id:
        q = q.where(SalesTarget.department_id == department_id)
    if target_type == "user":
        q = q.where(SalesTarget.user_id.isnot(None), SalesTarget.department_id.is_(None))
    elif target_type == "department":
        q = q.where(SalesTarget.department_id.isnot(None))

    # 销售目标逐行是「同事姓名 + 目标金额」，与业绩排行同级敏感：
    # 个人目标按 user_id 落在数据范围内；部门目标按「该部门至少有一名成员在范围内」，
    # 这样 dept 档能看到自己部门子树的目标，self 档只剩自己。管理员 / all 不受影响。
    scope = await _scope_owner_ids(db, tenant_id, _user)
    if scope is not None:
        from app.domains.organization.models import UserDepartment
        scoped_dept_ids = select(UserDepartment.department_id).where(
            UserDepartment.tenant_id == tenant_id,
            UserDepartment.user_id.in_(scope),
        )
        q = q.where(or_(
            SalesTarget.user_id.in_(scope),
            SalesTarget.department_id.in_(scoped_dept_ids),
        ))
    items = (await db.execute(q.order_by(SalesTarget.month, SalesTarget.user_name))).scalars().all()
    return ok([{
        "id": t.id, "user_id": t.user_id, "user_name": t.user_name,
        "department_id": t.department_id, "department_name": t.department_name,
        "year": t.year, "month": t.month,
        "target_amount": float(t.target_amount), "target_count": t.target_count,
    } for t in items])


@router.post("/targets")
async def upsert_target(
    body: SalesTargetBody,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("role:manage")),
):
    """Create or update a sales target (upsert by user+year+month or department+year+month)."""
    from app.common.exceptions import BusinessException
    if body.user_id and body.department_id:
        raise BusinessException(message="不能同时设置用户和部门")
    if not body.user_id and not body.department_id:
        raise BusinessException(message="必须指定用户或部门")

    if body.department_id:
        # Department-level target
        existing = (await db.execute(
            select(SalesTarget).where(
                SalesTarget.tenant_id == tenant_id,
                SalesTarget.department_id == body.department_id,
                SalesTarget.year == body.year,
                SalesTarget.month == body.month,
            )
        )).scalar()
    else:
        # User-level target
        existing = (await db.execute(
            select(SalesTarget).where(
                SalesTarget.tenant_id == tenant_id,
                SalesTarget.user_id == body.user_id,
                SalesTarget.year == body.year,
                SalesTarget.month == body.month,
            )
        )).scalar()

    if existing:
        existing.target_amount = body.target_amount
        existing.target_count = body.target_count
        if body.user_id:
            existing.user_name = body.user_name or existing.user_name
        if body.department_id:
            existing.department_name = body.department_name or existing.department_name
        await db.commit()
        await db.refresh(existing)
        t = existing
    else:
        t = SalesTarget(
            tenant_id=tenant_id,
            user_id=body.user_id, user_name=body.user_name,
            department_id=body.department_id, department_name=body.department_name,
            year=body.year, month=body.month,
            target_amount=body.target_amount, target_count=body.target_count,
        )
        db.add(t)
        await db.commit()
        await db.refresh(t)
    return ok({
        "id": t.id, "user_id": t.user_id, "department_id": t.department_id,
        "year": t.year, "month": t.month, "target_amount": float(t.target_amount),
    })


@router.delete("/targets/{target_id}")
async def delete_target(
    target_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("role:manage")),
):
    t = (await db.execute(
        select(SalesTarget).where(SalesTarget.tenant_id == tenant_id, SalesTarget.id == target_id)
    )).scalar()
    if not t:
        from app.common.exceptions import BusinessException
        raise BusinessException(message="目标不存在")
    await db.delete(t)
    await db.commit()
    return ok()


# ---- Sales Target Import (with name / department auto-matching) ----

TARGET_IMPORT_HEADERS = ["姓名", "部门", "年", "月", "目标金额", "目标单数"]

# Logical field -> accepted header names (order-independent matching)
_TARGET_HEADER_ALIASES = {
    "user_name": ["姓名", "销售人员", "员工", "员工姓名", "用户", "人员"],
    "department_name": ["部门", "部门名", "部门名称"],
    "year": ["年", "年份", "目标年份"],
    "month": ["月", "月份", "目标月份"],
    "target_amount": ["目标金额", "金额", "销售目标", "目标"],
    "target_count": ["目标单数", "目标数量", "目标单量", "单数"],
}


def _read_sheet_rows(content: bytes, filename: str) -> list:
    """Parse an uploaded .xlsx/.xls/.csv into a list of row tuples (incl. header)."""
    fname = (filename or "").lower()
    if fname.endswith(".csv"):
        import csv as csv_mod
        text = content.decode("utf-8-sig", errors="replace")
        return [tuple(r) for r in csv_mod.reader(text.splitlines())]
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return all_rows


def _build_target_header_map(header_row) -> dict:
    """Map logical field -> column index by matching header names."""
    norm = [str(c).strip() if c is not None else "" for c in header_row]
    field_idx: dict[str, int] = {}
    for field, aliases in _TARGET_HEADER_ALIASES.items():
        for idx, h in enumerate(norm):
            if h in aliases:
                field_idx[field] = idx
                break
    return field_idx


def _missing_required_headers(fmap: dict) -> list:
    missing = [f for f in ("year", "month", "target_amount") if f not in fmap]
    if "user_name" not in fmap and "department_name" not in fmap:
        missing.append("user_name_or_department")
    return missing


def _required_headers_error_msg(missing: list) -> str:
    label = {"year": "年", "month": "月", "target_amount": "目标金额",
             "user_name_or_department": "姓名或部门"}
    return "缺少必要列：" + "、".join(label.get(m, m) for m in missing)


async def _load_target_resolvers(db: AsyncSession, tenant_id: str):
    """Build lookup maps for a tenant: 姓名/用户名→用户、部门名→部门、用户→所属部门名集合。"""
    from app.domains.auth.models import User
    from app.domains.organization.models import Department, UserDepartment

    users = (await db.execute(select(User).where(User.tenant_id == tenant_id))).scalars().all()
    real_name_map: dict[str, list] = {}
    username_map: dict[str, object] = {}
    for u in users:
        if u.real_name:
            real_name_map.setdefault(u.real_name.strip(), []).append(u)
        if u.username:
            username_map[u.username.strip()] = u

    depts = (await db.execute(select(Department).where(Department.tenant_id == tenant_id))).scalars().all()
    dept_name_map: dict[str, list] = {}
    for d in depts:
        if d.name:
            dept_name_map.setdefault(d.name.strip(), []).append(d)

    # user_id -> {所属部门名}, 用于同名用户的部门消歧
    ud_rows = (await db.execute(
        select(UserDepartment.user_id, Department.name)
        .join(Department, Department.id == UserDepartment.department_id)
        .where(UserDepartment.tenant_id == tenant_id)
    )).all()
    user_dept_map: dict[str, set] = {}
    for uid, dname in ud_rows:
        if dname:
            user_dept_map.setdefault(uid, set()).add(dname.strip())

    return real_name_map, username_map, dept_name_map, user_dept_map


def _resolve_target_row(cells, fmap, real_name_map, username_map, dept_name_map, user_dept_map):
    """Resolve a data row to a target dict. Returns (parsed_dict, error_message).

    填了姓名 → 个人目标（"部门"列选填，仅在同名时用于区分该用户所属部门）；
    仅填部门 → 部门目标。
    """
    def cell(field: str) -> str:
        i = fmap.get(field)
        if i is None or i >= len(cells):
            return ""
        v = cells[i]
        return str(v).strip() if v is not None else ""

    name_v = cell("user_name")
    dept_v = cell("department_name")
    if not name_v and not dept_v:
        return None, "姓名和部门至少填写一个"

    year_v, month_v = cell("year"), cell("month")
    amt_v, cnt_v = cell("target_amount"), cell("target_count")
    try:
        year = int(float(year_v))
    except (ValueError, TypeError):
        return None, f"年份无效: {year_v or '空'}"
    if not (2020 <= year <= 2100):
        return None, f"年份超出范围: {year}"
    try:
        month = int(float(month_v))
    except (ValueError, TypeError):
        return None, f"月份无效: {month_v or '空'}"
    if not (1 <= month <= 12):
        return None, f"月份超出范围: {month}"
    try:
        amount = float(str(amt_v).replace(",", "").replace("，", "")) if amt_v else 0.0
    except (ValueError, TypeError):
        return None, f"目标金额无效: {amt_v}"
    if amount < 0:
        return None, "目标金额不能为负"
    count = None
    if cnt_v:
        try:
            count = int(float(cnt_v))
        except (ValueError, TypeError):
            return None, f"目标单数无效: {cnt_v}"

    parsed = {
        "year": year, "month": month, "target_amount": amount, "target_count": count,
        "user_id": None, "user_name": None, "department_id": None, "department_name": None,
    }

    if name_v:
        # 个人目标：按姓名匹配用户；"部门"列（如填）仅用于同名用户的消歧
        matches = list(real_name_map.get(name_v) or [])
        if len(matches) > 1 and dept_v:
            narrowed = [u for u in matches if dept_v in user_dept_map.get(u.id, ())]
            if len(narrowed) == 1:
                matches = narrowed
        if len(matches) == 1:
            u = matches[0]
        elif len(matches) > 1:
            # 姓名重复 → 尝试用唯一的用户名消歧
            if name_v in username_map:
                u = username_map[name_v]
            else:
                hint = "，可在“部门”列填写其所属部门以区分" if not dept_v else "，“部门”列仍无法唯一确定"
                return None, f"姓名“{name_v}”匹配到{len(matches)}个用户{hint}，或改用唯一用户名"
        elif name_v in username_map:
            u = username_map[name_v]
        else:
            return None, f"未找到姓名/用户名“{name_v}”对应的用户"
        parsed["user_id"] = u.id
        parsed["user_name"] = u.real_name or u.username
    else:
        matches = dept_name_map.get(dept_v)
        if not matches:
            return None, f"未找到部门“{dept_v}”"
        if len(matches) > 1:
            return None, f"部门“{dept_v}”重复（{len(matches)}个），请确保部门名唯一"
        d = matches[0]
        parsed["department_id"] = d.id
        parsed["department_name"] = d.name

    return parsed, None


@router.get("/targets/import/template")
async def targets_import_template(_user=Depends(get_current_user)):
    """Download the sales-target import template (.xlsx)."""
    buf = build_template(
        "销售目标导入",
        TARGET_IMPORT_HEADERS,
        sample_rows=[
            ["张三", "", 2026, 1, 500000, 5],            # 个人目标
            ["李四", "华东销售部", 2026, 1, 300000, 3],   # 个人目标（部门列仅用于同名区分）
            ["", "华东销售部", 2026, 1, 2000000, 20],     # 部门目标（仅填部门）
        ],
    )
    return excel_response(buf, "sales_targets_template.xlsx")


@router.post("/targets/import/preview")
async def targets_import_preview(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("role:manage")),
):
    """Parse + validate an import file, resolving 姓名/部门 to ids. Rows that fail to
    match are returned in `errors` (keyed by row index) so they show up in the preview.

    权限与真正的导入(/targets/import/excel)对齐：预览会把 姓名→用户、部门名→部门 在全租户
    范围内解析出来，只挂 get_current_user 的话，任何员工都能拿它当通讯录/组织架构探针。
    """
    from app.common.exceptions import BusinessException

    content = await file.read()
    all_rows = _read_sheet_rows(content, file.filename or "")
    if not all_rows:
        return ok({"headers": [], "rows": [], "duplicates": [], "errors": {}})

    fmap = _build_target_header_map(all_rows[0])
    missing = _missing_required_headers(fmap)
    if missing:
        raise BusinessException(message=_required_headers_error_msg(missing))

    headers = [str(c).strip() if c is not None else f"列{i+1}" for i, c in enumerate(all_rows[0])]
    data_rows = []
    for row in all_rows[1:]:
        if not row or not any(c not in (None, "") for c in row):
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        while len(cells) < len(headers):
            cells.append("")
        data_rows.append(cells[:len(headers)])

    real_name_map, username_map, dept_name_map, user_dept_map = await _load_target_resolvers(db, tenant_id)
    errors: dict[int, str] = {}
    for i, cells in enumerate(data_rows):
        _, err = _resolve_target_row(cells, fmap, real_name_map, username_map, dept_name_map, user_dept_map)
        if err:
            errors[i] = err

    # Upsert semantics → no rows are skipped as "duplicates"; existing targets are updated.
    return ok({"headers": headers, "rows": data_rows, "duplicates": [], "errors": errors})


@router.post("/targets/import/excel")
async def targets_import_excel(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("role:manage")),
):
    """Import sales targets from Excel/CSV, auto-matching 姓名→用户 and 部门名→部门, upserting per row.
    Rows that fail name/department resolution are skipped and reported in `errors`."""
    from app.common.exceptions import BusinessException

    content = await file.read()
    all_rows = _read_sheet_rows(content, file.filename or "")
    if len(all_rows) < 2:
        return ok({"created": 0, "skipped": 0, "errors": []})

    fmap = _build_target_header_map(all_rows[0])
    missing = _missing_required_headers(fmap)
    if missing:
        raise BusinessException(message=_required_headers_error_msg(missing))

    real_name_map, username_map, dept_name_map, user_dept_map = await _load_target_resolvers(db, tenant_id)

    created = 0
    skipped = 0
    errors: list[str] = []
    for idx, row in enumerate(all_rows[1:], start=2):
        if not row or not any(c not in (None, "") for c in row):
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        parsed, err = _resolve_target_row(cells, fmap, real_name_map, username_map, dept_name_map, user_dept_map)
        if err:
            errors.append(f"第{idx}行: {err}")
            skipped += 1
            continue
        try:
            if parsed["department_id"]:
                existing = (await db.execute(select(SalesTarget).where(
                    SalesTarget.tenant_id == tenant_id,
                    SalesTarget.department_id == parsed["department_id"],
                    SalesTarget.year == parsed["year"], SalesTarget.month == parsed["month"],
                ))).scalar()
            else:
                existing = (await db.execute(select(SalesTarget).where(
                    SalesTarget.tenant_id == tenant_id,
                    SalesTarget.user_id == parsed["user_id"],
                    SalesTarget.year == parsed["year"], SalesTarget.month == parsed["month"],
                ))).scalar()
            if existing:
                existing.target_amount = parsed["target_amount"]
                existing.target_count = parsed["target_count"]
                if parsed["user_id"]:
                    existing.user_name = parsed["user_name"]
                if parsed["department_id"]:
                    existing.department_name = parsed["department_name"]
            else:
                db.add(SalesTarget(
                    tenant_id=tenant_id,
                    user_id=parsed["user_id"], user_name=parsed["user_name"],
                    department_id=parsed["department_id"], department_name=parsed["department_name"],
                    year=parsed["year"], month=parsed["month"],
                    target_amount=parsed["target_amount"], target_count=parsed["target_count"],
                ))
            await db.commit()
            created += 1
        except Exception as e:  # noqa: BLE001
            await db.rollback()
            errors.append(f"第{idx}行: {str(e)[:80]}")
            skipped += 1

    return ok({"created": created, "skipped": skipped, "errors": errors})


@router.get("/target_achievement")
async def target_achievement(
    year: int = Query(...),
    month: int = Query(None),
    target_type: str = Query(None),  # "user" or "department"
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Get target vs actual achievement for users and/or departments."""
    from app.domains.organization.models import UserDepartment

    # 达成率表 = 目标 + 实际签单额 + 人名/部门名，与 /targets、/leaderboard 同一份敏感数据，
    # 三处必须同口径收窄，否则从任意一处都能把全员业绩拼回来。
    scope = await _scope_owner_ids(db, tenant_id, _user)

    # Get targets
    tq = select(SalesTarget).where(SalesTarget.tenant_id == tenant_id, SalesTarget.year == year)
    if month:
        tq = tq.where(SalesTarget.month == month)
    if scope is not None:
        scoped_dept_ids = select(UserDepartment.department_id).where(
            UserDepartment.tenant_id == tenant_id,
            UserDepartment.user_id.in_(scope),
        )
        tq = tq.where(or_(
            SalesTarget.user_id.in_(scope),
            SalesTarget.department_id.in_(scoped_dept_ids),
        ))
    targets = (await db.execute(tq)).scalars().all()

    # 实际业绩 = 已签合同金额（签单额口径）。
    # 金额取 contracts.amount_total（status='signed'），归期按签约日 signed_date；
    # 业绩归到该合同所属商机的负责人 owner_id，外部导入的无商机合同回退到合同负责人 assignee_id。
    credited_owner_id = func.coalesce(OpportunityProject.owner_id, Contract.assignee_id)
    credited_owner_name = func.coalesce(OpportunityProject.owner_name, Contract.assignee_name)
    aq = (
        select(
            credited_owner_id.label("owner_id"),
            credited_owner_name.label("owner_name"),
            func.coalesce(func.sum(Contract.amount_total), 0).label("actual_amount"),
            func.count(Contract.id).label("actual_count"),
        )
        .select_from(Contract)
        .outerjoin(
            OpportunityProject,
            and_(
                OpportunityProject.id == Contract.project_id,
                OpportunityProject.tenant_id == tenant_id,
            ),
        )
        .where(
            Contract.tenant_id == tenant_id,
            Contract.status == "signed",
            Contract.signed_date.isnot(None),
            credited_owner_id.isnot(None),
            extract("year", Contract.signed_date) == year,
            *([] if scope is None else [credited_owner_id.in_(scope)]),
        )
    )
    if month:
        aq = aq.where(extract("month", Contract.signed_date) == month)
    aq = aq.group_by(credited_owner_id, credited_owner_name)
    actuals = {r.owner_id: {"actual_amount": float(r.actual_amount), "actual_count": r.actual_count, "owner_name": r.owner_name}
               for r in (await db.execute(aq)).all()}

    result = []

    # ---- User-level achievements ----
    if target_type != "department":
        user_targets: dict = {}
        for t in targets:
            if not t.user_id or t.department_id:
                continue
            key = t.user_id
            if key not in user_targets:
                user_targets[key] = {"user_id": t.user_id, "user_name": t.user_name, "target_amount": 0, "target_count": 0}
            user_targets[key]["target_amount"] += float(t.target_amount)
            user_targets[key]["target_count"] += (t.target_count or 0)

        all_users = set(list(user_targets.keys()) + list(actuals.keys()))
        for uid in all_users:
            t_data = user_targets.get(uid, {"user_id": uid, "user_name": actuals.get(uid, {}).get("owner_name", ""), "target_amount": 0, "target_count": 0})
            a = actuals.get(uid, {"actual_amount": 0, "actual_count": 0})
            target_amt = t_data["target_amount"]
            actual_amt = a["actual_amount"]
            rate = round(actual_amt / target_amt * 100, 1) if target_amt > 0 else 0
            result.append({
                "type": "user",
                "user_id": uid,
                "user_name": t_data["user_name"],
                "target_amount": target_amt,
                "target_count": t_data["target_count"],
                "actual_amount": actual_amt,
                "actual_count": a["actual_count"],
                "achievement_rate": rate,
            })

    # ---- Department-level achievements ----
    if target_type != "user":
        dept_targets: dict = {}
        for t in targets:
            if not t.department_id:
                continue
            key = t.department_id
            if key not in dept_targets:
                dept_targets[key] = {"department_id": t.department_id, "department_name": t.department_name, "target_amount": 0, "target_count": 0}
            dept_targets[key]["target_amount"] += float(t.target_amount)
            dept_targets[key]["target_count"] += (t.target_count or 0)

        if dept_targets:
            # Find users in each department via UserDepartment junction
            dept_ids = list(dept_targets.keys())
            ud_rows = (await db.execute(
                select(UserDepartment.department_id, UserDepartment.user_id).where(
                    UserDepartment.tenant_id == tenant_id,
                    UserDepartment.department_id.in_(dept_ids),
                )
            )).all()
            dept_user_map: dict[str, list[str]] = {}
            for row in ud_rows:
                dept_user_map.setdefault(row.department_id, []).append(row.user_id)

            for did, dt in dept_targets.items():
                user_ids = dept_user_map.get(did, [])
                actual_amt = sum(actuals.get(uid, {}).get("actual_amount", 0) for uid in user_ids)
                actual_cnt = sum(actuals.get(uid, {}).get("actual_count", 0) for uid in user_ids)
                target_amt = dt["target_amount"]
                rate = round(actual_amt / target_amt * 100, 1) if target_amt > 0 else 0
                result.append({
                    "type": "department",
                    "department_id": did,
                    "department_name": dt["department_name"],
                    "target_amount": target_amt,
                    "target_count": dt["target_count"],
                    "actual_amount": actual_amt,
                    "actual_count": actual_cnt,
                    "achievement_rate": rate,
                })

    result.sort(key=lambda x: x["achievement_rate"], reverse=True)
    return ok(result)


@router.get("/customer_region_stats")
async def customer_region_stats(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Customer distribution by region. 优先按结构化省份分组，回退 legacy region 文本。"""
    region_expr = func.coalesce(Customer.province, Customer.region)
    cust_where = await _customer_scope_where(db, tenant_id, _user)
    rows = (await db.execute(
        select(
            region_expr.label("region"),
            func.count(Customer.id).label("count"),
        ).where(
            Customer.tenant_id == tenant_id,
            Customer.is_deleted == False,
            region_expr.isnot(None),
            region_expr != "",
            *cust_where,
        ).group_by(region_expr)
        .order_by(func.count(Customer.id).desc())
    )).all()

    return ok([{"region": r.region, "count": r.count} for r in rows])


@router.get("/calendar_events")
async def calendar_events(
    year: int = Query(...),
    month: int = Query(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """Aggregate events for a given month: follow-ups, payment dues, contract expiry, milestones."""
    from datetime import date as date_type, timedelta
    import calendar as cal_mod
    uid = user["sub"]
    first_day = date_type(year, month, 1)
    last_day = date_type(year, month, cal_mod.monthrange(year, month)[1])

    # 日历里回款/合同/里程碑三类原本是全租户的，跟进计划才按本人过滤——
    # 结果是「我的日历」上排满了别人的回款金额和合同号。三类统一按数据范围收窄。
    scope = await _scope_owner_ids(db, tenant_id, user)
    plan_where = _child_scope_where(PaymentPlan, tenant_id, user, scope)
    contract_where = _child_scope_where(Contract, tenant_id, user, scope)
    milestone_where = _child_scope_where(DeliveryMilestone, tenant_id, user, scope)

    events = []

    # Follow-up activities with next_follow_date
    follow_rows = (await db.execute(
        select(Activity.id, Activity.subject, Activity.next_follow_date).where(
            Activity.tenant_id == tenant_id,
            Activity.created_by_id == uid,
            Activity.next_follow_date.isnot(None),
            Activity.next_follow_date >= first_day,
            Activity.next_follow_date <= last_day,
        )
    )).all()
    for r in follow_rows:
        events.append({
            "id": r.id, "date": str(r.next_follow_date),
            "type": "follow_up", "title": r.subject or "跟进计划",
            "color": "#3b82f6",
        })

    # Payment plans due
    pay_rows = (await db.execute(
        select(PaymentPlan.id, PaymentPlan.due_date, PaymentPlan.amount, PaymentPlan.status).where(
            PaymentPlan.tenant_id == tenant_id,
            PaymentPlan.due_date >= first_day,
            PaymentPlan.due_date <= last_day,
            *plan_where,
        )
    )).all()
    for r in pay_rows:
        events.append({
            "id": r.id, "date": str(r.due_date),
            "type": "payment_due",
            "title": f"回款¥{float(r.amount)/10000:.1f}万",
            "color": "#f59e0b" if r.status == "pending" else "#ef4444" if r.status == "overdue" else "#10b981",
        })

    # Contract expiry
    contract_rows = (await db.execute(
        select(Contract.id, Contract.contract_no, Contract.end_date).where(
            Contract.tenant_id == tenant_id,
            Contract.status == "signed",
            Contract.end_date.isnot(None),
            Contract.end_date >= first_day,
            Contract.end_date <= last_day,
            *contract_where,
        )
    )).all()
    for r in contract_rows:
        events.append({
            "id": r.id, "date": str(r.end_date),
            "type": "contract_expiry",
            "title": f"合同到期 {r.contract_no}",
            "color": "#ef4444",
        })

    # Delivery milestones
    milestone_rows = (await db.execute(
        select(DeliveryMilestone.id, DeliveryMilestone.name, DeliveryMilestone.plan_date).where(
            DeliveryMilestone.tenant_id == tenant_id,
            DeliveryMilestone.plan_date.isnot(None),
            DeliveryMilestone.plan_date >= first_day,
            DeliveryMilestone.plan_date <= last_day,
            *milestone_where,
        )
    )).all()
    for r in milestone_rows:
        events.append({
            "id": r.id, "date": str(r.plan_date),
            "type": "milestone",
            "title": r.name or "里程碑",
            "color": "#8b5cf6",
        })

    return ok(events)


@router.get("/contract_expiry")
async def contract_expiry(
    days: int = Query(90, ge=1, le=365),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Contracts expiring within N days."""
    from datetime import timedelta
    now = datetime.now(timezone.utc).date()
    cutoff = now + timedelta(days=days)

    # 到期预警逐行带 contract_no / 金额 / 客户（商机）名 / 负责人姓名，
    # 不按范围过滤等于把全租户合同台账贴到工作台上。
    scope = await _scope_owner_ids(db, tenant_id, _user)
    contract_where = _child_scope_where(Contract, tenant_id, _user, scope)

    rows = (await db.execute(
        select(
            Contract.id, Contract.contract_no, Contract.project_id, Contract.status,
            Contract.signed_date, Contract.end_date, Contract.amount_total,
            OpportunityProject.name.label("project_name"),
            OpportunityProject.owner_name,
        ).join(
            OpportunityProject, OpportunityProject.id == Contract.project_id
        ).where(
            Contract.tenant_id == tenant_id,
            Contract.status == "signed",
            Contract.end_date.isnot(None),
            Contract.end_date <= cutoff,
            *contract_where,
        ).order_by(Contract.end_date.asc())
        .limit(50)
    )).all()

    items = []
    for r in rows:
        days_left = (r.end_date - now).days if r.end_date else 0
        items.append({
            "id": r.id,
            "contract_no": r.contract_no,
            "project_id": r.project_id,
            "project_name": r.project_name,
            "owner_name": r.owner_name,
            "amount_total": float(r.amount_total) if r.amount_total else 0,
            "signed_date": str(r.signed_date) if r.signed_date else None,
            "end_date": str(r.end_date) if r.end_date else None,
            "days_left": days_left,
            "urgency": "expired" if days_left < 0 else "critical" if days_left <= 7 else "warning" if days_left <= 30 else "normal",
        })
    return ok(items)


# ---------- Export helpers ----------

async def _gather_export_data(db: AsyncSession, tenant_id: str, user: dict, start_date: Optional[str], end_date: Optional[str]):
    """Gather analytics summary data for export.

    报表中心的 Excel/PDF 导出走的就是这一份聚合，所以数据范围必须在这里就卡住——
    否则受限用户在页面上看到的是收窄后的数字，一点「导出」又把全租户口径落到文件里。
    """
    # created_at 是 timestamptz，asyncpg 下必须与 datetime 比较（与字符串比较会报 timestamptz>=text 类型错误）
    scope = await _scope_owner_ids(db, tenant_id, user)
    proj_where = _project_scope_where(tenant_id, user, scope)
    contract_where = _child_scope_where(Contract, tenant_id, user, scope)
    cust_where = await _customer_scope_where(db, tenant_id, user)

    filters = [OpportunityProject.tenant_id == tenant_id, *proj_where]
    if start_date:
        filters.append(OpportunityProject.created_at >= datetime.fromisoformat(start_date[:10]).replace(tzinfo=timezone.utc))
    if end_date:
        filters.append(OpportunityProject.created_at <= datetime.fromisoformat(end_date[:10] + "T23:59:59").replace(tzinfo=timezone.utc))

    # Funnel
    stage_map = {"S1": "线索确认", "S2": "需求分析", "S3": "方案报价", "S4": "商务谈判", "S5": "合同签订", "S6": "交付验收"}
    funnel_rows = (await db.execute(
        select(OpportunityProject.stage_code.label("stage"), func.count(OpportunityProject.id).label("cnt"),
               func.coalesce(func.sum(OpportunityProject.amount_expect), 0).label("amt"))
        .where(*filters).group_by(OpportunityProject.stage_code)
    )).all()
    funnel = [{"stage": r.stage, "label": stage_map.get(r.stage, r.stage), "count": r.cnt, "amount": float(r.amt)} for r in funnel_rows]

    # Win/Loss
    won_count = (await db.execute(select(func.count(OpportunityProject.id)).where(*filters, OpportunityProject.status == "won"))).scalar() or 0
    lost_count = (await db.execute(select(func.count(OpportunityProject.id)).where(*filters, OpportunityProject.status == "lost"))).scalar() or 0
    # 赢单金额 = 已签合同额（按签约日 signed_date 落在导出区间内）
    # signed_date 是 Date 列，asyncpg 下必须与 date 对象比较（与字符串比较会报 date>=text 类型错误）
    signed_filters = [Contract.signed_date.isnot(None), *contract_where]
    if start_date:
        signed_filters.append(Contract.signed_date >= date.fromisoformat(start_date[:10]))
    if end_date:
        signed_filters.append(Contract.signed_date <= date.fromisoformat(end_date[:10]))
    won_amount = await _signed_contract_amount(db, tenant_id, *signed_filters)
    lost_amount = float((await db.execute(select(func.coalesce(func.sum(OpportunityProject.amount_expect), 0)).where(*filters, OpportunityProject.status == "lost"))).scalar() or 0)
    total_decided = won_count + lost_count
    win_rate = round(won_count / total_decided * 100, 1) if total_decided else 0

    # Top customers by 成交额（已签合同额；客户取合同直连客户，回退到关联商机客户）
    cust_id_expr = func.coalesce(Contract.customer_id, OpportunityProject.customer_id)
    top_rows = (await db.execute(
        select(Customer.name, func.count(Contract.id).label("cnt"),
               func.coalesce(func.sum(Contract.amount_total), 0).label("amt"))
        .select_from(Contract)
        .outerjoin(OpportunityProject, and_(
            OpportunityProject.id == Contract.project_id,
            OpportunityProject.tenant_id == tenant_id,
        ))
        .join(Customer, and_(Customer.id == cust_id_expr, Customer.tenant_id == tenant_id))
        .where(Contract.tenant_id == tenant_id, Contract.status == "signed", *signed_filters, *cust_where)
        .group_by(Customer.name).order_by(func.sum(Contract.amount_total).desc()).limit(10)
    )).all()
    top_customers = [{"name": r.name, "project_count": r.cnt, "total_amount": float(r.amt)} for r in top_rows]

    # Region（优先结构化省份，回退 legacy region 文本）
    region_expr = func.coalesce(Customer.province, Customer.region)
    region_rows = (await db.execute(
        select(region_expr.label("region"), func.count(Customer.id).label("count"))
        .where(Customer.tenant_id == tenant_id, Customer.is_deleted == False,
               region_expr.isnot(None), region_expr != "", *cust_where)
        .group_by(region_expr).order_by(func.count(Customer.id).desc())
    )).all()
    regions = [{"region": r.region, "count": r.count} for r in region_rows]

    return {
        "funnel": funnel,
        "won_count": won_count, "lost_count": lost_count,
        "won_amount": won_amount, "lost_amount": lost_amount, "win_rate": win_rate,
        "top_customers": top_customers,
        "regions": regions,
    }


@router.get("/export/excel")
async def export_excel(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    data = await _gather_export_data(db, tenant_id, _user, start_date, end_date)
    wb = Workbook()

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
    )

    def style_header(ws, cols):
        for col_idx, title in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # Sheet 1: Funnel
    ws1 = wb.active
    ws1.title = "销售漏斗"
    style_header(ws1, ["阶段", "名称", "商机数", "金额"])
    for i, f in enumerate(data["funnel"], 2):
        ws1.cell(row=i, column=1, value=f["stage"])
        ws1.cell(row=i, column=2, value=f["label"])
        ws1.cell(row=i, column=3, value=f["count"])
        ws1.cell(row=i, column=4, value=f["amount"])
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 16

    # Sheet 2: Win/Loss
    ws2 = wb.create_sheet("赢单分析")
    style_header(ws2, ["指标", "值"])
    for i, (k, v) in enumerate([
        ("赢单数", data["won_count"]), ("丢单数", data["lost_count"]),
        ("赢单金额", data["won_amount"]), ("丢单金额", data["lost_amount"]),
        ("赢单率(%)", data["win_rate"]),
    ], 2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 16

    # Sheet 3: Top Customers
    ws3 = wb.create_sheet("客户TOP10")
    style_header(ws3, ["排名", "客户名称", "商机数", "总金额"])
    for i, c in enumerate(data["top_customers"], 2):
        ws3.cell(row=i, column=1, value=i - 1)
        ws3.cell(row=i, column=2, value=c["name"])
        ws3.cell(row=i, column=3, value=c["project_count"])
        ws3.cell(row=i, column=4, value=c["total_amount"])
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 16

    # Sheet 4: Region
    ws4 = wb.create_sheet("区域分布")
    style_header(ws4, ["区域", "客户数"])
    for i, r in enumerate(data["regions"], 2):
        ws4.cell(row=i, column=1, value=r["region"])
        ws4.cell(row=i, column=2, value=r["count"])
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"analytics_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/pdf")
async def export_pdf(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # Try to register a Chinese font; fall back to Helvetica
    font_name = "Helvetica"
    for font_path in [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    ]:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont("CJK", font_path))
                font_name = "CJK"
                break
            except Exception:
                continue

    data = await _gather_export_data(db, tenant_id, _user, start_date, end_date)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=15 * mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title_CJK", parent=styles["Title"], fontName=font_name, fontSize=18)
    h2_style = ParagraphStyle("H2_CJK", parent=styles["Heading2"], fontName=font_name, fontSize=13, spaceAfter=6)
    normal_style = ParagraphStyle("Normal_CJK", parent=styles["Normal"], fontName=font_name, fontSize=9)

    elements = []
    period = ""
    if start_date and end_date:
        period = f"  ({start_date} ~ {end_date})"
    elements.append(Paragraph(f"Analytics Report{period}", title_style))
    elements.append(Spacer(1, 8 * mm))

    table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FC")]),
    ])

    # Funnel
    elements.append(Paragraph("Sales Funnel", h2_style))
    t_data = [["Stage", "Label", "Count", "Amount"]]
    for f in data["funnel"]:
        t_data.append([f["stage"], f["label"], str(f["count"]), f"{f['amount']:,.0f}"])
    t = Table(t_data, colWidths=[60, 80, 60, 90])
    t.setStyle(table_style)
    elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # Win/Loss
    elements.append(Paragraph("Win/Loss Analysis", h2_style))
    t_data = [["Metric", "Value"]]
    t_data.append(["Won", f"{data['won_count']} ({data['won_amount']:,.0f})"])
    t_data.append(["Lost", f"{data['lost_count']} ({data['lost_amount']:,.0f})"])
    t_data.append(["Win Rate", f"{data['win_rate']}%"])
    t = Table(t_data, colWidths=[100, 150])
    t.setStyle(table_style)
    elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # Top Customers
    elements.append(Paragraph("Top Customers", h2_style))
    t_data = [["#", "Customer", "Projects", "Amount"]]
    for i, c in enumerate(data["top_customers"], 1):
        t_data.append([str(i), c["name"], str(c["project_count"]), f"{c['total_amount']:,.0f}"])
    t = Table(t_data, colWidths=[30, 160, 60, 90])
    t.setStyle(table_style)
    elements.append(t)
    elements.append(Spacer(1, 6 * mm))

    # Region
    if data["regions"]:
        elements.append(Paragraph("Customer Region Distribution", h2_style))
        t_data = [["Region", "Count"]]
        for r in data["regions"]:
            t_data.append([r["region"], str(r["count"])])
        t = Table(t_data, colWidths=[150, 80])
        t.setStyle(table_style)
        elements.append(t)

    doc.build(elements)
    buf.seek(0)
    filename = f"analytics_{date.today().isoformat()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# --- Win Forecast ---

STAGE_PROBABILITIES = {"S1": 0.10, "S2": 0.20, "S3": 0.40, "S4": 0.60, "S5": 0.80, "S6": 0.95}


@router.get("/win_forecast")
async def win_forecast(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Calculate weighted pipeline forecast by stage."""
    scope = await _scope_owner_ids(db, tenant_id, _user)
    projects = (await db.execute(
        select(
            OpportunityProject.stage_code,
            func.count(OpportunityProject.id).label("count"),
            func.coalesce(func.sum(OpportunityProject.amount_expect), 0).label("total"),
        ).where(
            OpportunityProject.tenant_id == tenant_id,
            OpportunityProject.is_deleted == False,
            OpportunityProject.status == "active",
            *_project_scope_where(tenant_id, _user, scope),
        ).group_by(OpportunityProject.stage_code)
    )).all()

    stages = []
    weighted_total = 0.0
    pipeline_total = 0.0
    for row in projects:
        prob = STAGE_PROBABILITIES.get(row.stage_code, 0.5)
        amount = float(row.total or 0)
        weighted = amount * prob
        weighted_total += weighted
        pipeline_total += amount
        stages.append({
            "stage": row.stage_code,
            "count": row.count,
            "amount": amount,
            "probability": prob,
            "weighted_amount": round(weighted, 2),
        })

    return ok({
        "stages": sorted(stages, key=lambda x: x["stage"]),
        "pipeline_total": round(pipeline_total, 2),
        "weighted_total": round(weighted_total, 2),
    })


@router.get("/rate_limit_stats")
async def rate_limit_stats(
    _user=Depends(require_permissions("role:manage")),
):
    """Return rate limiter stats for admin dashboard."""
    try:
        from app.middleware.rate_limiter import RateLimitMiddleware
        from app.main import app as fastapi_app
        current = fastapi_app
        for _ in range(20):
            if isinstance(current, RateLimitMiddleware):
                return ok(current.get_stats())
            current = getattr(current, "app", None)
            if current is None:
                break
        return ok({"rpm_limit": 120, "active_clients": 0, "total_rejected": 0, "clients": []})
    except Exception:
        return ok({"rpm_limit": 120, "active_clients": 0, "total_rejected": 0, "clients": []})


@router.get("/stage_duration")
async def stage_duration(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(get_current_user),
):
    """Analyze average time spent in each project stage."""
    from app.domains.project.models import ProjectStageHistory
    from collections import defaultdict

    # 阶段耗时按可见商机的推进记录统计（ProjectStageHistory 无 owner_id，只能靠父商机）
    scope = await _scope_owner_ids(db, tenant_id, _user)
    history_where = ([] if scope is None else
                     [ProjectStageHistory.project_id.in_(_visible_project_ids(tenant_id, _user, scope))])
    rows = (await db.execute(
        select(ProjectStageHistory).where(
            ProjectStageHistory.tenant_id == tenant_id,
            *history_where,
        ).order_by(ProjectStageHistory.project_id, ProjectStageHistory.created_at)
    )).scalars().all()

    projects: dict[str, list] = defaultdict(list)
    for r in rows:
        projects[r.project_id].append(r)

    stage_durations: dict[str, list[float]] = defaultdict(list)
    for pid, transitions in projects.items():
        for i, t in enumerate(transitions):
            entered = t.created_at
            if i + 1 < len(transitions):
                exited = transitions[i + 1].created_at
            else:
                exited = datetime.now(timezone.utc)
            days = (exited - entered).total_seconds() / 86400
            stage_durations[t.to_stage].append(days)

    stages = []
    for stage in ["S1", "S2", "S3", "S4", "S5", "S6"]:
        durations = stage_durations.get(stage, [])
        if durations:
            stages.append({
                "stage": stage,
                "avg_days": round(sum(durations) / len(durations), 1),
                "min_days": round(min(durations), 1),
                "max_days": round(max(durations), 1),
                "count": len(durations),
            })
        else:
            stages.append({"stage": stage, "avg_days": 0, "min_days": 0, "max_days": 0, "count": 0})

    return ok(stages)


# 说明：这里原本还有第二份 GET/POST/DELETE /targets（守卫写成 dashboard:view）。
# FastAPI 首个匹配优先，实际生效的一直是文件上方那一份（写操作守 role:manage），
# 第二份从未被路由到，属于死代码，已删除。不要改用 dashboard:view 当写守卫——
# 它在核心权限集里人人都有，等于放开任意员工增删销售目标。

# ── Dashboard Snapshot Sharing ─────────────────────────────────────

class SnapshotCreate(BaseModel):
    title: str = Field(..., max_length=200)
    snapshot_data: dict
    card_visibility: Optional[dict] = None
    card_order: Optional[list] = None
    expires_hours: Optional[int] = None  # None = never expires


@router.post("/snapshots")
async def create_snapshot(
    body: SnapshotCreate,
    tenant_id: str = Depends(get_tenant_id),
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    import secrets, json
    from datetime import timedelta

    token = secrets.token_urlsafe(32)
    expires_at = None
    if body.expires_hours:
        expires_at = datetime.now(timezone.utc) + timedelta(hours=body.expires_hours)

    snap = DashboardSnapshot(
        tenant_id=tenant_id,
        share_token=token,
        title=body.title,
        created_by=current_user["sub"],
        created_by_name=current_user.get("real_name") or current_user.get("username", ""),
        snapshot_json=json.dumps(body.snapshot_data, ensure_ascii=False),
        card_visibility_json=json.dumps(body.card_visibility) if body.card_visibility else None,
        card_order_json=json.dumps(body.card_order) if body.card_order else None,
        expires_at=expires_at,
    )
    db.add(snap)
    await db.commit()
    return ok({"id": snap.id, "share_token": token, "expires_at": expires_at.isoformat() if expires_at else None})


@router.get("/snapshots")
async def list_snapshots(
    tenant_id: str = Depends(get_tenant_id),
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    import json
    rows = (await db.execute(
        select(DashboardSnapshot).where(
            DashboardSnapshot.tenant_id == tenant_id,
            DashboardSnapshot.created_by == current_user["sub"],
        ).order_by(DashboardSnapshot.created_at.desc()).limit(50)
    )).scalars().all()
    return ok([{
        "id": s.id, "title": s.title, "share_token": s.share_token,
        "created_at": s.created_at.isoformat() if s.created_at else None,
        "expires_at": s.expires_at.isoformat() if s.expires_at else None,
    } for s in rows])


@router.get("/snapshots/{token}")
async def get_snapshot(
    token: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    import json
    from app.common.exceptions import BusinessException

    snap = (await db.execute(
        select(DashboardSnapshot).where(
            DashboardSnapshot.share_token == token,
            DashboardSnapshot.tenant_id == current_user["tenant_id"],
        )
    )).scalar_one_or_none()
    if not snap:
        raise BusinessException(code=404, message="快照不存在")
    if snap.expires_at and snap.expires_at < datetime.now(timezone.utc):
        raise BusinessException(code=410, message="快照已过期")

    return ok({
        "id": snap.id,
        "title": snap.title,
        "created_by_name": snap.created_by_name,
        "created_at": snap.created_at.isoformat() if snap.created_at else None,
        "snapshot_data": json.loads(snap.snapshot_json),
        "card_visibility": json.loads(snap.card_visibility_json) if snap.card_visibility_json else None,
        "card_order": json.loads(snap.card_order_json) if snap.card_order_json else None,
    })


@router.delete("/snapshots/{snapshot_id}")
async def delete_snapshot(
    snapshot_id: str,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.common.exceptions import BusinessException
    snap = (await db.execute(
        select(DashboardSnapshot).where(
            DashboardSnapshot.id == snapshot_id,
            DashboardSnapshot.tenant_id == current_user["tenant_id"],
            DashboardSnapshot.created_by == current_user["sub"],
        )
    )).scalar_one_or_none()
    if not snap:
        raise BusinessException(code=404, message="快照不存在")
    await db.delete(snap)
    await db.commit()
    return ok(None)
