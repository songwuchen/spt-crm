import io
import csv

from fastapi import APIRouter, Depends, Query, UploadFile, File
from openpyxl import load_workbook
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_db, get_tenant_id, require_permissions
from app.common.schemas import ok
from app.common.exceptions import BusinessException
from app.common.export import build_excel, excel_response
from app.common.field_mask import load_mask_policies, apply_field_mask
from app.domains.payment import service
from app.domains.payment.models import PaymentPlan, PaymentRecord, Invoice
from app.domains.payment.schemas import (
    InvoiceCreate, InvoiceUpdate, PaymentPlanCreate, PaymentPlanUpdate, PaymentRecordCreate,
    PaymentPlanBulkCreate,
)
from app.domains.project.models import OpportunityProject

router = APIRouter(tags=["回款管理"])


def _inv_dict(i) -> dict:
    return {
        "id": i.id, "project_id": i.project_id,
        "invoice_no": i.invoice_no,
        "amount": float(i.amount) if i.amount is not None else None,
        "invoice_date": str(i.invoice_date) if i.invoice_date else None,
        "status": i.status, "erp_ref_json": i.erp_ref_json, "remark": i.remark,
        "created_by_id": i.created_by_id, "created_by_name": i.created_by_name,
        "created_at": i.created_at.isoformat() if i.created_at else "",
    }


def _plan_dict(p) -> dict:
    return {
        "id": p.id, "project_id": p.project_id,
        "plan_no": p.plan_no,
        "due_date": str(p.due_date) if p.due_date else None,
        "amount": float(p.amount) if p.amount is not None else None,
        "trigger_milestone_code": p.trigger_milestone_code,
        "source_contract_id": p.source_contract_id,
        "status": p.status, "remark": p.remark,
        "assignee_id": p.assignee_id, "assignee_name": p.assignee_name,
        "department_id": p.department_id, "department_name": p.department_name,
        "created_at": p.created_at.isoformat() if p.created_at else "",
    }


def _rec_dict(r) -> dict:
    return {
        "id": r.id, "project_id": r.project_id,
        "received_date": str(r.received_date) if r.received_date else None,
        "amount": float(r.amount) if r.amount is not None else None,
        "channel": r.channel, "reference_no": r.reference_no,
        "matched_plan_id": r.matched_plan_id, "remark": r.remark,
        "created_by_id": r.created_by_id, "created_by_name": r.created_by_name,
        "custom_fields_json": r.custom_fields_json,
        "created_at": r.created_at.isoformat() if r.created_at else "",
    }


# --- Invoice ---

@router.get("/api/v1/projects/{project_id}/invoices")
async def list_invoices(
    project_id: str, tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db), _user=Depends(require_permissions("payment:view")),
):
    items = await service.list_invoices(db, tenant_id, project_id, _user)
    return ok([_inv_dict(i) for i in items])


@router.post("/api/v1/projects/{project_id}/invoices")
async def create_invoice(
    project_id: str, body: InvoiceCreate,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:edit")),
):
    inv = await service.create_invoice(db, tenant_id, project_id, body, current_user)
    return ok(_inv_dict(inv))


@router.put("/api/v1/invoices/{invoice_id}")
async def update_invoice(
    invoice_id: str, body: InvoiceUpdate,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:edit")),
):
    inv = await service.update_invoice(db, tenant_id, invoice_id, body, current_user)
    return ok(_inv_dict(inv))


@router.delete("/api/v1/invoices/{invoice_id}")
async def delete_invoice(
    invoice_id: str,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:edit")),
):
    await service.delete_invoice(db, tenant_id, invoice_id, current_user)
    return ok(None)


# --- PaymentPlan ---

@router.get("/api/v1/projects/{project_id}/payment_plans")
async def list_plans(
    project_id: str, tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db), _user=Depends(require_permissions("payment:view")),
):
    items = await service.list_plans(db, tenant_id, project_id, _user)
    return ok([_plan_dict(p) for p in items])


@router.post("/api/v1/projects/{project_id}/payment_plans")
async def create_plan(
    project_id: str, body: PaymentPlanCreate,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:edit")),
):
    plan = await service.create_plan(db, tenant_id, project_id, body, current_user)
    return ok(_plan_dict(plan))


@router.post("/api/v1/projects/{project_id}/payment_plans/bulk")
async def bulk_create_plans(
    project_id: str, body: PaymentPlanBulkCreate,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:edit")),
):
    """Bulk-create payment plans (e.g. generated from a contract's payment terms)."""
    plans = await service.bulk_create_plans(db, tenant_id, project_id, body, current_user)
    return ok([_plan_dict(p) for p in plans])


@router.put("/api/v1/payment_plans/{plan_id}")
async def update_plan(
    plan_id: str, body: PaymentPlanUpdate,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:edit")),
):
    plan = await service.update_plan(db, tenant_id, plan_id, body, current_user)
    return ok(_plan_dict(plan))


@router.delete("/api/v1/payment_plans/{plan_id}")
async def delete_plan(
    plan_id: str,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:edit")),
):
    await service.delete_plan(db, tenant_id, plan_id, current_user)
    return ok(None)


# --- PaymentRecord ---

@router.get("/api/v1/projects/{project_id}/payment_records")
async def list_records(
    project_id: str, tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db), _user=Depends(require_permissions("payment:view")),
):
    items = await service.list_records(db, tenant_id, project_id, _user)
    return ok([_rec_dict(r) for r in items])


@router.post("/api/v1/projects/{project_id}/payment_records")
async def create_record(
    project_id: str, body: PaymentRecordCreate,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:edit")),
):
    rec = await service.create_record(db, tenant_id, project_id, body, current_user)
    return ok(_rec_dict(rec))


@router.delete("/api/v1/payment_records/{record_id}")
async def delete_record(
    record_id: str,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:edit")),
):
    await service.delete_record(db, tenant_id, record_id, current_user)
    return ok(None)


# --- 到账记录批量导入 ---

_REC_IMPORT_COLS = ["商机编号", "到账日期", "金额", "渠道", "凭证号", "备注"]


@router.get("/api/v1/payment/records/import/template")
async def payment_records_template(_user=Depends(require_permissions("payment:view"))):
    """下载到账记录导入模板（表头 + 示例行）。商机编号用于把每条到账关联到对应商机。"""
    example = ["P-2026-0001", "2026-06-01", 100000, "电汇", "TT20260601", "首付款"]
    buf = build_excel("到账记录导入模板", _REC_IMPORT_COLS, [example])
    return excel_response(buf, "payment_records_template.xlsx")


@router.post("/api/v1/payment/records/import")
async def import_payment_records(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:edit")),
):
    """批量导入到账记录。列顺序：商机编号, 到账日期, 金额, 渠道, 凭证号, 备注。
    每行按"商机编号"关联到对应商机；找不到编号的行计入错误，不中断其余行。"""
    content = await file.read()
    fname = (file.filename or "").lower()
    try:
        if fname.endswith(".csv"):
            text = content.decode("utf-8-sig", errors="replace")
            all_rows = [tuple(r) for r in csv.reader(io.StringIO(text))]
        else:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True)) if ws is not None else []
            wb.close()
    except Exception:
        raise BusinessException(message="无法解析文件，请使用导入模板（.xlsx 或 .csv）")

    if len(all_rows) < 2:
        return ok({"created": 0, "skipped": 0, "errors": []})
    data_rows = all_rows[1:]

    # 预取商机编号 -> id 映射（仅本租户）
    codes = {str(r[0]).strip() for r in data_rows if r and len(r) > 0 and r[0]}
    code_to_id: dict[str, str] = {}
    if codes:
        rows = (await db.execute(
            select(OpportunityProject.project_code, OpportunityProject.id).where(
                OpportunityProject.tenant_id == tenant_id,
                OpportunityProject.project_code.in_(codes),
            )
        )).all()
        code_to_id = {c: i for c, i in rows}

    created, skipped, errors = 0, 0, []
    for idx, row in enumerate(data_rows, 2):
        if not row or not any(row):
            continue
        code = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        if not code:
            skipped += 1
            continue
        pid = code_to_id.get(code)
        if not pid:
            errors.append(f"第{idx}行: 商机编号「{code}」不存在")
            continue

        def _cell(i: int):
            return str(row[i]).strip() if len(row) > i and row[i] is not None else None

        amount = None
        try:
            amount = float(row[2]) if len(row) > 2 and row[2] is not None and str(row[2]).strip() != "" else None
        except (ValueError, TypeError):
            errors.append(f"第{idx}行: 金额格式错误")
            continue
        try:
            body = PaymentRecordCreate(
                received_date=_cell(1), amount=amount,
                channel=_cell(3), reference_no=_cell(4), remark=_cell(5),
            )
            await service.create_record(db, tenant_id, pid, body, current_user)
            created += 1
        except Exception as e:
            await db.rollback()
            errors.append(f"第{idx}行: {str(e)[:80]}")
    return ok({"created": created, "skipped": skipped, "errors": errors})


# --- Cross-project listing ---

@router.get("/api/v1/payment/plans")
async def list_all_plans(
    pageNo: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    status: str = Query(None),
    keyword: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:view")),
):
    """List all payment plans across projects with pagination."""
    q = select(
        PaymentPlan, OpportunityProject.name.label("project_name"),
        OpportunityProject.project_code.label("project_code"),
    ).outerjoin(OpportunityProject, OpportunityProject.id == PaymentPlan.project_id).where(
        PaymentPlan.tenant_id == tenant_id
    )
    count_q = select(func.count(PaymentPlan.id)).where(PaymentPlan.tenant_id == tenant_id)

    if status:
        q = q.where(PaymentPlan.status == status)
        count_q = count_q.where(PaymentPlan.status == status)
    if keyword:
        kw = f"%{keyword}%"
        flt = or_(PaymentPlan.plan_no.ilike(kw), PaymentPlan.remark.ilike(kw))
        q = q.where(flt)
        count_q = count_q.where(flt)

    from app.common.data_scope import apply_project_child_scope
    q, count_q = await apply_project_child_scope(q, count_q, db, tenant_id, current_user, PaymentPlan)
    total = (await db.execute(count_q)).scalar() or 0
    rows = (await db.execute(
        q.order_by(PaymentPlan.due_date.asc())
        .offset((pageNo - 1) * pageSize).limit(pageSize)
    )).all()

    items = []
    for row in rows:
        plan = row[0]
        d = _plan_dict(plan)
        d["project_name"] = row.project_name
        d["project_code"] = row.project_code
        items.append(d)

    # 回款计划没有 custom_fields_json 列、不属于 entity_type="payment"，
    # 其金额改由按权限脱敏的 field_mask 覆盖（与报价成本/毛利同一套引擎）
    items = apply_field_mask(items, "payment_plan",
                             current_user.get("permissions", []),
                             await load_mask_policies(db, tenant_id))
    return ok({"items": items, "total": total, "pageNo": pageNo, "pageSize": pageSize})


@router.get("/api/v1/payment/records")
async def list_all_records(
    pageNo: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    keyword: str = Query(None),
    filter: str = Query(None, description="高级筛选 FilterDsl(JSON)"),
    sort_by: str = Query(None),
    sort_order: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:view")),
):
    """List all payment records across projects with pagination."""
    q = select(
        PaymentRecord, OpportunityProject.name.label("project_name"),
        OpportunityProject.project_code.label("project_code"),
    ).outerjoin(OpportunityProject, OpportunityProject.id == PaymentRecord.project_id).where(
        PaymentRecord.tenant_id == tenant_id
    )
    count_q = select(func.count(PaymentRecord.id)).where(PaymentRecord.tenant_id == tenant_id)

    if keyword:
        kw = f"%{keyword}%"
        flt = or_(PaymentRecord.reference_no.ilike(kw), PaymentRecord.channel.ilike(kw), PaymentRecord.remark.ilike(kw))
        q = q.where(flt)
        count_q = count_q.where(flt)

    # 高级筛选（多字段/多条件，含自定义扩展字段）
    from app.common.search import (
        entity_search_context, filter_clause_from_schema_or_400, resolve_sort_from_schema,
    )
    search_schema = await entity_search_context("payment", db, tenant_id)
    clause = filter_clause_from_schema_or_400(search_schema, filter, {"user_id": current_user.get("sub")})
    if clause is not None:
        q = q.where(clause)
        count_q = count_q.where(clause)

    from app.common.data_scope import apply_project_child_scope
    q, count_q = await apply_project_child_scope(q, count_q, db, tenant_id, current_user, PaymentRecord)
    total = (await db.execute(count_q)).scalar() or 0
    order = resolve_sort_from_schema(search_schema, sort_by, sort_order, PaymentRecord.received_date.desc())
    rows = (await db.execute(
        q.order_by(order)
        .offset((pageNo - 1) * pageSize).limit(pageSize)
    )).all()

    items = []
    for row in rows:
        rec = row[0]
        d = _rec_dict(rec)
        d["project_name"] = row.project_name
        d["project_code"] = row.project_code
        items.append(d)

    from app.domains.lowcode.field_permission import strip_entity_dicts
    await strip_entity_dicts(db, tenant_id, "payment", items, current_user.get("roles"))  # 字段级权限：剔除隐藏扩展字段
    return ok({"items": items, "total": total, "pageNo": pageNo, "pageSize": pageSize})


@router.get("/api/v1/payment/invoices")
async def list_all_invoices(
    pageNo: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    status: str = Query(None),
    keyword: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("payment:view")),
):
    """List all invoices across projects with pagination."""
    q = select(
        Invoice, OpportunityProject.name.label("project_name"),
        OpportunityProject.project_code.label("project_code"),
    ).outerjoin(OpportunityProject, OpportunityProject.id == Invoice.project_id).where(
        Invoice.tenant_id == tenant_id
    )
    count_q = select(func.count(Invoice.id)).where(Invoice.tenant_id == tenant_id)

    if status:
        q = q.where(Invoice.status == status)
        count_q = count_q.where(Invoice.status == status)
    if keyword:
        kw = f"%{keyword}%"
        flt = or_(Invoice.invoice_no.ilike(kw), Invoice.remark.ilike(kw))
        q = q.where(flt)
        count_q = count_q.where(flt)

    from app.common.data_scope import apply_project_child_scope
    q, count_q = await apply_project_child_scope(q, count_q, db, tenant_id, current_user, Invoice)
    total = (await db.execute(count_q)).scalar() or 0
    rows = (await db.execute(
        q.order_by(Invoice.created_at.desc())
        .offset((pageNo - 1) * pageSize).limit(pageSize)
    )).all()

    items = []
    for row in rows:
        inv = row[0]
        d = _inv_dict(inv)
        d["project_name"] = row.project_name
        d["project_code"] = row.project_code
        items.append(d)

    items = apply_field_mask(items, "invoice",
                             current_user.get("permissions", []),
                             await load_mask_policies(db, tenant_id))
    return ok({"items": items, "total": total, "pageNo": pageNo, "pageSize": pageSize})


@router.get("/api/v1/payment/export/excel")
async def export_payments_excel(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("payment:view")),
):
    """Export payment plans + records to Excel."""
    from app.config import settings
    from app.common.data_scope import apply_project_child_scope
    # 导出与列表同口径按所属商机过滤，否则「列表看不到但能导出来」就是绕过范围的后门
    # Plans
    plan_q = (
        select(PaymentPlan, OpportunityProject.name.label("project_name"))
        .outerjoin(OpportunityProject, OpportunityProject.id == PaymentPlan.project_id)
        .where(PaymentPlan.tenant_id == tenant_id)
    )
    plan_q, _ = await apply_project_child_scope(plan_q, plan_q, db, tenant_id, _user, PaymentPlan)
    plan_rows = (await db.execute(
        plan_q.order_by(PaymentPlan.due_date.asc()).limit(settings.MAX_EXPORT_ROWS)
    )).all()
    # Records
    rec_q = (
        select(PaymentRecord, OpportunityProject.name.label("project_name"))
        .outerjoin(OpportunityProject, OpportunityProject.id == PaymentRecord.project_id)
        .where(PaymentRecord.tenant_id == tenant_id)
    )
    rec_q, _ = await apply_project_child_scope(rec_q, rec_q, db, tenant_id, _user, PaymentRecord)
    rec_rows = (await db.execute(
        rec_q.order_by(PaymentRecord.received_date.desc()).limit(settings.MAX_EXPORT_ROWS)
    )).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    wb = Workbook()
    # Sheet 1: Plans
    ws1 = wb.active
    ws1.title = "回款计划"
    h1 = ["计划编号", "项目名称", "到期日", "金额", "状态", "里程碑", "备注"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
    )
    for ci, h in enumerate(h1, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for ri, row in enumerate(plan_rows, 2):
        p = row[0]
        vals = [p.plan_no, row.project_name or "", str(p.due_date) if p.due_date else "",
                float(p.amount) if p.amount else 0, p.status or "",
                p.trigger_milestone_code or "", p.remark or ""]
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=v)
            cell.border = thin_border

    # Sheet 2: Records
    ws2 = wb.create_sheet("回款记录")
    h2 = ["项目名称", "到账日", "金额", "渠道", "参考号", "备注"]
    for ci, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    # 到账记录导出与列表同口径脱敏（金额/渠道/凭证号是典型的按角色控制对象）。
    # 回款计划(PaymentPlan)不属于 entity_type="payment"，无扩展字段列，故不在此裁剪范围内。
    from app.domains.lowcode.field_permission import entity_field_restrictions, export_cell
    rst = await entity_field_restrictions(db, tenant_id, "payment", _user.get("roles"))
    for ri, row in enumerate(rec_rows, 2):
        r = row[0]
        vals = [row.project_name or "",
                export_cell(rst, "received_date", str(r.received_date) if r.received_date else ""),
                export_cell(rst, "amount", float(r.amount) if r.amount else 0),
                export_cell(rst, "channel", r.channel or ""),
                export_cell(rst, "reference_no", r.reference_no or ""),
                export_cell(rst, "remark", r.remark or "")]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return excel_response(buf, "payments.xlsx")


@router.post("/api/v1/payment/check_overdue")
async def check_overdue(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("payment:view")),
):
    """Manually trigger overdue check and notifications. Can also be called by cron/worker."""
    notified = await service.check_overdue_and_notify(db, tenant_id)
    return ok({"notified_projects": notified})
