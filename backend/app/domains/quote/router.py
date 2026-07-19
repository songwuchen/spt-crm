import io

from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import Response
from openpyxl import load_workbook
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_db, get_tenant_id, require_permissions
from app.common.schemas import ok
from app.common.export import build_template, excel_response
from app.common.field_mask import load_mask_policies, apply_field_mask, masked_number
from app.domains.quote import service
from app.domains.quote.schemas import QuoteCreate, QuoteUpdate, QuoteVersionUpdate, QuoteLineCreate, QuoteLineUpdate, CostSnapshotCreate, QuoteSendLogCreate

router = APIRouter(tags=["报价管理"])

# 报价行项目导入列（顺序即模板列顺序）
QUOTE_LINE_IMPORT_HEADERS = ["类型", "编码", "品名", "规格", "数量", "单位", "单价", "估计成本", "交期(天)"]

# 类型列既接受编码也接受中文标签
_ITEM_TYPE_BY_LABEL = {
    "标准品": "standard", "standard": "standard",
    "非标品": "nonstandard", "nonstandard": "nonstandard",
    "服务": "service", "service": "service",
    "备件": "spare", "spare": "spare",
}


def _quote_dict(q) -> dict:
    return {
        "id": q.id, "project_id": q.project_id, "quote_no": q.quote_no,
        "current_version_no": q.current_version_no, "status": q.status,
        "created_by_id": q.created_by_id, "created_by_name": q.created_by_name,
        "assignee_id": q.assignee_id, "assignee_name": q.assignee_name,
        "department_id": q.department_id, "department_name": q.department_name,
        "custom_fields_json": q.custom_fields_json,
        "created_at": q.created_at.isoformat() if q.created_at else "",
        "updated_at": q.updated_at.isoformat() if q.updated_at else "",
    }


def _version_dict(v) -> dict:
    return {
        "id": v.id, "quote_id": v.quote_id, "version_no": v.version_no,
        "title": v.title,
        "price_total": float(v.price_total) if v.price_total is not None else None,
        "tax_rate": float(v.tax_rate) if v.tax_rate is not None else None,
        "tax_total": float(v.tax_total) if v.tax_total is not None else None,
        "discount_total": float(v.discount_total) if v.discount_total is not None else None,
        "margin_rate": float(v.margin_rate) if v.margin_rate is not None else None,
        "delivery_promise_date": str(v.delivery_promise_date) if v.delivery_promise_date else None,
        "validity_days": v.validity_days,
        "terms_summary_json": v.terms_summary_json,
        "status": v.status,
        "created_at": v.created_at.isoformat() if v.created_at else "",
    }


def _line_dict(l) -> dict:
    return {
        "id": l.id, "quote_version_id": l.quote_version_id, "line_no": l.line_no,
        "item_type": l.item_type, "item_name": l.item_name,
        "item_code": l.item_code, "spec": l.spec,
        "qty": float(l.qty) if l.qty is not None else None,
        "unit": l.unit,
        "unit_price": float(l.unit_price) if l.unit_price is not None else None,
        "line_total": float(l.line_total) if l.line_total is not None else None,
        "cost_est": float(l.cost_est) if l.cost_est is not None else None,
        "leadtime_days": l.leadtime_days,
    }


# --- List all quotes (tenant-wide) ---
@router.get("/api/v1/quotes")
async def list_quotes(
    pageNo: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    status: str = Query(None),
    keyword: str = Query(None),
    filter: str = Query(None, description="高级筛选 FilterDsl(JSON)"),
    sort_by: str = Query(None),
    sort_order: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:view")),
):
    from app.domains.quote.models import Quote, QuoteVersion
    q = select(Quote).where(Quote.tenant_id == tenant_id)
    cq = select(func.count(Quote.id)).where(Quote.tenant_id == tenant_id)
    if status:
        q = q.where(Quote.status == status)
        cq = cq.where(Quote.status == status)
    if keyword:
        like = f"%{keyword}%"
        q = q.where(Quote.quote_no.ilike(like))
        cq = cq.where(Quote.quote_no.ilike(like))
    # 高级筛选（多字段/多条件，含自定义扩展字段）
    from app.common.search import (
        entity_search_context, filter_clause_from_schema_or_400, resolve_sort_from_schema,
    )
    search_schema = await entity_search_context("quote", db, tenant_id)
    clause = filter_clause_from_schema_or_400(search_schema, filter, {"user_id": current_user.get("sub")})
    if clause is not None:
        q = q.where(clause)
        cq = cq.where(clause)
    from app.common.data_scope import apply_project_child_scope
    q, cq = await apply_project_child_scope(q, cq, db, tenant_id, current_user, Quote)
    total = (await db.execute(cq)).scalar() or 0
    order = resolve_sort_from_schema(search_schema, sort_by, sort_order, Quote.created_at.desc())
    quotes = (await db.execute(
        q.order_by(order)
        .offset((pageNo - 1) * pageSize).limit(pageSize)
    )).scalars().all()

    # Pull the current version of each listed quote in one query for price/margin.
    cur_by_quote: dict = {}
    qids = [x.id for x in quotes]
    if qids:
        versions = (await db.execute(
            select(QuoteVersion).where(
                QuoteVersion.tenant_id == tenant_id, QuoteVersion.quote_id.in_(qids)
            )
        )).scalars().all()
        for v in versions:
            cur_by_quote[(v.quote_id, v.version_no)] = v

    from app.common.list_enrich import project_names_map
    name_map = await project_names_map(db, tenant_id, [x.project_id for x in quotes])

    rows = []
    for x in quotes:
        d = _quote_dict(x)
        cv = cur_by_quote.get((x.id, x.current_version_no))
        d["price_total"] = float(cv.price_total) if cv and cv.price_total is not None else None
        d["margin_rate"] = float(cv.margin_rate) if cv and cv.margin_rate is not None else None
        d["version_status"] = cv.status if cv else None
        d.update(name_map.get(x.project_id) or {})
        rows.append(d)

    perms = current_user.get("permissions", [])
    policies = await load_mask_policies(db, tenant_id)
    rows = apply_field_mask(rows, "quote", perms, policies)
    from app.domains.lowcode.field_permission import strip_entity_dicts
    await strip_entity_dicts(db, tenant_id, "quote", rows, current_user.get("roles"))  # 字段级权限：剔除隐藏扩展字段
    return ok({"items": rows, "total": total})


# --- Project-scoped routes ---
@router.get("/api/v1/projects/{project_id}/quotes")
async def list_project_quotes(
    project_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("quote:view")),
):
    items = await service.list_quotes_by_project(db, tenant_id, project_id)
    return ok([_quote_dict(q) for q in items])


@router.post("/api/v1/projects/{project_id}/quotes")
async def create_quote(
    project_id: str,
    body: QuoteCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:create")),
):
    result = await service.create_quote(db, tenant_id, project_id, body, current_user)
    return ok({
        "quote": _quote_dict(result["quote"]),
        "version": _version_dict(result["version"]),
    })


# --- Quote routes ---
@router.get("/api/v1/quotes/{quote_id}")
async def get_quote(
    quote_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:view")),
):
    quote = await service.get_quote(db, tenant_id, quote_id)
    versions = await service.get_versions_by_quote(db, tenant_id, quote_id)

    # Get current version lines
    current_ver = next((v for v in versions if v.version_no == quote.current_version_no), None)
    lines = []
    if current_ver:
        lines = await service.list_lines(db, tenant_id, current_ver.id)

    # Apply field masking
    perms = current_user.get("permissions", [])
    policies = await load_mask_policies(db, tenant_id)
    version_dicts = apply_field_mask([_version_dict(v) for v in versions], "quote_version", perms, policies)
    cur_ver_dict = apply_field_mask(_version_dict(current_ver), "quote_version", perms, policies) if current_ver else None
    line_dicts = apply_field_mask([_line_dict(l) for l in lines], "quote_line", perms, policies)

    # 字段级权限：详情此前漏了这一步（列表有），隐藏/脱敏的扩展字段可从详情页直接读到
    from app.domains.lowcode.field_permission import strip_entity_dicts
    quote_dict = _quote_dict(quote)
    await strip_entity_dicts(db, tenant_id, "quote", [quote_dict], current_user.get("roles"))

    return ok({
        **quote_dict,
        "versions": version_dicts,
        "current_version": cur_ver_dict,
        "lines": line_dicts,
    })


@router.put("/api/v1/quotes/{quote_id}")
async def update_quote(
    quote_id: str,
    body: QuoteUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:edit")),
):
    q = await service.update_quote(db, tenant_id, quote_id, body, current_user)
    return ok(_quote_dict(q))


@router.delete("/api/v1/quotes/{quote_id}")
async def delete_quote(
    quote_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:delete")),
):
    await service.delete_quote(db, tenant_id, quote_id, current_user)
    return ok()


@router.post("/api/v1/quotes/{quote_id}/new_version")
async def new_version(
    quote_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:edit")),
):
    v = await service.new_version(db, tenant_id, quote_id, current_user)
    return ok(_version_dict(v))


# --- Version routes ---
@router.get("/api/v1/quote_versions/{version_id}")
async def get_version(
    version_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:view")),
):
    version = await service.get_version(db, tenant_id, version_id)
    lines = await service.list_lines(db, tenant_id, version_id)
    perms = current_user.get("permissions", [])
    policies = await load_mask_policies(db, tenant_id)
    ver_dict = apply_field_mask(_version_dict(version), "quote_version", perms, policies)
    line_dicts = apply_field_mask([_line_dict(l) for l in lines], "quote_line", perms, policies)
    return ok({**ver_dict, "lines": line_dicts})


@router.put("/api/v1/quote_versions/{version_id}")
async def update_version(
    version_id: str,
    body: QuoteVersionUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:edit")),
):
    v = await service.update_version(db, tenant_id, version_id, body, current_user)
    return ok(_version_dict(v))


@router.post("/api/v1/quote_versions/{version_id}/lines")
async def add_line(
    version_id: str,
    body: QuoteLineCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:edit")),
):
    l = await service.add_line(db, tenant_id, version_id, body, current_user)
    return ok(_line_dict(l))


@router.put("/api/v1/quote_lines/{line_id}")
async def update_line(
    line_id: str,
    body: QuoteLineUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:edit")),
):
    l = await service.update_line(db, tenant_id, line_id, body, current_user)
    return ok(_line_dict(l))


@router.delete("/api/v1/quote_lines/{line_id}")
async def delete_line(
    line_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:edit")),
):
    await service.delete_line(db, tenant_id, line_id, current_user)
    return ok()


@router.get("/api/v1/quote_versions/{version_id}/lines/import/template")
async def download_quote_line_import_template(
    version_id: str,
    _user=Depends(require_permissions("quote:edit")),
):
    """下载报价行项目导入模板 (issue #93)。"""
    sample = [["标准品", "PRD-001", "示例产品", "500×300×200", 2, "台", 1500, 1200, 15]]
    buf = build_template("报价行项目导入模板", QUOTE_LINE_IMPORT_HEADERS, sample)
    return excel_response(buf, "quote_line_import_template.xlsx")


def _num_cell(v):
    """把单元格转为 float；空/无法解析返回 None。"""
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


@router.post("/api/v1/quote_versions/{version_id}/lines/import")
async def import_quote_lines(
    version_id: str,
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:edit")),
):
    """从 Excel 批量导入报价行项目 (issue #93)。
    列顺序：类型, 编码, 品名, 规格, 数量, 单位, 单价, 估计成本, 交期(天)。
    导入前校验版本存在（且属于本租户）。"""
    # 校验版本存在，避免把行项目挂到不存在/越权的版本上
    await service.get_version(db, tenant_id, version_id)

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    errors = []

    def cell(row, i):
        return row[i] if len(row) > i and row[i] not in (None, "") else None

    for idx, row in enumerate(rows, 2):
        # 跳过完全空行
        if not row or all(c in (None, "") for c in row):
            continue
        try:
            qty = _num_cell(cell(row, 4))
            unit_price = _num_cell(cell(row, 6))
            if qty is None or qty <= 0:
                raise ValueError("数量必须为正数")
            if unit_price is None or unit_price < 0:
                raise ValueError("单价必须为非负数")
            item_type_raw = cell(row, 0)
            body = QuoteLineCreate(
                item_type=_ITEM_TYPE_BY_LABEL.get(str(item_type_raw).strip()) if item_type_raw else None,
                item_code=str(cell(row, 1)).strip() if cell(row, 1) else None,
                item_name=str(cell(row, 2)).strip() if cell(row, 2) else None,
                spec=str(cell(row, 3)).strip() if cell(row, 3) else None,
                qty=qty,
                unit=str(cell(row, 5)).strip() if cell(row, 5) else None,
                unit_price=unit_price,
                cost_est=_num_cell(cell(row, 7)),
                leadtime_days=int(_num_cell(cell(row, 8))) if _num_cell(cell(row, 8)) is not None else None,
            )
            await service.add_line(db, tenant_id, version_id, body, current_user)
            created += 1
        except Exception as e:
            errors.append(f"第{idx}行: {str(e)[:80]}")
    wb.close()
    return ok({"created": created, "skipped": 0, "errors": errors})


# --- Version Comparison ---
@router.get("/api/v1/quotes/{quote_id}/compare")
async def compare_versions(
    quote_id: str,
    version_a: str = Query(...),
    version_b: str = Query(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("quote:view")),
):
    result = await service.compare_versions(db, tenant_id, version_a, version_b)
    return ok(result)


# --- Cost Snapshots ---
def _snapshot_dict(s) -> dict:
    return {
        "id": s.id, "quote_version_id": s.quote_version_id,
        "snapshot_type": s.snapshot_type,
        "price_total": float(s.price_total) if s.price_total is not None else None,
        "cost_total": float(s.cost_total) if s.cost_total is not None else None,
        "margin_rate": float(s.margin_rate) if s.margin_rate is not None else None,
        "breakdown_json": s.breakdown_json,
        "line_snapshot_json": s.line_snapshot_json,
        "note": s.note,
        "created_by_name": s.created_by_name,
        "created_at": s.created_at.isoformat() if s.created_at else "",
    }


def _send_log_dict(l) -> dict:
    return {
        "id": l.id, "quote_id": l.quote_id, "quote_version_id": l.quote_version_id,
        "channel": l.channel, "to_list_json": l.to_list_json,
        "subject": l.subject, "body": l.body,
        "attachments_json": l.attachments_json,
        "status": l.status,
        "sent_by_id": l.sent_by_id, "sent_by_name": l.sent_by_name,
        "created_at": l.created_at.isoformat() if l.created_at else "",
    }


@router.get("/api/v1/quote_versions/{version_id}/cost_snapshots")
async def list_cost_snapshots(
    version_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:view")),
):
    items = await service.list_cost_snapshots(db, tenant_id, version_id)
    perms = current_user.get("permissions", [])
    policies = await load_mask_policies(db, tenant_id)
    return ok(apply_field_mask([_snapshot_dict(s) for s in items], "cost_snapshot", perms, policies))


@router.post("/api/v1/quote_versions/{version_id}/cost_snapshots")
async def create_cost_snapshot(
    version_id: str,
    body: CostSnapshotCreate = CostSnapshotCreate(),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:edit")),
):
    s = await service.create_cost_snapshot(db, tenant_id, version_id, current_user, data=body)
    return ok(_snapshot_dict(s))


# --- Quote Send Log ---
@router.post("/api/v1/quote_versions/{version_id}/send")
async def send_quote(
    version_id: str,
    body: QuoteSendLogCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("quote:edit")),
):
    # Get quote_id from version
    version = await service.get_version(db, tenant_id, version_id)
    log = await service.create_send_log(db, tenant_id, version.quote_id, version_id, body, current_user)
    return ok(_send_log_dict(log))


@router.get("/api/v1/quotes/{quote_id}/send_logs")
async def list_send_logs(
    quote_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("quote:view")),
):
    items = await service.list_send_logs(db, tenant_id, quote_id)
    return ok([_send_log_dict(l) for l in items])


@router.get("/api/v1/quotes/{quote_id}/export/pdf")
async def export_quote_pdf(
    quote_id: str,
    version_id: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("quote:view")),
):
    """Export quote as PDF."""
    from app.common.pdf_builder import build_quote_pdf

    quote = await service.get_quote(db, tenant_id, quote_id)
    versions = await service.get_versions_by_quote(db, tenant_id, quote_id)

    if version_id:
        ver = next((v for v in versions if v.id == version_id), None)
    else:
        ver = next((v for v in versions if v.version_no == quote.current_version_no), None)

    if not ver:
        from app.common.exceptions import BusinessException
        raise BusinessException(message="未找到报价版本")

    lines = await service.list_lines(db, tenant_id, ver.id)

    # 导出同样受字段脱敏约束：此前页面上显示 *** 的折扣，导出 PDF 却打印真实金额，
    # 等于给了一条绕过脱敏的后门。
    # 注意必须**透传脱敏后的值**，不能只判断是否等于 "***" 再回头读 ver.discount_total ——
    # mask_type 还有 null / zero 两种，那样写这两种类型仍会打印真值。
    masked_ver = apply_field_mask(
        _version_dict(ver), "quote_version",
        _user.get("permissions", []), await load_mask_policies(db, tenant_id),
    )

    pdf_bytes = build_quote_pdf(
        quote_no=quote.quote_no,
        version_title=ver.title or "",
        version_no=ver.version_no,
        price_total=float(ver.price_total or 0),
        tax_rate=float(ver.tax_rate or 0),
        tax_total=float(ver.tax_total or 0),
        # 脱敏后取不到数值 → 传 0，builder 便不渲染折扣行
        discount_total=masked_number(masked_ver.get("discount_total"), 0.0) or 0.0,
        delivery_promise_date=str(ver.delivery_promise_date) if ver.delivery_promise_date else None,
        validity_days=ver.validity_days,
        lines=[_line_dict(l) for l in lines],
        created_by_name=quote.created_by_name or "",
        created_at=quote.created_at.isoformat() if quote.created_at else "",
    )

    filename = f"quote_{quote.quote_no}_v{ver.version_no}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/api/v1/quotes/batch_export/pdf")
async def batch_export_quote_pdf(
    body: dict,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("quote:view")),
):
    """Batch export multiple quotes as a ZIP of PDFs."""
    import io
    import zipfile
    from app.common.pdf_builder import build_quote_pdf

    ids = body.get("ids", [])
    if not ids:
        from app.common.exceptions import BusinessException
        raise BusinessException(message="请选择要导出的报价")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for qid in ids[:50]:  # cap at 50
            try:
                quote = await service.get_quote(db, tenant_id, qid)
                versions = await service.get_versions_by_quote(db, tenant_id, qid)
                ver = next((v for v in versions if v.version_no == quote.current_version_no), None)
                if not ver:
                    continue
                lines = await service.list_lines(db, tenant_id, ver.id)
                pdf_bytes = build_quote_pdf(
                    quote_no=quote.quote_no,
                    version_title=ver.title or "",
                    version_no=ver.version_no,
                    price_total=float(ver.price_total or 0),
                    tax_rate=float(ver.tax_rate or 0),
                    tax_total=float(ver.tax_total or 0),
                    discount_total=float(ver.discount_total or 0),
                    delivery_promise_date=str(ver.delivery_promise_date) if ver.delivery_promise_date else None,
                    validity_days=ver.validity_days,
                    lines=[_line_dict(l) for l in lines],
                    created_by_name=quote.created_by_name or "",
                    created_at=quote.created_at.isoformat() if quote.created_at else "",
                )
                zf.writestr(f"quote_{quote.quote_no}_v{ver.version_no}.pdf", pdf_bytes)
            except Exception:
                continue

    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="quotes_export.zip"'},
    )
