import io

from fastapi import APIRouter, Depends, Query, UploadFile, File
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_db, get_tenant_id, require_permissions
from app.common.schemas import ok
from app.common.export import build_excel, build_template, excel_response
from app.domains.service_ticket import service
from app.domains.lowcode.field_permission import strip_entity_dicts
from app.domains.service_ticket.schemas import (
    ServiceTicketCreate, ServiceTicketUpdate, RenewalCreate, RenewalUpdate,
)
from app.domains.order import service as order_service
from app.domains.customer.models import Customer

router = APIRouter(tags=["售后管理"])

# 工单类型 / 优先级：code <-> 中文标签（与前端 constants/labels 保持一致），供导入解析用
TYPE_LABELS = {"fault": "故障", "maintenance": "维保", "training": "培训", "spare": "备件", "upgrade": "升级改造"}
PRIORITY_LABELS = {"low": "低", "medium": "中", "high": "高", "critical": "紧急"}
_TYPE_BY_LABEL = {v: k for k, v in TYPE_LABELS.items()}
_PRIORITY_BY_LABEL = {v: k for k, v in PRIORITY_LABELS.items()}
# 导入列（顺序即模板列顺序）
IMPORT_HEADERS = ["工单类型", "优先级", "关联客户", "关联订单号", "负责人", "问题描述"]


def _ticket_dict(t) -> dict:
    return {
        "id": t.id, "customer_id": t.customer_id, "project_id": t.project_id,
        "order_id": t.order_id,
        "ticket_no": t.ticket_no, "type": t.type,
        "priority": t.priority, "status": t.status,
        "description": t.description, "resolution": t.resolution,
        "ai_summary_json": t.ai_summary_json,
        "assigned_to_id": t.assigned_to_id, "assigned_to_name": t.assigned_to_name,
        "created_by_id": t.created_by_id, "created_by_name": t.created_by_name,
        "sla_respond_by": t.sla_respond_by.isoformat() if t.sla_respond_by else None,
        "sla_resolve_by": t.sla_resolve_by.isoformat() if t.sla_resolve_by else None,
        "sla_responded_at": t.sla_responded_at.isoformat() if t.sla_responded_at else None,
        "sla_resolved_at": t.sla_resolved_at.isoformat() if t.sla_resolved_at else None,
        "satisfaction_score": t.satisfaction_score,
        "satisfaction_comment": t.satisfaction_comment,
        "satisfaction_at": t.satisfaction_at.isoformat() if t.satisfaction_at else None,
        "custom_fields_json": t.custom_fields_json,
        "created_at": t.created_at.isoformat() if t.created_at else "",
        "updated_at": t.updated_at.isoformat() if t.updated_at else "",
    }


def _renewal_dict(r) -> dict:
    return {
        "id": r.id, "customer_id": r.customer_id,
        "name": r.name,
        "amount_expect": float(r.amount_expect) if r.amount_expect is not None else None,
        "close_date_expect": str(r.close_date_expect) if r.close_date_expect else None,
        "probability": r.probability,
        "related_asset_json": r.related_asset_json,
        "status": r.status,
        "owner_id": r.owner_id, "owner_name": r.owner_name,
        "remark": r.remark,
        "created_at": r.created_at.isoformat() if r.created_at else "",
        "updated_at": r.updated_at.isoformat() if r.updated_at else "",
    }


# --- Export ---

@router.get("/api/v1/service_tickets/export/excel")
async def export_tickets_excel(
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("service:view")),
):
    from app.config import settings
    items, _ = await service.list_tickets(db, tenant_id, page_size=settings.MAX_EXPORT_ROWS)
    headers = ["工单编号", "类型", "优先级", "状态", "描述", "处理结果", "负责人", "创建人", "创建时间"]
    rows = []
    for t in items:
        rows.append([
            t.ticket_no, t.type or "", t.priority or "", t.status or "",
            t.description or "", t.resolution or "",
            t.assigned_to_name or "", t.created_by_name or "",
            t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
        ])
    buf = build_excel("售后工单", headers, rows)
    return excel_response(buf, "service_tickets.xlsx")


# --- 关联订单下拉（售后专用轻量查询，issue #85）---
# 独立于订单模块：用 service 权限鉴权，售后角色即使没有 order:view 也能在新建工单时
# 按客户/关键字挑选要关联的订单（原先前端直接调 /api/v1/orders 需 order:view，导致缺权限报错）。

@router.get("/api/v1/service_tickets/order_options")
async def order_options_for_ticket(
    customer_id: str | None = Query(None),
    keyword: str | None = Query(None),
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("service:view")),
):
    items, _ = await order_service.list_orders(
        db, tenant_id, page_no=1, page_size=20,
        customer_id=customer_id or None, keyword=keyword or None)
    return ok({"items": [{"id": o.id, "order_no": o.order_no, "title": o.title} for o in items]})


# --- 导入（含模板下载，issue #85）---

@router.get("/api/v1/service_tickets/import/template")
async def download_ticket_import_template(
    _user=Depends(require_permissions("service:create")),
):
    """下载售后工单导入模板。"""
    sample = [["故障", "高", "示例客户A", "", "张三", "设备无法启动，请尽快处理"]]
    buf = build_template("售后工单导入模板", IMPORT_HEADERS, sample)
    return excel_response(buf, "service_ticket_import_template.xlsx")


def _read_import_rows(content: bytes, filename: str) -> list[tuple]:
    """解析上传的 .xlsx/.csv，返回含表头的行列表。"""
    if (filename or "").lower().endswith(".csv"):
        import csv as csv_mod
        return [tuple(r) for r in csv_mod.reader(content.decode("utf-8-sig").splitlines())]
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


@router.post("/api/v1/service_tickets/import/preview")
async def import_tickets_preview(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("service:create")),
):
    """解析导入文件并返回表头 + 行 + 校验错误（工单无唯一键，故不做重复检测）。"""
    all_rows = _read_import_rows(await file.read(), file.filename or "")
    if not all_rows:
        return ok({"headers": [], "rows": [], "duplicates": [], "errors": {}})
    headers = [str(c).strip() if c else f"列{i+1}" for i, c in enumerate(all_rows[0])]
    data_rows: list[list[str]] = []
    for row in all_rows[1:]:
        if not row or not any(row):
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        while len(cells) < len(headers):
            cells.append("")
        data_rows.append(cells[:len(headers)])
    errors: dict[int, str] = {}
    for i, row in enumerate(data_rows):
        if len(row) < 6 or not row[5].strip():
            errors[i] = "问题描述不能为空"
    return ok({"headers": headers, "rows": data_rows, "duplicates": [], "errors": errors})


@router.post("/api/v1/service_tickets/import/excel")
async def import_tickets_excel(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("service:create")),
):
    """导入售后工单。列：工单类型, 优先级, 关联客户, 关联订单号, 负责人, 问题描述。
    类型/优先级支持中文或英文 code；客户/订单/负责人按名称匹配，匹配不到则留空不报错。"""
    from app.domains.order.models import Order
    from app.domains.auth.models import User as AuthUser

    all_rows = _read_import_rows(await file.read(), file.filename or "")
    if len(all_rows) < 2:
        return ok({"created": 0, "skipped": 0, "errors": []})

    created = 0
    errors: list[str] = []
    for idx, row in enumerate(all_rows[1:], 2):
        if not row or not any(row):
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]

        def cell(i: int) -> str:
            return cells[i] if i < len(cells) else ""

        desc = cell(5)
        if not desc:
            errors.append(f"第{idx}行: 问题描述不能为空")
            continue
        try:
            type_code = _TYPE_BY_LABEL.get(cell(0)) or (cell(0) if cell(0) in TYPE_LABELS else "fault")
            prio_code = _PRIORITY_BY_LABEL.get(cell(1)) or (cell(1) if cell(1) in PRIORITY_LABELS else "medium")
            customer_id = None
            if cell(2):
                cust = (await db.execute(select(Customer).where(
                    Customer.tenant_id == tenant_id, Customer.name == cell(2),
                    Customer.is_deleted == False))).scalars().first()
                customer_id = cust.id if cust else None
            order_id = None
            if cell(3):
                order = (await db.execute(select(Order).where(
                    Order.tenant_id == tenant_id, Order.order_no == cell(3),
                    Order.is_deleted == False))).scalars().first()
                order_id = order.id if order else None
            assigned_to_id = assigned_to_name = None
            if cell(4):
                u = (await db.execute(select(AuthUser).where(
                    AuthUser.tenant_id == tenant_id,
                    (AuthUser.real_name == cell(4)) | (AuthUser.username == cell(4))))).scalars().first()
                if u:
                    assigned_to_id, assigned_to_name = u.id, (u.real_name or u.username)
            body = ServiceTicketCreate(
                type=type_code, priority=prio_code, description=desc,
                customer_id=customer_id, order_id=order_id,
                assigned_to_id=assigned_to_id, assigned_to_name=assigned_to_name,
            )
            await service.create_ticket(db, tenant_id, body, current_user)
            created += 1
        except Exception as e:
            errors.append(f"第{idx}行: {str(e)[:80]}")
    return ok({"created": created, "skipped": 0, "errors": errors})


# --- ServiceTicket ---

@router.get("/api/v1/service_tickets")
async def list_tickets(
    customer_id: str | None = Query(None), project_id: str | None = Query(None),
    keyword: str | None = Query(None), status: str | None = Query(None),
    priority: str | None = Query(None), type: str | None = Query(None),
    filter: str | None = Query(None, description="高级筛选 FilterDsl(JSON)"),
    sort_by: str | None = Query(None), sort_order: str | None = Query(None),
    pageNo: int = Query(1, ge=1), pageSize: int = Query(20, ge=1, le=100),
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("service:view")),
):
    items, total = await service.list_tickets(
        db, tenant_id, customer_id=customer_id, project_id=project_id,
        keyword=keyword, status=status, priority=priority, ticket_type=type,
        page=pageNo, page_size=pageSize,
        current_user=_user, adv_filter=filter, sort_by=sort_by, sort_order=sort_order,
    )
    from app.common.list_enrich import customer_names_map, order_names_map
    cust_names = await customer_names_map(db, tenant_id, [t.customer_id for t in items])
    order_names = await order_names_map(db, tenant_id, [t.order_id for t in items])
    rows = [{
        **_ticket_dict(t),
        "customer_name": cust_names.get(t.customer_id),
        "order_name": order_names.get(t.order_id),
    } for t in items]
    await strip_entity_dicts(db, tenant_id, "service_ticket", rows, _user.get("roles"))  # 字段级权限：读取剔除隐藏扩展字段
    return ok({"items": rows, "total": total, "pageNo": pageNo, "pageSize": pageSize})


@router.post("/api/v1/service_tickets")
async def create_ticket(
    body: ServiceTicketCreate,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("service:create")),
):
    t = await service.create_ticket(db, tenant_id, body, current_user)
    return ok(_ticket_dict(t))


@router.get("/api/v1/service_tickets/{ticket_id}")
async def get_ticket(
    ticket_id: str, tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db), _user=Depends(require_permissions("service:view")),
):
    t = await service.get_ticket(db, tenant_id, ticket_id)
    d = _ticket_dict(t)
    await strip_entity_dicts(db, tenant_id, "service_ticket", [d], _user.get("roles"))  # 字段级权限：读取剔除隐藏扩展字段
    return ok(d)


@router.put("/api/v1/service_tickets/{ticket_id}")
async def update_ticket(
    ticket_id: str, body: ServiceTicketUpdate,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("service:edit")),
):
    t = await service.update_ticket(db, tenant_id, ticket_id, body, current_user)
    return ok(_ticket_dict(t))


@router.post("/api/v1/service_tickets/{ticket_id}/submit")
async def submit_ticket(
    ticket_id: str,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("service:edit")),
):
    """提交售后审批（内勤发起）。按 service_ticket 审批策略自动建流
    （内勤发起→生产主任审批分配售后人员→售后人员完成填写工作内容）。"""
    t = await service.submit_for_approval(db, tenant_id, ticket_id, current_user)
    return ok(_ticket_dict(t))


@router.delete("/api/v1/service_tickets/{ticket_id}")
async def delete_ticket(
    ticket_id: str,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("service:delete")),
):
    await service.delete_ticket(db, tenant_id, ticket_id, current_user)
    return ok(None)


# --- Satisfaction Rating ---

@router.post("/api/v1/service_tickets/{ticket_id}/rate")
async def rate_ticket(
    ticket_id: str,
    body: dict,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("service:view")),
):
    """Rate a resolved/closed ticket with 1-5 stars and optional comment."""
    from datetime import datetime, timezone
    from app.domains.service_ticket.models import ServiceTicket

    score = body.get("score")
    if not score or score < 1 or score > 5:
        from app.common.exceptions import BusinessException
        raise BusinessException(message="评分必须为1-5")

    t = (await db.execute(
        select(ServiceTicket).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.id == ticket_id,
        )
    )).scalar_one_or_none()
    if not t:
        from app.common.exceptions import BusinessException
        raise BusinessException(message="工单不存在")
    if t.status not in ("resolved", "closed"):
        from app.common.exceptions import BusinessException
        raise BusinessException(message="仅已解决或已关闭的工单可评价")

    t.satisfaction_score = score
    t.satisfaction_comment = body.get("comment", "")
    t.satisfaction_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(t)
    return ok(_ticket_dict(t))


# --- RenewalOpportunity ---

@router.get("/api/v1/renewal_opportunities")
async def list_renewals(
    customer_id: str | None = Query(None),
    status: str | None = Query(None),
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("service:view")),
):
    items = await service.list_renewals(db, tenant_id, customer_id=customer_id, status=status)
    # Batch lookup customer names
    cust_ids = list({r.customer_id for r in items if r.customer_id})
    cust_names: dict[str, str] = {}
    if cust_ids:
        rows = (await db.execute(
            select(Customer.id, Customer.name).where(Customer.id.in_(cust_ids), Customer.tenant_id == tenant_id)
        )).all()
        cust_names = {r.id: r.name for r in rows}
    result = []
    for r in items:
        d = _renewal_dict(r)
        d["customer_name"] = cust_names.get(r.customer_id, "")
        result.append(d)
    return ok(result)


@router.post("/api/v1/renewal_opportunities")
async def create_renewal(
    body: RenewalCreate,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("service:create")),
):
    r = await service.create_renewal(db, tenant_id, body, current_user)
    return ok(_renewal_dict(r))


@router.get("/api/v1/renewal_opportunities/{renewal_id}")
async def get_renewal(
    renewal_id: str, tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db), _user=Depends(require_permissions("service:view")),
):
    r = await service.get_renewal(db, tenant_id, renewal_id)
    return ok(_renewal_dict(r))


@router.put("/api/v1/renewal_opportunities/{renewal_id}")
async def update_renewal(
    renewal_id: str, body: RenewalUpdate,
    tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("service:edit")),
):
    r = await service.update_renewal(db, tenant_id, renewal_id, body, current_user)
    return ok(_renewal_dict(r))


# --- SLA Management ---

# Default SLA hours by priority
DEFAULT_SLA_HOURS = {
    "critical": 4,
    "high": 8,
    "medium": 24,
    "low": 72,
}


@router.get("/api/v1/service_tickets/sla/stats")
async def sla_stats(
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("service:view")),
):
    """SLA statistics: on-time rate, average resolution time, breach count."""
    from sqlalchemy import func, case
    from datetime import datetime, timezone, timedelta
    from app.domains.service_ticket.models import ServiceTicket

    # Open/active tickets with SLA status
    open_tickets = (await db.execute(
        select(func.count(ServiceTicket.id)).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.status.in_(["open", "assigned", "in_progress"]),
        )
    )).scalar() or 0

    resolved_tickets = (await db.execute(
        select(func.count(ServiceTicket.id)).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.status.in_(["resolved", "closed"]),
        )
    )).scalar() or 0

    # Calculate SLA breaches for open tickets
    now = datetime.now(timezone.utc)
    breach_count = 0
    near_breach_count = 0

    open_items = (await db.execute(
        select(ServiceTicket).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.status.in_(["open", "assigned", "in_progress"]),
        )
    )).scalars().all()

    for t in open_items:
        sla_hours = DEFAULT_SLA_HOURS.get(t.priority, 24)
        if t.created_at:
            created = t.created_at.replace(tzinfo=timezone.utc) if t.created_at.tzinfo is None else t.created_at
            deadline = created + timedelta(hours=sla_hours)
            if now > deadline:
                breach_count += 1
            elif now > deadline - timedelta(hours=max(1, sla_hours * 0.2)):
                near_breach_count += 1

    # By priority distribution
    priority_rows = (await db.execute(
        select(ServiceTicket.priority, func.count(ServiceTicket.id).label("count")).where(
            ServiceTicket.tenant_id == tenant_id,
            ServiceTicket.status.in_(["open", "assigned", "in_progress"]),
        ).group_by(ServiceTicket.priority)
    )).all()

    by_priority = {r.priority: r.count for r in priority_rows}

    total = open_tickets + resolved_tickets
    on_time_rate = round((total - breach_count) / total * 100, 1) if total > 0 else 100

    return ok({
        "open_tickets": open_tickets,
        "resolved_tickets": resolved_tickets,
        "breach_count": breach_count,
        "near_breach_count": near_breach_count,
        "on_time_rate": on_time_rate,
        "sla_config": DEFAULT_SLA_HOURS,
        "by_priority": by_priority,
    })


@router.get("/api/v1/service_tickets/knowledge")
async def knowledge_search(
    keyword: str = Query(..., min_length=2),
    ticket_type: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("service:view")),
):
    """Search resolved tickets as knowledge base for similar solutions."""
    from sqlalchemy import or_
    from app.domains.service_ticket.models import ServiceTicket

    q = select(ServiceTicket).where(
        ServiceTicket.tenant_id == tenant_id,
        ServiceTicket.status.in_(["resolved", "closed"]),
        ServiceTicket.resolution != None,
        or_(
            ServiceTicket.description.ilike(f"%{keyword}%"),
            ServiceTicket.resolution.ilike(f"%{keyword}%"),
            ServiceTicket.ticket_no.ilike(f"%{keyword}%"),
        ),
    )
    if ticket_type:
        q = q.where(ServiceTicket.type == ticket_type)
    q = q.order_by(ServiceTicket.updated_at.desc()).limit(20)

    items = (await db.execute(q)).scalars().all()
    return ok([{
        "id": t.id,
        "ticket_no": t.ticket_no,
        "type": t.type,
        "priority": t.priority,
        "description": (t.description or "")[:200],
        "resolution": (t.resolution or "")[:500],
        "updated_at": t.updated_at.isoformat() if t.updated_at else "",
    } for t in items])
