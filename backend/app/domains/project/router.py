from fastapi import APIRouter, Depends, Query, UploadFile, File
from sqlalchemy import select as sa_select, func, case
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
import io

from app.dependencies import get_db, get_tenant_id, require_permissions, get_data_scope
from app.common.schemas import ok
from app.common.export import build_excel, excel_response
from app.domains.project import service
from app.domains.project.schemas import (
    ProjectCreate, ProjectUpdate, ProjectTransfer, StageAdvance, StageRollback,
    ProjectMemberAdd, ProjectMemberUpdate,
)

router = APIRouter(prefix="/api/v1/projects", tags=["商机项目"])


def _project_dict(p, customer_name: str | None = None, delivery: tuple | None = None) -> dict:
    return {
        "id": p.id, "project_code": p.project_code,
        "customer_id": p.customer_id, "customer_name": customer_name, "name": p.name,
        "stage_code": p.stage_code,
        "amount_expect": float(p.amount_expect) if p.amount_expect is not None else None,
        "probability": p.probability,
        "close_date_expect": str(p.close_date_expect) if p.close_date_expect else None,
        "competitors_json": p.competitors_json,
        "key_requirements_json": p.key_requirements_json,
        "risk_level": p.risk_level,
        "has_guarantee": p.has_guarantee,
        "has_weight_requirement": p.has_weight_requirement,
        "uses_idle_equipment": p.uses_idle_equipment,
        "payment_method": p.payment_method,
        "owner_id": p.owner_id, "owner_name": p.owner_name,
        "created_by_id": p.created_by_id, "created_by_name": p.created_by_name,
        "status": p.status, "remark": p.remark, "custom_fields_json": p.custom_fields_json,
        # 交付进度(里程碑 已完成/总数)，让列表能一眼区分"已交付/交付中/未开始"(issue #65)
        "delivery_total": delivery[0] if delivery else None,
        "delivery_done": delivery[1] if delivery else None,
        "created_at": p.created_at.isoformat() if p.created_at else "",
        "updated_at": p.updated_at.isoformat() if p.updated_at else "",
    }


async def _delivery_progress(db, tenant_id: str, project_ids: list[str]) -> dict:
    """商机id -> (里程碑总数, 已完成数)。用于列表展示交付进度(issue #65)。"""
    if not project_ids:
        return {}
    from app.domains.delivery.models import DeliveryMilestone
    rows = (await db.execute(
        sa_select(
            DeliveryMilestone.project_id,
            func.count().label("total"),
            func.sum(case((DeliveryMilestone.status == "done", 1), else_=0)).label("done"),
        ).where(
            DeliveryMilestone.tenant_id == tenant_id,
            DeliveryMilestone.project_id.in_(project_ids),
        ).group_by(DeliveryMilestone.project_id)
    )).all()
    return {pid: (int(total or 0), int(done or 0)) for pid, total, done in rows}


async def _customer_names(db, tenant_id: str, items) -> dict:
    """批量取 customer_id -> name 映射，供列表/详情回填客户名。

    商机表只存 customer_id，客户名按 id 实时关联，避免前端靠"拉前 N 个客户"
    猜映射（客户总数 > 该页大小或受数据范围限制时会查不到名字）。"""
    from app.domains.customer.models import Customer
    ids = {p.customer_id for p in items if p.customer_id}
    if not ids:
        return {}
    rows = (await db.execute(
        sa_select(Customer.id, Customer.name).where(
            Customer.tenant_id == tenant_id, Customer.id.in_(ids)
        )
    )).all()
    return {cid: name for cid, name in rows}


def _history_dict(h) -> dict:
    return {
        "id": h.id, "project_id": h.project_id,
        "from_stage": h.from_stage, "to_stage": h.to_stage,
        "changed_by_id": h.changed_by_id, "changed_by_name": h.changed_by_name,
        "note": h.note,
        "created_at": h.created_at.isoformat() if h.created_at else "",
    }


@router.get("")
async def list_projects(
    pageNo: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    keyword: str = Query(None),
    stage_code: str = Query(None),
    customer_id: str = Query(None),
    status: str = Query(None),
    owner_id: str = Query(None),
    filter: str = Query(None, description="高级筛选 FilterDsl(JSON)"),
    sort_by: str = Query(None),
    sort_order: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("project:view")),
    data_scope: str | None = Depends(get_data_scope),
):
    # data_scope is None for admins (data:view_all) -> skip scope so they see everything.
    # For non-admins, pass current_user so the data scope (owner/member/share) applies, and
    # any explicit owner_id query param still narrows further.
    scope_user = None if data_scope is None else _user
    items, total = await service.list_projects(
        db, tenant_id, pageNo, pageSize, keyword, stage_code, customer_id, status,
        owner_id=owner_id, current_user=scope_user,
        adv_filter=filter, sort_by=sort_by, sort_order=sort_order,
        filter_user_id=_user.get("sub"),
    )
    name_map = await _customer_names(db, tenant_id, items)
    prog = await _delivery_progress(db, tenant_id, [p.id for p in items])
    return ok({"items": [_project_dict(p, name_map.get(p.customer_id), prog.get(p.id)) for p in items], "total": total, "pageNo": pageNo, "pageSize": pageSize})


@router.get("/export/excel")
async def export_projects_excel(
    keyword: str = Query(None),
    stage_code: str = Query(None),
    customer_id: str = Query(None),
    status: str = Query(None),
    owner_id: str = Query(None),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("project:view")),
):
    from app.config import settings
    items, _ = await service.list_projects(db, tenant_id, 1, settings.MAX_EXPORT_ROWS, keyword, stage_code, customer_id, status, owner_id)
    name_map = await _customer_names(db, tenant_id, items)
    headers = ["项目编码", "项目名称", "客户", "阶段", "预计金额", "概率(%)", "预计关闭日", "风险等级", "负责人", "状态", "创建时间"]
    rows = []
    for p in items:
        rows.append([
            p.project_code, p.name or "", name_map.get(p.customer_id) or "", p.stage_code or "",
            float(p.amount_expect) if p.amount_expect is not None else "",
            p.probability or "", str(p.close_date_expect) if p.close_date_expect else "",
            p.risk_level or "", p.owner_name or "", p.status or "",
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
        ])
    buf = build_excel("商机项目", headers, rows)
    return excel_response(buf, "projects.xlsx")


@router.post("/import/preview")
async def import_preview(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("project:create")),
):
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return ok({"headers": [], "rows": [], "duplicates": []})

    headers = [str(c).strip() if c else f"列{i+1}" for i, c in enumerate(all_rows[0])]
    data_rows = []
    names = []
    for row in all_rows[1:]:
        if not row or not any(row):
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        while len(cells) < len(headers):
            cells.append("")
        data_rows.append(cells[:len(headers)])
        if cells:
            names.append(cells[0])

    from app.domains.project.models import OpportunityProject
    existing = set()
    if names:
        result = await db.execute(
            sa_select(OpportunityProject.name).where(
                OpportunityProject.tenant_id == tenant_id,
                OpportunityProject.name.in_(names),
            )
        )
        existing = {r[0] for r in result.all()}

    duplicates = [i for i, name in enumerate(names) if name in existing]
    return ok({"headers": headers, "rows": data_rows, "duplicates": duplicates})


@router.post("/import/excel")
async def import_projects_excel(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:create")),
):
    """Import projects. Columns: 项目名称, 预计金额, 概率(%), 预计关闭日, 风险等级, 备注"""
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(all_rows) < 2:
        return ok({"created": 0, "skipped": 0, "errors": []})

    data_rows = all_rows[1:]

    from app.domains.project.models import OpportunityProject
    names = [str(r[0]).strip() for r in data_rows if r and r[0]]
    existing_names = set()
    if names:
        result = await db.execute(
            sa_select(OpportunityProject.name).where(
                OpportunityProject.tenant_id == tenant_id,
                OpportunityProject.name.in_(names),
            )
        )
        existing_names = {r[0] for r in result.all()}

    created = 0
    skipped = 0
    errors = []
    for idx, row in enumerate(data_rows, 2):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if name in existing_names:
            skipped += 1
            continue
        try:
            amount = None
            if len(row) > 1 and row[1]:
                try:
                    amount = float(row[1])
                except (ValueError, TypeError):
                    pass
            prob = None
            if len(row) > 2 and row[2]:
                try:
                    prob = int(float(row[2]))
                except (ValueError, TypeError):
                    pass
            data = ProjectCreate(
                name=name,
                amount_expect=amount,
                probability=prob,
                close_date_expect=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                risk_level=str(row[4]).strip() if len(row) > 4 and row[4] else None,
                remark=str(row[5]).strip() if len(row) > 5 and row[5] else None,
            )
            await service.create_project(db, tenant_id, data, current_user)
            created += 1
            existing_names.add(name)
        except Exception as e:
            errors.append(f"第{idx}行: {str(e)[:80]}")
    return ok({"created": created, "skipped": skipped, "errors": errors})


@router.post("")
async def create_project(
    body: ProjectCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:create")),
):
    p = await service.create_project(db, tenant_id, body, current_user)
    return ok(_project_dict(p))


@router.get("/{project_id}")
async def get_project(
    project_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("project:view")),
):
    p = await service.get_project(db, tenant_id, project_id)
    name_map = await _customer_names(db, tenant_id, [p])
    return ok(_project_dict(p, name_map.get(p.customer_id)))


@router.put("/{project_id}")
async def update_project(
    project_id: str,
    body: ProjectUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:edit")),
):
    p = await service.update_project(db, tenant_id, project_id, body, current_user)
    return ok(_project_dict(p))


@router.post("/{project_id}/transfer")
async def transfer_project_owner(
    project_id: str,
    body: ProjectTransfer,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:transfer")),
):
    """转移商机负责人。需 project:transfer 权限（默认仅主管/管理员持有），
    与普通编辑(project:edit)分离，避免任意成员私自改派以逃避数据范围监控或抢单。"""
    p = await service.transfer_owner(db, tenant_id, project_id, body.owner_id, body.note, current_user)
    return ok(_project_dict(p))


@router.delete("/{project_id}")
async def delete_project(
    project_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:delete")),
):
    await service.delete_project(db, tenant_id, project_id, current_user)
    return ok()


@router.post("/{project_id}/advance")
async def advance_stage(
    project_id: str,
    body: StageAdvance,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:advance")),
):
    p = await service.advance_stage(db, tenant_id, project_id, body.to_stage, body.note, current_user, force=body.force or False)
    return ok(_project_dict(p))


@router.post("/{project_id}/rollback")
async def rollback_stage(
    project_id: str,
    body: StageRollback,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:advance")),
):
    p = await service.rollback_stage(db, tenant_id, project_id, body.to_stage, body.note, current_user)
    return ok(_project_dict(p))


@router.get("/{project_id}/gate_check")
async def gate_check(
    project_id: str,
    to_stage: str = Query(...),
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("project:view")),
):
    project = await service.get_project(db, tenant_id, project_id)
    failed = await service.check_gate_rules(db, tenant_id, project, to_stage)
    return ok({"pass": len(failed) == 0, "failed_rules": failed})


@router.get("/{project_id}/health")
async def project_health(
    project_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("project:view")),
):
    result = await service.calculate_health_score(db, tenant_id, project_id)
    return ok(result)


@router.get("/{project_id}/stage_history")
async def list_stage_history(
    project_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("project:view")),
):
    items = await service.list_stage_history(db, tenant_id, project_id)
    return ok([_history_dict(h) for h in items])


# ---- ACL Shares (project-level) ----
@router.get("/{project_id}/shares")
async def list_project_shares(
    project_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("project:view")),
):
    from app.domains.customer.service import list_shares
    items = await list_shares(db, tenant_id, "project", project_id)
    return ok([{
        "id": s.id, "biz_type": s.biz_type, "biz_id": s.biz_id,
        "shared_to_type": s.shared_to_type, "shared_to_id": s.shared_to_id,
        "shared_to_name": s.shared_to_name, "permission": s.permission,
        "shared_by_name": s.shared_by_name,
        "created_at": s.created_at.isoformat() if s.created_at else "",
    } for s in items])


@router.post("/{project_id}/shares")
async def create_project_share(
    project_id: str,
    body: dict,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:edit")),
):
    from app.domains.customer.service import create_share
    body["biz_type"] = "project"
    body["biz_id"] = project_id
    s = await create_share(db, tenant_id, body, current_user)
    return ok({"id": s.id, "shared_to_name": s.shared_to_name})


@router.delete("/{project_id}/shares/{share_id}")
async def delete_project_share(
    project_id: str,
    share_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:edit")),
):
    from app.domains.customer.service import delete_share
    await delete_share(db, tenant_id, share_id, current_user)
    return ok()


# ---- Project Members (多部门 / 多人协作) ----
def _member_dict(m) -> dict:
    return {
        "id": m.id, "project_id": m.project_id,
        "user_id": m.user_id, "user_name": m.user_name,
        "member_role": m.member_role,
        "department_id": m.department_id, "department_name": m.department_name,
        "permission": m.permission,
        "added_by_id": m.added_by_id, "added_by_name": m.added_by_name,
        "created_at": m.created_at.isoformat() if m.created_at else "",
    }


@router.get("/{project_id}/members")
async def list_project_members(
    project_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    _user=Depends(require_permissions("project:view")),
):
    items = await service.list_members(db, tenant_id, project_id)
    return ok([_member_dict(m) for m in items])


@router.post("/{project_id}/members")
async def add_project_member(
    project_id: str,
    body: ProjectMemberAdd,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:edit")),
):
    m = await service.add_member(db, tenant_id, project_id, body, current_user)
    return ok(_member_dict(m))


@router.put("/{project_id}/members/{member_id}")
async def update_project_member(
    project_id: str,
    member_id: str,
    body: ProjectMemberUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:edit")),
):
    m = await service.update_member(db, tenant_id, project_id, member_id, body, current_user)
    return ok(_member_dict(m))


@router.delete("/{project_id}/members/{member_id}")
async def remove_project_member(
    project_id: str,
    member_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(require_permissions("project:edit")),
):
    await service.remove_member(db, tenant_id, project_id, member_id, current_user)
    return ok()
