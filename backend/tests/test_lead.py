"""Lead API integration tests."""

from httpx import AsyncClient

from tests.lead_intel_helpers import approve_lead_intel_include


async def test_list_leads(client: AsyncClient, auth_headers: dict):
    resp = await client.get("/api/v1/leads", headers=auth_headers)
    data = resp.json()
    assert data["code"] == 0
    assert "items" in data["data"]


async def test_list_leads_reactivation_active(client: AsyncClient, auth_headers: dict):
    resp = await client.get(
        "/api/v1/leads",
        headers=auth_headers,
        params={"reactivation_active": True, "pageSize": 5},
    )
    data = resp.json()
    assert data["code"] == 0
    for item in data["data"]["items"]:
        assert item["reactivation_status"] in (
            "awaiting_reporter", "awaiting_filler", "pending_review",
        )


async def test_list_tasks_biz_type_filter(client: AsyncClient, auth_headers: dict):
    resp = await client.get(
        "/api/v1/tasks",
        headers=auth_headers,
        params={"biz_type": "lead_reactivation", "pageSize": 5},
    )
    assert resp.json()["code"] == 0


async def test_lead_crud(client: AsyncClient, auth_headers: dict):
    # as_draft：提交后会启审并可能卡在「业务员确认」待办，整单不可编辑
    resp = await client.post("/api/v1/leads", headers=auth_headers, json={
        "title": "测试线索_自动化", "company_name": "测试公司", "source": "website",
        "as_draft": True,
    })
    data = resp.json()
    assert data["code"] == 0
    lid = data["data"]["id"]
    assert data["data"]["lead_code"] is not None

    resp = await client.get(f"/api/v1/leads/{lid}", headers=auth_headers)
    assert resp.json()["code"] == 0

    resp = await client.put(f"/api/v1/leads/{lid}", headers=auth_headers, json={
        "title": "测试线索_已更新", "score": 80,
    })
    assert resp.json()["code"] == 0

    await client.delete(f"/api/v1/leads/{lid}", headers=auth_headers)


async def test_lead_custom_fields_survive_read_and_edit(client: AsyncClient, auth_headers: dict):
    """扩展字段值必须出现在出参里，且不被「不带该字段的更新」清空。

    回归：_lead_dict 曾漏掉 custom_fields_json，前端编辑表单读到空对象后原样回提，
    每次保存都会静默清空已存的扩展字段值。
    """
    h = auth_headers
    cf = {"f_industry_note": "选矿", "f_visit_count": 3}
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "扩展字段线索", "company_name": "扩展字段公司", "custom_fields_json": cf,
        "as_draft": True,
    })).json()["data"]["id"]

    # 详情与列表都应回传扩展字段值
    detail = (await client.get(f"/api/v1/leads/{lid}", headers=h)).json()["data"]
    assert detail["custom_fields_json"] == cf

    listed = (await client.get("/api/v1/leads", headers=h,
                               params={"keyword": "扩展字段线索"})).json()["data"]["items"]
    assert any(i["id"] == lid and i["custom_fields_json"] == cf for i in listed)

    # 不携带 custom_fields_json 的更新不应清空已存值
    await client.put(f"/api/v1/leads/{lid}", headers=h, json={"title": "扩展字段线索_改名"})
    after = (await client.get(f"/api/v1/leads/{lid}", headers=h)).json()["data"]
    assert after["custom_fields_json"] == cf

    await client.delete(f"/api/v1/leads/{lid}", headers=h)


async def test_lead_date_field_filter(client: AsyncClient, auth_headers: dict):
    """日期区间可切换按 biz_date 筛选（默认仍按 created_at）。"""
    h = auth_headers
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "业务日期线索", "company_name": "业务日期公司", "biz_date": "2020-03-15",
    })).json()["data"]["id"]

    def ids(items):
        return [i["id"] for i in items]

    hit = (await client.get("/api/v1/leads", headers=h, params={
        "date_field": "biz_date", "start_date": "2020-03-01", "end_date": "2020-03-31",
    })).json()["data"]["items"]
    assert lid in ids(hit)

    miss = (await client.get("/api/v1/leads", headers=h, params={
        "date_field": "biz_date", "start_date": "2020-04-01", "end_date": "2020-04-30",
    })).json()["data"]["items"]
    assert lid not in ids(miss)

    # 默认维度仍是 created_at：按 2020 年筛创建时间应筛不到今天刚建的这条
    default_dim = (await client.get("/api/v1/leads", headers=h, params={
        "start_date": "2020-03-01", "end_date": "2020-03-31",
    })).json()["data"]["items"]
    assert lid not in ids(default_dim)

    await client.delete(f"/api/v1/leads/{lid}", headers=h)


async def test_lead_export(client: AsyncClient, auth_headers: dict):
    resp = await client.get("/api/v1/leads/export/excel", headers=auth_headers)
    assert resp.status_code == 200


async def test_lead_export_maps_codes_to_chinese(client: AsyncClient, auth_headers: dict):
    """导出客户类型/行业/状态应为中文，不能裸出字典码。"""
    import io
    from openpyxl import load_workbook

    h = auth_headers
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "导出中文映射线索",
        "company_name": "映射测试公司",
        "source": "import",
        "customer_type": "terminal_private",
        "industry": "screening_chemical",
    })).json()["data"]["id"]

    resp = await client.get(
        "/api/v1/leads/export/excel", headers=h,
        params={"keyword": "导出中文映射线索"},
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    row = next(ws.iter_rows(min_row=2, max_row=2))
    by = {header[i]: row[i].value for i in range(len(header))}
    assert by["客户类型"] == "终端客户-一般民企"
    assert by["行业"] == "筛分分选-化工"
    assert by["状态"] == "新建"
    assert by["来源"] == "导入"
    assert by["客户类型"] != "terminal_private"
    assert by["行业"] != "screening_chemical"
    assert "客户类型（新/老）" in header
    assert "项目最终状态" in header
    # 新建未审：最终状态多为草稿/待审；新/老可能为空
    assert by["项目最终状态"] in ("草稿", "待审", "收录", "已驳回", "袭击", "")

    await client.delete(f"/api/v1/leads/{lid}", headers=h)


async def test_qualify_marks_lead_without_opportunity_when_disabled(
    client: AsyncClient, auth_headers: dict, db, lead_intel_user,
):
    h = auth_headers
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "仅标记转化线索", "company_name": "测试公司", "source": "website",
    })).json()["data"]["id"]
    await approve_lead_intel_include(db, lid, lead_intel_user)
    res = (await client.post(f"/api/v1/leads/{lid}/qualify", headers=h,
                             json={"create_opportunity": False})).json()
    assert res["code"] == 0
    assert res["data"].get("customer_id") is None
    assert res["data"].get("project_id") is None

    lead = (await client.get(f"/api/v1/leads/{lid}", headers=h)).json()["data"]
    assert lead["status"] == "qualified"

    upd = (await client.put(f"/api/v1/leads/{lid}", headers=h, json={"title": "改名"})).json()
    assert upd["code"] == 0
    assert upd["data"]["title"] == "改名"


async def test_qualify_creates_opportunity_by_default(
    client: AsyncClient, auth_headers: dict, db, lead_intel_user,
):
    h = auth_headers
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "默认转商机线索", "company_name": "默认商机公司", "source": "website",
    })).json()["data"]["id"]
    await approve_lead_intel_include(db, lid, lead_intel_user)
    res = (await client.post(f"/api/v1/leads/{lid}/qualify", headers=h)).json()
    assert res["code"] == 0
    assert res["data"].get("customer_id") is None
    pid = res["data"].get("project_id")
    assert pid, "默认应创建商机"
    assert res["data"].get("project_code")

    proj = (await client.get(f"/api/v1/projects/{pid}", headers=h)).json()["data"]
    assert proj["customer_id"] is None
    assert proj.get("lead_id") == lid

    await client.delete(f"/api/v1/projects/{pid}", headers=h)


async def test_qualify_with_create_opportunity_carries_context(
    client: AsyncClient, auth_headers: dict, db, lead_intel_user,
):
    h = auth_headers
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "大型振动筛采购", "company_name": "矿业集团",
        "source": "expo", "demand_summary": "需要 3 台大型直线振动筛，含保函",
        "budget_range": "200-300万",
    })).json()["data"]["id"]
    await approve_lead_intel_include(db, lid, lead_intel_user)

    res = (await client.post(f"/api/v1/leads/{lid}/qualify", headers=h,
                             json={"create_opportunity": True})).json()
    assert res["code"] == 0
    assert res["data"].get("customer_id") is None
    pid = res["data"].get("project_id")
    assert pid, "勾选后应创建商机"
    assert res["data"].get("project_code")

    lead = (await client.get(f"/api/v1/leads/{lid}", headers=h)).json()["data"]
    lead_code = lead["lead_code"]

    # 商机不自动建档客户，并带入需求摘要；编号走商机规则，同时保留来源线索号
    proj = (await client.get(f"/api/v1/projects/{pid}", headers=h)).json()["data"]
    assert proj["customer_id"] is None
    assert proj["stage_code"] == "S1"
    assert (proj.get("key_requirements_json") or {}).get("summary") == "需要 3 台大型直线振动筛，含保函"
    assert proj["project_code"] != f"PRJ{lead_code}"
    assert proj["project_code"].startswith("PRJ")
    assert "-" not in proj["project_code"]
    assert proj.get("lead_id") == lid
    assert proj.get("lead_code") == lead_code

    await client.delete(f"/api/v1/projects/{pid}", headers=h)
