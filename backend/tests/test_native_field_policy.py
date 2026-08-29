"""原生字段策略：租户可为内置字段配置必填/条件显隐/只读，后端强制。

通过「发布线索实体系统模板」写入原生字段覆盖项，再打线索 API 验证行为。
"""
import pytest
from httpx import AsyncClient

from app.domains.lowcode.native_field_catalog import merge_native_overrides, get_native_fields, merge_system_rules, get_system_rules


# ===== 目录合并（纯函数，无需 DB） =====

def test_contract_detail_tables_in_catalog():
    fields = get_native_fields("contract")
    by_id = {f["id"]: f for f in fields}
    assert by_id["line_items"]["type"] == "detail_table"
    assert by_id["line_items"]["entity_storage"] == "key_clauses_json"
    assert by_id["payment_terms"]["entity_storage"] == "payment_terms_json"

    line_ids = [c["id"] for c in by_id["line_items"]["detail_table_columns"]]
    pay_ids = [c["id"] for c in by_id["payment_terms"]["detail_table_columns"]]
    # 与简道云/旧 ContractTerms 字段 id 对齐（前端 FALLBACK_* 须同序同 id）
    assert line_ids == [
        "is_fx", "product_type", "name", "spec", "unit", "qty",
        "fx_price", "fx_rate", "price", "amount", "fx_amount",
        "elec_ctrl", "standard", "line_remark",
    ]
    assert pay_ids == ["due_date", "kind", "ratio", "amount", "remind", "note"]
    fx_price = next(c for c in by_id["line_items"]["detail_table_columns"] if c["id"] == "fx_price")
    assert fx_price["props"]["show_when"] == {"field": "is_fx", "equals": ["是"]}
    assert fx_price.get("required") is True
    assert fx_price["props"].get("system_column") is True
    assert "_widget_" in (fx_price["props"].get("aliases") or [""])[0]
    assert by_id["project_name"].get("required") is True
    assert by_id["peer_contract_no"].get("required") is True
    assert by_id["payment_desc"].get("required") is True
    assert by_id["application_field"].get("required") is True

    merged = merge_native_overrides("contract", [{
        "id": "line_items", "native": True,
        "detail_table_columns": [
            {"id": "name", "label": "品名改", "type": "text", "props": {}},
        ],
    }])
    li = next(f for f in merged if f["id"] == "line_items")
    assert li["detail_table_columns"][0]["label"] == "品名改"
    assert "detail_table_columns" in __import__(
        "app.domains.lowcode.native_field_catalog", fromlist=["OVERRIDABLE_KEYS"]
    ).OVERRIDABLE_KEYS


def test_customer_required_defaults_align_jdy():
    """对齐简道云：内贸/智能化分区字段目录默认必填；开关/行业/地址由表单硬必填（API 可简写建档）。"""
    by = {f["id"]: f for f in get_native_fields("customer")}
    assert by["name"].get("required") is True
    # 始终展示项不设目录 default_required，避免 pytest/openapi 只传 name 被拦
    for fid in ("is_smart_filing", "is_foreign_trade", "industry", "address", "owner_id"):
        assert by[fid].get("required") is not True, fid
    domestic = [
        "registered_capital", "paid_in_capital", "founded_year", "parent_company_note",
        "customer_nature", "customer_relation", "level", "primary_contact_title",
        "wage_insurance_status", "taxpayer_id", "is_company_customer",
    ]
    smart = ["legal_person", "headcount", "smart_industry_category"]
    for fid in domestic + smart:
        assert by[fid].get("required") is True, fid
    for fid in ("short_name", "country", "foreign_customer_type", "customer_email"):
        assert by[fid].get("required") is not True, fid
    rules = get_system_rules("customer")
    assert any(r["id"] == "__sys_customer_foreign_star_required" for r in rules)


def test_customer_system_rules_align_jdy_show_rules():
    """客户显隐对齐简道云：外贸/智能化开关控制内贸·开票 / 外贸 / 备案分区。"""
    from app.domains.lowcode.rule_engine import compute_field_states

    fields = get_native_fields("customer")
    rules = get_system_rules("customer")
    assert [r["id"] for r in rules] == [
        "__sys_customer_smart_when_yes",
        "__sys_customer_domestic_when_not_foreign",
        "__sys_customer_foreign_when_yes",
        "__sys_customer_foreign_star_required",
    ]
    by_id = {f["id"] for f in fields}
    assert {"is_foreign_trade", "is_smart_filing", "country", "taxpayer_id", "legal_person"} <= by_id

    domestic = compute_field_states(
        fields, {"is_foreign_trade": False, "is_smart_filing": False}, rules,
    )
    assert domestic["registered_capital"]["visible"] is True
    assert domestic["taxpayer_id"]["visible"] is True
    assert domestic["country"]["visible"] is False
    assert domestic["legal_person"]["visible"] is False
    assert domestic.get("region", {}).get("visible", True) is False
    assert domestic.get("website", {}).get("visible", True) is False

    foreign = compute_field_states(
        fields, {"is_foreign_trade": True, "is_smart_filing": False}, rules,
    )
    assert foreign["registered_capital"]["visible"] is False
    assert foreign["country"]["visible"] is True
    assert foreign["region"]["visible"] is True
    assert foreign["website"]["visible"] is True
    assert foreign["source"]["visible"] is True

    smart_cn = compute_field_states(
        fields, {"is_foreign_trade": "否", "is_smart_filing": "是"}, rules,
    )
    assert smart_cn["legal_person"]["visible"] is True
    assert smart_cn["taxpayer_id"]["visible"] is True
    assert smart_cn["short_name"]["visible"] is False

    # 必填：隐藏区清掉；外贸=是时简称/国别等条件必填
    assert domestic["registered_capital"]["required"] is True
    assert domestic["legal_person"]["required"] is False
    assert domestic["short_name"]["required"] is False
    assert foreign["registered_capital"]["required"] is False
    assert foreign["short_name"]["required"] is True
    assert foreign["country"]["required"] is True
    assert foreign["customer_email"]["required"] is True
    assert smart_cn["legal_person"]["required"] is True
    assert smart_cn["headcount"]["required"] is True


def test_merge_system_rules_override_and_disable():
    defaults = get_system_rules("contract")
    assert defaults, "合同应有系统规则"
    first_id = defaults[0]["id"]
    merged = merge_system_rules("contract", [
        {
            "id": first_id,
            "type": "visibility",
            "target_field_id": "change_reason",
            "condition": {"field": "change_type", "operator": "eq", "value": "custom"},
            "action": {"visible": True},
            "enabled": False,
        },
        {"id": "tenant_extra", "type": "visibility", "target_field_id": "x",
         "condition": {}, "action": {"visible": True}},
    ])
    overridden = next(r for r in merged if r["id"] == first_id)
    assert overridden["enabled"] is False
    assert overridden["condition"]["value"] == "custom"
    assert any(r["id"] == "tenant_extra" for r in merged)
    # 未覆盖的系统规则仍在，且排在租户规则前
    assert merged[-1]["id"] == "tenant_extra"
    assert len([r for r in merged if str(r["id"]).startswith("__sys_")]) == len(defaults)


def test_override_can_make_optional_field_required():
    merged = merge_native_overrides("lead", [
        {"id": "industry", "native": True, "required": True},
    ])
    industry = next(f for f in merged if f["id"] == "industry")
    assert industry["required"] is True
    assert industry["type"] == "select", "type 必须来自目录，不受覆盖项影响"


def test_system_required_cannot_be_downgraded():
    """title 是 NOT NULL 列，租户不得把它改成非必填。"""
    merged = merge_native_overrides("lead", [
        {"id": "title", "native": True, "required": False},
    ])
    title = next(f for f in merged if f["id"] == "title")
    assert title["required"] is True
    assert title["system_required"] is True


def test_label_override_only_set_when_tenant_actually_renamed():
    """未改标签时不得透出 label_override，否则业务表单会被目录默认名覆盖既有文案。"""
    untouched = merge_native_overrides("lead", [])
    assert all("label_override" not in f for f in untouched)

    # 只改了必填、没动标签 → 仍不应有 label_override
    partial = merge_native_overrides("lead", [
        {"id": "biz_date", "native": True, "required": True},
    ])
    biz = next(f for f in partial if f["id"] == "biz_date")
    assert "label_override" not in biz
    assert biz["label"] == "业务日期"  # 目录默认名仍在，供设计器展示

    renamed = merge_native_overrides("lead", [
        {"id": "biz_date", "native": True, "label": "跟进日期"},
    ])
    assert next(f for f in renamed if f["id"] == "biz_date")["label_override"] == "跟进日期"


async def test_non_form_editable_field_can_be_masked_but_not_required(
    client: AsyncClient, auth_headers: dict,
):
    """form_editable=False 的字段（如合同签约日期，由签署流程写入）：
    仍可配隐藏/脱敏，但配了必填不得阻断保存 —— 用户根本没有填它的入口。
    """
    h = auth_headers
    cust_id = (await client.post("/api/v1/customers", json={"name": "签约日期测试客户"},
                                 headers=h)).json()["data"]["id"]
    proj_id = (await client.post("/api/v1/projects", json={
        "name": "签约日期测试商机", "customer_id": cust_id, "stage_code": "S1",
    }, headers=h)).json()["data"]["id"]

    tpl = (await client.get("/api/v1/lc/entity-templates/contract", headers=h)).json()["data"]

    async def publish(defs):
        await client.post(f"/api/v1/lc/form-templates/{tpl['id']}/design", headers=h, json={
            "field_definitions": defs, "layout_definition": {}, "rule_definitions": []})
        await client.post(f"/api/v1/lc/form-templates/{tpl['id']}/publish", headers=h)

    try:
        await publish([{"id": "signed_date", "native": True, "label": "签订日期",
                        "type": "date", "required": True}])
        r = (await client.post(f"/api/v1/projects/{proj_id}/contracts",
                               json={"contract_no": f"CT-POLICY-{proj_id[:8].upper()}", "amount_total": 100}, headers=h)).json()
        assert r["code"] == 0, f"表单上填不了的字段不该拦住保存: {r}"
        await client.delete(f"/api/v1/contracts/{r['data']['contract']['id']}", headers=h)
    finally:
        await publish([])


async def test_edit_only_field_does_not_block_creation(
    client: AsyncClient, auth_headers: dict,
):
    """available_on_create=False 的字段（如工单「解决方案」）配了必填也不得挡住新建 ——
    新建工单时还谈不上解决方案，界面上本就没有该输入项。
    """
    h = auth_headers
    tpl = (await client.get("/api/v1/lc/entity-templates/service_ticket", headers=h)).json()["data"]

    async def publish(defs):
        await client.post(f"/api/v1/lc/form-templates/{tpl['id']}/design", headers=h, json={
            "field_definitions": defs, "layout_definition": {}, "rule_definitions": []})
        await client.post(f"/api/v1/lc/form-templates/{tpl['id']}/publish", headers=h)

    try:
        await publish([{"id": "resolution", "native": True, "label": "解决方案",
                        "type": "textarea", "required": True}])
        r = (await client.post("/api/v1/service_tickets", headers=h, json={
            "type": "fault", "description": "设备无法启动",
        })).json()
        assert r["code"] == 0, f"新建工单不该被「解决方案」必填拦住: {r}"

        # 但编辑时该字段可见可填，必填照常生效
        tid = r["data"]["id"]
        upd = (await client.put(f"/api/v1/service_tickets/{tid}", headers=h,
                                json={"resolution": ""})).json()
        assert upd["code"] != 0 and "解决方案" in upd["message"]
        await client.delete(f"/api/v1/service_tickets/{tid}", headers=h)
    finally:
        await publish([])


def test_override_cannot_change_id_or_type():
    merged = merge_native_overrides("lead", [
        {"id": "industry", "native": True, "type": "detail_table", "label": "所属行业"},
    ])
    industry = next(f for f in merged if f["id"] == "industry")
    assert industry["type"] == "select", "改类型会写坏与业务列的映射，必须被忽略"
    assert industry["label"] == "所属行业", "label 属于可覆盖白名单"


def test_stale_override_for_removed_field_is_ignored():
    merged = merge_native_overrides("lead", [
        {"id": "__no_such_field__", "native": True, "required": True},
    ])
    assert all(f["id"] != "__no_such_field__" for f in merged)


def test_all_field_permission_keys_are_overridable():
    """字段级权限的三个键都必须在覆盖白名单里。

    回归：unmask_roles 曾漏加，导致租户在设计器里配了脱敏、合并时被静默丢弃 ——
    界面上看着配好了，实际一点也不生效。
    """
    from app.domains.lowcode.native_field_catalog import OVERRIDABLE_KEYS
    for key in ("visible_roles", "unmask_roles", "edit_roles", "download_roles"):
        assert key in OVERRIDABLE_KEYS, f"{key} 未列入 OVERRIDABLE_KEYS，租户配置会被丢弃"


def test_props_override_limited_to_hidden_and_readonly():
    merged = merge_native_overrides("lead", [
        {"id": "remark", "native": True, "props": {"readonly": True, "evil": "x"}},
    ])
    remark = next(f for f in merged if f["id"] == "remark")
    assert remark["props"]["readonly"] is True
    assert "evil" not in remark["props"]


def test_catalog_fields_all_have_a_form_control():
    """已接表单的实体（FORM_WIRED），其目录字段都必须有对应的 PolicyItem。

    否则会出现「后端按策略拦下、界面上却找不到该字段可填」的死角 —— 省市区就是因此被
    刻意排除的（它们由 RegionCascader 整体选择，只有隐藏桩）。
    未接表单的实体只服务读取路径，不受此约束。
    """
    import re
    from pathlib import Path
    from app.domains.lowcode.native_field_catalog import FORM_WIRED

    # 一个实体可能有多处表单，且新建与编辑能填的字段并不相同：
    # (相对路径, 该表单覆盖的场景)  场景 ∈ {"create", "edit", "both"}
    forms = {
        "lead": [
            ("frontend/src/pages/lead/LeadForm.tsx", "both"),
            ("frontend/src/pages/mobile/MobileLeadForm.tsx", "create"),
        ],
        "customer": [
            ("frontend/src/pages/customer/CustomerForm.tsx", "both"),
            ("frontend/src/pages/mobile/MobileCustomerForm.tsx", "create"),
        ],
        "contact": [
            ("frontend/src/pages/customer/CustomerDetail.tsx", "both"),
            ("frontend/src/pages/customer/ContactList.tsx", "create"),
        ],
        "project": [
            ("frontend/src/pages/opportunity/OpportunityForm.tsx", "both"),
            ("frontend/src/pages/mobile/MobileOpportunityForm.tsx", "create"),
        ],
        "contract": [("frontend/src/components/ContractRegistrationFields.tsx", "create")],
        "order": [("frontend/src/pages/order/OrderList.tsx", "both")],
        "service_ticket": [
            ("frontend/src/pages/service/ServiceTicketList.tsx", "create"),
            ("frontend/src/pages/service/ServiceTicketDetail.tsx", "edit"),
        ],
        "payment": [("frontend/src/pages/payment/PaymentPage.tsx", "create")],
        "solution": [("frontend/src/pages/solution/SolutionList.tsx", "create")],
    }
    root = Path(__file__).resolve().parents[2]

    def _policy_item_names(src: str) -> set[str]:
        # <PolicyItem name="foo"> 或属性顺序不限
        return set(re.findall(r"<(?:PolicyItem|MField)\b[^>]*\bname=\"([^\"]+)\"", src))

    def _contract_registration_keys(source: str | None = None) -> set[str]:
        """从 contractRegistration 常量收集 key；source 为 'native'|'reg'|None(全部)。"""
        csrc = (root / "frontend/src/constants/contractRegistration.ts").read_text(encoding="utf-8")
        keys: set[str] = set()
        for m in re.finditer(r"key:\s*'([^']+)'", csrc):
            start = m.end()
            nxt = re.search(r"\bkey:\s*'", csrc[start:start + 800])
            window = csrc[start:start + (nxt.start() if nxt else 800)]
            if source is None:
                keys.add(m.group(1))
            elif re.search(rf"source:\s*['\"]{source}['\"]", window):
                keys.add(m.group(1))
        return keys

    for entity in FORM_WIRED:
        assert entity in forms, f"{entity} 已声明接入表单，但测试不知道它的表单文件在哪"
        for rel, scope in forms[entity]:
            src = (root / rel).read_text(encoding="utf-8")
            # MField 是移动端样式的 PolicyItem 别名，两者都算「该字段可填」
            rendered = _policy_item_names(src)
            if rel.endswith("ContractRegistrationFields.tsx"):
                # native 与 json_storage(reg) 均走 PolicyItem name={f.key} / 嵌套路径
                rendered |= _contract_registration_keys("native")
                rendered |= _contract_registration_keys("reg")
            missing = []
            for fd in get_native_fields(entity):
                if not fd.get("form_editable", True):
                    continue  # 由系统/专用流程写入，表单上本就没有输入项
                if scope == "create" and not fd.get("available_on_create", True):
                    continue  # 只在记录建立后才出现的字段（如工单解决方案）
                # 明细子表由业务插槽（ContractTerms）渲染，不是 PolicyItem
                if fd.get("type") == "detail_table" or fd.get("entity_storage"):
                    continue
                if fd["id"] not in rendered:
                    missing.append(fd["id"])
            assert not missing, f"{entity} @ {rel}({scope}): 目录里有但表单没有对应 PolicyItem: {missing}"


# 目录里的实体 -> (模型模块, 类名)，用于校验字段 id 都是真实列
_ENTITY_MODELS = {
    "lead": ("app.domains.lead.models", "Lead"),
    "customer": ("app.domains.customer.models", "Customer"),
    "contact": ("app.domains.customer.models", "Contact"),
    "project": ("app.domains.project.models", "OpportunityProject"),
    "contract": ("app.domains.contract.models", "Contract"),
    "contract_review": ("app.domains.contract_review.models", "ContractReview"),
    "quote": ("app.domains.quote.models", "Quote"),
    "order": ("app.domains.order.models", "Order"),
    "service_ticket": ("app.domains.service_ticket.models", "ServiceTicket"),
    "payment": ("app.domains.payment.models", "PaymentRecord"),
    "solution": ("app.domains.solution.models", "Solution"),
}


@pytest.mark.parametrize("entity", sorted(_ENTITY_MODELS))
def test_catalog_ids_are_real_columns(entity):
    """目录里的 id 必须是该实体表上真实存在的列（或声明为 JSON 列内键）。

    回归：被删掉的旧 field_rules UI 里 18 个字段有 12 个是不存在的列（customer.phone、
    contract.amount、quote.margin_rate 等），配了规则也永远匹配不到 —— 新目录不能重蹈覆辙。
    json_storage 字段落在 JSON 列（如 registration_json）内，id 是对象 key 而非表列。
    """
    import importlib
    from app.domains.lowcode.native_field_catalog import CATALOG

    mod, cls = _ENTITY_MODELS[entity]
    model = getattr(importlib.import_module(mod), cls)
    columns = set(model.__table__.columns.keys())
    for fd in get_native_fields(entity):
        storage = fd.get("json_storage")
        if storage:
            assert storage in columns, (
                f"{entity}.{fd['id']} 声明 json_storage={storage!r}，"
                f"但 {model.__tablename__} 无此列"
            )
            continue
        entity_storage = fd.get("entity_storage")
        if entity_storage:
            # 整字段 JSON 列：可能在主表，也可能在版本表（如合同明细 key_clauses_json）
            if entity_storage in columns:
                continue
            if entity == "contract" and entity_storage == "key_clauses_json":
                from app.domains.contract.models import ContractVersion
                ver_cols = set(ContractVersion.__table__.columns.keys())
                assert entity_storage in ver_cols, (
                    f"contract.{fd['id']} entity_storage={entity_storage!r} "
                    f"不在 Contract 也不在 ContractVersion"
                )
                continue
            assert False, (
                f"{entity}.{fd['id']} 声明 entity_storage={entity_storage!r}，"
                f"但 {model.__tablename__} 无此列"
            )
        assert fd["id"] in columns, f"{entity}.{fd['id']} 不是 {model.__tablename__} 表的列"
    assert entity in CATALOG, f"{entity} 应在 CATALOG 中"


def test_contract_drawing_no_visible_on_create():
    """合同登记新建页应展示只读图纸编号（预生成预览）。"""
    from app.domains.lowcode.native_field_catalog import CATALOG
    fd = next(f for f in CATALOG["contract"] if f["id"] == "drawing_no")
    assert fd.get("available_on_create") is True
    assert fd.get("form_editable") is False


def test_every_catalog_entity_is_covered_by_column_check():
    """新增实体目录时必须同步补 _ENTITY_MODELS，否则字段名校验会静默跳过它。"""
    from app.domains.lowcode.native_field_catalog import CATALOG
    missing = set(CATALOG) - set(_ENTITY_MODELS)
    assert not missing, f"这些实体有目录但没纳入列名校验: {missing}"


# ===== 端到端：经 API 配置后生效 =====

async def _publish_lead_native_override(client: AsyncClient, h: dict, overrides: list[dict],
                                        rules: list[dict] | None = None):
    tpl = (await client.get("/api/v1/lc/entity-templates/lead", headers=h)).json()["data"]
    body = {"field_definitions": overrides, "layout_definition": {}, "rule_definitions": rules or []}
    r = await client.post(f"/api/v1/lc/form-templates/{tpl['id']}/design", headers=h, json=body)
    assert r.json()["code"] == 0, r.text
    r = await client.post(f"/api/v1/lc/form-templates/{tpl['id']}/publish", headers=h)
    assert r.json()["code"] == 0, r.text
    return tpl["id"]


@pytest.fixture
async def reset_lead_template(client: AsyncClient, auth_headers: dict):
    """前后各清空发布一次线索实体模板。

    这些用例改的是种子租户里真实的实体模板，测试库又是共用的。只在结束时清理不够：
    上一轮若被中断（Ctrl-C / 进程被杀），残留的必填或脱敏覆盖会让后续 test_lead.py
    莫名其妙地失败，且跨进程持续存在。故进入时也先清一次。
    """
    await _publish_lead_native_override(client, auth_headers, [])
    yield
    await _publish_lead_native_override(client, auth_headers, [])


async def test_tenant_configured_required_is_enforced(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    h = auth_headers
    await _publish_lead_native_override(client, h, [
        {"id": "industry", "native": True, "required": True, "label": "行业", "type": "select"},
    ])

    # 缺行业 → 被后端拦下
    resp = (await client.post("/api/v1/leads", headers=h, json={
        "title": "缺行业线索", "company_name": "某公司",
    })).json()
    assert resp["code"] != 0
    assert "行业" in resp["message"]

    # 补上行业 → 通过
    ok = (await client.post("/api/v1/leads", headers=h, json={
        "title": "有行业线索", "company_name": "某公司", "industry": "mining",
    })).json()
    assert ok["code"] == 0
    await client.delete(f"/api/v1/leads/{ok['data']['id']}", headers=h)


async def test_company_name_default_required_but_tenant_can_relax(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    """company_name 列可空，改造前表单硬编码必填；改造后默认仍必填，但租户可关掉。"""
    h = auth_headers

    # 出厂默认：不填公司名被拦（保持改造前行为）
    resp = (await client.post("/api/v1/leads", headers=h, json={"title": "无公司名线索"})).json()
    assert resp["code"] != 0 and "公司名称" in resp["message"]

    # 租户把它改成非必填 → 放行
    await _publish_lead_native_override(client, h, [
        {"id": "company_name", "native": True, "required": False, "label": "公司名称", "type": "text"},
    ])
    ok = (await client.post("/api/v1/leads", headers=h, json={"title": "无公司名线索"})).json()
    assert ok["code"] == 0, f"租户已关掉必填，不应再被拦: {ok}"
    await client.delete(f"/api/v1/leads/{ok['data']['id']}", headers=h)


async def test_conditional_visibility_suppresses_required(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    """国家名设为必填、但仅在国别=国外时显示：国内提交不应被必填拦住（防死锁）。"""
    h = auth_headers
    await _publish_lead_native_override(client, h, [
        {"id": "country_name", "native": True, "required": True, "label": "国家", "type": "text"},
    ], rules=[{
        "id": "r1", "type": "visibility", "target_field_id": "country_name",
        "condition": {"field": "country_type", "operator": "eq", "value": "overseas"},
        "action": {"visible": True},
    }])

    domestic = (await client.post("/api/v1/leads", headers=h, json={
        "title": "国内线索", "company_name": "某公司", "country_type": "domestic",
    })).json()
    assert domestic["code"] == 0, f"字段被规则隐藏时不得报必填: {domestic}"
    await client.delete(f"/api/v1/leads/{domestic['data']['id']}", headers=h)

    overseas = (await client.post("/api/v1/leads", headers=h, json={
        "title": "国外线索", "company_name": "某公司", "country_type": "overseas",
    })).json()
    assert overseas["code"] != 0 and "国家" in overseas["message"]


async def test_partial_update_skips_untouched_required_fields(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    """批量改派这类局部更新，不应因历史数据缺少「后来才设为必填」的字段而失败。

    收录后整单内容已锁；草稿 PUT 又会 skip_required。故用「无 running 的 pending」
    （等同撤回后再改）测 payload 范围必填：局部更新放行、显式留空仍拦。
    """
    import app.database as db_module
    from sqlalchemy import select
    from app.domains.lead.models import Lead

    h = auth_headers
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "历史线索", "company_name": "老公司", "as_draft": True,
    })).json()["data"]["id"]

    async with db_module.async_session_factory() as db:
        lead = (await db.execute(select(Lead).where(Lead.id == lid))).scalar_one()
        lead.review_status = "pending"
        await db.commit()

    await _publish_lead_native_override(client, h, [
        {"id": "industry", "native": True, "required": True, "label": "行业", "type": "select"},
    ])

    # 只改状态：不带 industry → 放行
    r = (await client.put(f"/api/v1/leads/{lid}", headers=h, json={"status": "following"})).json()
    assert r["code"] == 0, f"局部更新不应被未提交的必填字段拦住: {r}"

    # 表单整体提交且把 industry 显式留空 → 拦下
    r = (await client.put(f"/api/v1/leads/{lid}", headers=h, json={
        "title": "历史线索", "company_name": "老公司", "industry": None,
    })).json()
    assert r["code"] != 0 and "行业" in r["message"]

    await client.delete(f"/api/v1/leads/{lid}", headers=h)


async def test_native_field_masking_on_read_paths(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    """角色键控脱敏：无「可见明文角色」的用户在列表与详情上都只拿到 ***。

    这是被删掉的 field_rules Tab 曾经承诺、却从未真正实现的能力（那套后端零执行点）。
    """
    h = auth_headers
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "脱敏线索", "company_name": "某公司", "industry": "screening_mining",
    })).json()["data"]["id"]

    # 明文可见角色设为一个当前用户不具备的角色 → 该用户应只看到 ***
    await _publish_lead_native_override(client, h, [
        {"id": "industry", "native": True, "label": "行业", "type": "select",
         "unmask_roles": ["__finance_only__"]},
    ])

    detail = (await client.get(f"/api/v1/leads/{lid}", headers=h)).json()["data"]
    assert detail["industry"] == "***", f"详情未脱敏: {detail['industry']!r}"

    listed = (await client.get("/api/v1/leads", headers=h,
                               params={"keyword": "脱敏线索"})).json()["data"]["items"]
    row = next(i for i in listed if i["id"] == lid)
    assert row["industry"] == "***", f"列表未脱敏: {row['industry']!r}"

    # 撤掉限制 → 恢复明文，确认脱敏未把真实值写坏
    await _publish_lead_native_override(client, h, [])
    after = (await client.get(f"/api/v1/leads/{lid}", headers=h)).json()["data"]
    assert after["industry"] == "screening_mining", "脱敏只应影响出参，不得改动库里的真实值"

    await client.delete(f"/api/v1/leads/{lid}", headers=h)


async def test_masked_and_required_is_not_a_deadlock(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    """脱敏 + 必填不得让记录永远存不下去 —— 看不到明文的人无法填写该字段。"""
    h = auth_headers
    await _publish_lead_native_override(client, h, [
        {"id": "industry", "native": True, "label": "行业", "type": "select",
         "required": True, "unmask_roles": ["__finance_only__"]},
    ])
    r = (await client.post("/api/v1/leads", headers=h, json={
        "title": "脱敏必填线索", "company_name": "某公司",
    })).json()
    assert r["code"] == 0, f"脱敏字段不应报必填(用户无从填写): {r}"
    await client.delete(f"/api/v1/leads/{r['data']['id']}", headers=h)


async def test_country_name_required_does_not_block_domestic_lead(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    """「国家」只在国别=国外时才在表单上出现，内置规则须让国内线索免于该必填。

    否则租户一旦给「国家」勾必填，国内线索会被后端拦下、界面上却找不到这个字段可填。
    """
    h = auth_headers
    await _publish_lead_native_override(client, h, [
        {"id": "country_name", "native": True, "label": "国家", "type": "text", "required": True},
    ])

    domestic = (await client.post("/api/v1/leads", headers=h, json={
        "title": "国内线索", "company_name": "某公司", "country_type": "domestic",
    })).json()
    assert domestic["code"] == 0, f"国内线索不应被「国家」必填拦住: {domestic}"
    await client.delete(f"/api/v1/leads/{domestic['data']['id']}", headers=h)

    # 国外线索该字段确实出现在表单上，必填照常生效
    overseas = (await client.post("/api/v1/leads", headers=h, json={
        "title": "国外线索", "company_name": "某公司", "country_type": "overseas",
    })).json()
    assert overseas["code"] != 0 and "国家" in overseas["message"]


async def test_masking_covers_derived_display_fields(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    """脱敏 reporter_id 必须连带 reporter_name —— 列表页渲染的正是后者。"""
    h = auth_headers
    me = (await client.get("/api/v1/auth/me", headers=h)).json()["data"]
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "派生字段脱敏线索", "company_name": "某公司",
        "reporter_id": me["id"],
    })).json()["data"]["id"]

    before = (await client.get(f"/api/v1/leads/{lid}", headers=h)).json()["data"]
    assert before["reporter_name"], "前置条件：该线索应有申报人姓名"

    await _publish_lead_native_override(client, h, [
        {"id": "reporter_id", "native": True, "label": "申报人", "type": "person",
         "unmask_roles": ["__manager_only__"]},
    ])
    after = (await client.get(f"/api/v1/leads/{lid}", headers=h)).json()["data"]
    assert after["reporter_id"] == "***"
    assert after["reporter_name"] == "***", "只裁 reporter_id 而漏了 reporter_name，脱敏等于没配"

    await client.delete(f"/api/v1/leads/{lid}", headers=h)


async def test_export_respects_field_policy(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    """导出必须与页面同口径脱敏，否则「看不到但导得出」= 绕过字段权限的后门。"""
    from openpyxl import load_workbook
    import io as _io

    h = auth_headers
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "导出脱敏线索", "company_name": "某公司", "industry": "screening_mining",
    })).json()["data"]["id"]

    def industry_cells(content: bytes):
        wb = load_workbook(_io.BytesIO(content))
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = header.index("行业")
        return [row[idx].value for row in ws.iter_rows(min_row=2)]

    # 未配置时导出明文中文标签（库内存英文码）
    plain = (await client.get("/api/v1/leads/export/excel", headers=h,
                              params={"keyword": "导出脱敏线索"})).content
    assert "筛分分选-矿山" in industry_cells(plain)
    assert "screening_mining" not in industry_cells(plain)

    # 配置为仅特定角色可见明文 → 导出应变成 ***
    await _publish_lead_native_override(client, h, [
        {"id": "industry", "native": True, "label": "行业", "type": "select",
         "unmask_roles": ["__finance_only__"]},
    ])
    masked = (await client.get("/api/v1/leads/export/excel", headers=h,
                               params={"keyword": "导出脱敏线索"})).content
    cells = industry_cells(masked)
    assert "screening_mining" not in cells, "导出泄露了页面上已脱敏的字段"
    assert "筛分分选-矿山" not in cells, "脱敏后不应再出现明文行业"
    assert "***" in cells

    await client.delete(f"/api/v1/leads/{lid}", headers=h)


async def test_masked_native_field_write_is_discarded(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    """被脱敏的字段一律不可编辑 —— 否则用户会把 "***" 当成真值提交回去。"""
    h = auth_headers
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "脱敏只读线索", "company_name": "某公司", "industry": "screening_mining",
        "as_draft": True,
    })).json()["data"]["id"]

    await _publish_lead_native_override(client, h, [
        {"id": "industry", "native": True, "label": "行业", "type": "select",
         "unmask_roles": ["__finance_only__"]},
    ])
    await client.put(f"/api/v1/leads/{lid}", headers=h, json={"industry": "***"})

    await _publish_lead_native_override(client, h, [])
    after = (await client.get(f"/api/v1/leads/{lid}", headers=h)).json()["data"]
    assert after["industry"] == "screening_mining", "脱敏字段的写入必须被丢弃，不能用 *** 覆盖真值"

    await client.delete(f"/api/v1/leads/{lid}", headers=h)


async def test_readonly_native_field_ignores_user_write(
    client: AsyncClient, auth_headers: dict, reset_lead_template,
):
    h = auth_headers
    lid = (await client.post("/api/v1/leads", headers=h, json={
        "title": "只读字段线索", "company_name": "原始公司", "remark": "原始备注",
        "as_draft": True,
    })).json()["data"]["id"]

    await _publish_lead_native_override(client, h, [
        {"id": "remark", "native": True, "label": "备注", "type": "textarea",
         "props": {"readonly": True}},
    ])

    await client.put(f"/api/v1/leads/{lid}", headers=h, json={"remark": "偷改的备注"})
    after = (await client.get(f"/api/v1/leads/{lid}", headers=h)).json()["data"]
    assert after["remark"] == "原始备注", "只读原生字段的写入必须被后端丢弃"

    await client.delete(f"/api/v1/leads/{lid}", headers=h)


def test_validate_detail_table_rows_contract_lines():
    from app.common.exceptions import BusinessException
    from app.domains.lowcode.field_permission import _validate_detail_table_rows
    from app.domains.lowcode.native_field_catalog import get_native_fields

    cols = next(f for f in get_native_fields("contract") if f["id"] == "line_items")["detail_table_columns"]

    with pytest.raises(BusinessException) as exc:
        _validate_detail_table_rows("合同明细", [], cols)
    assert "合同明细" in exc.value.message

    with pytest.raises(BusinessException) as exc2:
        _validate_detail_table_rows("合同明细", [{
            "is_fx": "否", "product_type": "复频筛", "name": "筛", "spec": "X", "unit": "台", "qty": 1,
        }], cols)
    assert "单价" in exc2.value.message

    _validate_detail_table_rows("合同明细", [{
        "is_fx": "否", "product_type": "复频筛", "name": "筛", "spec": "X", "unit": "台", "qty": 1, "price": 100,
    }], cols)
